from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "stepwise_pathway_batch25_20260712"
MASTER_XLSX = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch25.xlsx"
MASTER_CSV = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch25.csv"
UNRESOLVED_PRODUCT = "Microbial transformation products not resolved in current database source"


def row(pesticide, microorganism, product, reaction_label, enzyme, gene, evidence_type, doi, title, note):
    return {
        "pesticide": pesticide,
        "pathway_name": f"{microorganism} {pesticide} degradation evidence",
        "microorganism": microorganism,
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": pesticide,
        "product": product,
        "reaction_label": reaction_label,
        "enzyme": enzyme,
        "gene": gene,
        "evidence_type": evidence_type,
        "doi": doi,
        "reference_title": title,
        "source_pdf": "Database reference from pesticide_data.xlsx / PBDB_master_with_ids.xlsx",
        "evidence_note": note,
    }


ROWS = [
    row(
        "Captan",
        "Bacillus circulans",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.1021/jf1030339",
        "Database DOI record for captan degradation by Bacillus circulans",
        "This DOI is present in pesticide_data.xlsx. The database supports microbial captan degradation; exact transformation products are not standardized in the current source table.",
    ),
    row(
        "Chlordane",
        "Phanerochaete chrysosporium BKM-F-1767",
        "Fungal oxidative transformation products not resolved in current database source",
        "White-rot fungal transformation",
        "",
        "",
        "WHOLE_CELL",
        "10.1128/aem.56.8.2347-2353.1990",
        "Database DOI record for chlordane degradation by Phanerochaete chrysosporium",
        "This DOI is present in pesticide_data.xlsx. The database supports fungal chlordane degradation; exact metabolites are not standardized in the current source table.",
    ),
    row(
        "Cyprodinil",
        "Acinetobacter johnsonii LXL_C1",
        "Cytochrome P450-associated transformation products",
        "Genomic cytochrome P450-associated transformation",
        "Cytochrome P450 monooxygenase",
        "frmA; ADH5; adhC; catA",
        "GENETIC",
        "10.1016/j.micpath.2018.11.016",
        "Database DOI record for cyprodinil-associated genomic evidence in Acinetobacter johnsonii",
        "This DOI is present in pesticide_data.xlsx. The database supports genomic/P450-associated evidence; exact products are not standardized in the current source table.",
    ),
    row(
        "Dimethomorph",
        "Bacillus cereus WL08",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.1016/j.jhazmat.2020.122806",
        "Database DOI record for dimethomorph degradation by Bacillus cereus WL08",
        "This DOI is present in pesticide_data.xlsx. The database supports microbial degradation; exact transformation products are not standardized in the current source table.",
    ),
    row(
        "Diquat",
        "Meyerozyma guilliermondii strain Wyslmt",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation with reported gene lead",
        "",
        "DN676",
        "GENETIC",
        "10.3389/fmicb.2022.993721",
        "Database DOI record for diquat degradation by Meyerozyma guilliermondii",
        "This DOI is present in pesticide_data.xlsx. The database supports microbial degradation with a gene lead; exact transformation products are not standardized in the current source table.",
    ),
    row(
        "Emamectin benzoate",
        "Aeromonas taiwanensis ZJB-18044",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.1007/s10532-020-09909-8",
        "Database DOI record for emamectin benzoate degradation by Aeromonas taiwanensis",
        "This DOI is present in pesticide_data.xlsx. The database supports microbial degradation; exact products are not standardized in the current source table.",
    ),
    row(
        "Fenhexamid",
        "Serratia sarumanii strain GBS19",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.1007/s10532-025-10144-2",
        "Database DOI record for fenhexamid degradation by Serratia sarumanii",
        "This DOI is present in pesticide_data.xlsx. The database supports microbial degradation; exact products are not standardized in the current source table.",
    ),
    row(
        "Fluazinam",
        "Paenibacillus polymyxa",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.1016/j.ecoenv.2018.10.093",
        "Database DOI record for fluazinam degradation by Paenibacillus polymyxa",
        "This DOI is present in pesticide_data.xlsx. The database supports microbial degradation; exact products are not standardized in the current source table.",
    ),
    row(
        "Haloxyfop",
        "Myrothecium verrucaria",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.1016/j.scitotenv.2024.178012",
        "Database DOI record for haloxyfop degradation by Myrothecium verrucaria",
        "This DOI is present in pesticide_data.xlsx. The database supports microbial degradation; exact products are not standardized in the current source table.",
    ),
    row(
        "Heptachlor",
        "Shigella sp.",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.2166/wst.2018.127",
        "Database DOI record for heptachlor degradation by Shigella sp.",
        "This DOI is present in pesticide_data.xlsx. The database supports microbial degradation; exact transformation products are not standardized in the current source table.",
    ),
]


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 72)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(ROWS)
    decisions = pd.DataFrame(
        [
            {
                "pesticide": pesticide,
                "decision": "Integrated with conservative product label",
                "reason": "Reference DOI is present in pesticide_data.xlsx. Product names were kept conservative where exact metabolites are not standardized in the current source data.",
            }
            for pesticide in sorted({r["pesticide"] for r in ROWS})
        ]
    )
    df.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch25_final.csv", index=False)
    decisions.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch25_screening_decisions.csv", index=False)
    df.to_csv(MASTER_CSV, index=False)
    with pd.ExcelWriter(OUT_DIR / "pesticidedb_stepwise_pathway_batch25_final.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 25", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 25", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    print(f"rows={len(df)}")
    print(f"pesticides={df['pesticide'].nunique()}")


if __name__ == "__main__":
    main()
