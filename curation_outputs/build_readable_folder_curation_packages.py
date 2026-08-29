from __future__ import annotations

from pathlib import Path
import re
import subprocess
import textwrap
from datetime import date

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
QUEUE_FILE = PROJECT_ROOT / "curation_outputs" / "evidence_pesticide_pdf_inventory_20260707" / "evidence_pesticide_readable_pdf_queue.csv"
EVIDENCE_FILE = PROJECT_ROOT / "pesticide_data.xlsx"
PDFTOTEXT = "/Users/nana/.cache/codex-runtimes/codex-primary-runtime/dependencies/native/poppler/poppler/bin/pdftotext"
TODAY = "20260707"

BIODEGRADATION_TERMS = [
    "biodegradation",
    "biodegrade",
    "degradation",
    "degrade",
    "mineralization",
    "metabolite",
    "metabolites",
    "transformation",
    "biotransformation",
    "catabolism",
    "hydrolysis",
    "dehalogenation",
    "oxidation",
]

MOLECULAR_TERMS = [
    "gene",
    "enzyme",
    "protein",
    "monooxygenase",
    "dioxygenase",
    "hydrolase",
    "dehydrogenase",
    "dehalogenase",
    "esterase",
    "amidase",
    "nitrile hydratase",
]

PATHWAY_TERMS = [
    "pathway",
    "intermediate",
    "product",
    "products",
    "metabolite",
    "lc-ms",
    "gc-ms",
    "hplc",
]

DOI_RE = re.compile(r"\b10\.\d{4,9}/[-._;()/:A-Z0-9]+\b", re.I)
YEAR_RE = re.compile(r"\b(19[7-9]\d|20[0-2]\d)\b")
STRAIN_RE = re.compile(
    r"\b(?:strain|isolate)\s+([A-Z][A-Za-z0-9_.-]*(?:\s+[A-Za-z0-9_.-]+){0,3})",
    re.I,
)


def clean_text(text: str) -> str:
    text = text.replace("\x00", " ")
    text = re.sub(r"[\x01-\x08\x0B-\x0C\x0E-\x1F]", " ", text)
    text = re.sub(r"-\n", "", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def excel_safe(value: object) -> object:
    if not isinstance(value, str):
        return value
    return re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]", " ", value)


def frame_safe(rows: list[dict[str, object]] | pd.DataFrame) -> pd.DataFrame:
    df = rows if isinstance(rows, pd.DataFrame) else pd.DataFrame(rows)
    return df.map(excel_safe) if not df.empty else df


def safe_name(name: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "_", name).strip("_")


def classify_file(path: Path) -> str:
    try:
        head = path.read_bytes()[:64]
    except OSError:
        return "read_error"
    if head.startswith(b"%PDF"):
        return "real_pdf"
    if head.lstrip().startswith(b"<"):
        return "html_placeholder"
    return "other"


def extract_pdf_text(pdf_path: Path) -> tuple[str, str]:
    try:
        result = subprocess.run(
            [PDFTOTEXT, "-layout", str(pdf_path), "-"],
            check=False,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=45,
        )
    except subprocess.TimeoutExpired:
        return "", "timeout"
    if result.returncode != 0:
        return "", f"pdftotext_error: {result.stderr.strip()[:180]}"
    return result.stdout, ""


def probable_title(raw_text: str, filename: str) -> str:
    lines = [re.sub(r"\s+", " ", line).strip() for line in raw_text.splitlines()[:80]]
    candidates = []
    for line in lines:
        if len(line) < 18 or len(line) > 220:
            continue
        lower = line.lower()
        if any(skip in lower for skip in ["abstract", "introduction", "journal", "copyright", "downloaded from"]):
            continue
        if sum(ch.isalpha() for ch in line) < 12:
            continue
        candidates.append(line)
    if candidates:
        return candidates[0]
    return filename.rsplit(".", 1)[0]


def best_snippet(text: str, pesticide: str, terms: list[str], window: int = 260) -> str:
    lower = text.lower()
    search_terms = [pesticide.lower(), *terms]
    hits = []
    for term in search_terms:
        idx = lower.find(term)
        if idx >= 0:
            start = max(0, idx - window)
            end = min(len(text), idx + len(term) + window)
            hits.append(text[start:end])
    if not hits:
        return ""
    return textwrap.shorten(clean_text(" ... ".join(hits[:3])), width=1200, placeholder=" ...")


def evidence_level(text: str) -> str:
    lower = text.lower()
    has_molecular = any(term in lower for term in MOLECULAR_TERMS)
    has_pathway = any(term in lower for term in PATHWAY_TERMS)
    has_bio = any(term in lower for term in BIODEGRADATION_TERMS)
    if has_molecular and has_pathway:
        return "candidate_gene/enzyme_plus_pathway"
    if has_molecular:
        return "candidate_gene/enzyme"
    if has_pathway:
        return "candidate_pathway/metabolite"
    if has_bio:
        return "candidate_whole-cell_or_transformation"
    return "low_signal"


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells[:80])
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 58)


def main() -> None:
    queue = pd.read_csv(QUEUE_FILE)
    evidence = pd.read_excel(EVIDENCE_FILE)
    evidence["Pesticide_clean"] = evidence["Pesticide"].astype(str).str.strip().str.casefold()
    manifest_rows = []

    for _, qrow in queue.iterrows():
        pesticide = str(qrow["pesticide"]).strip()
        folder = Path(str(qrow["usb_folder"]))
        out_dir = folder / f"PesticideDB_curation_{safe_name(pesticide)}_{TODAY}"
        out_dir.mkdir(parents=True, exist_ok=True)

        paper_rows = []
        evidence_rows = []
        snippet_rows = []
        real_pdfs = [p for p in sorted(folder.glob("*.pdf"), key=lambda p: p.name.casefold()) if classify_file(p) == "real_pdf"]

        for idx, pdf in enumerate(real_pdfs, start=1):
            raw, error = extract_pdf_text(pdf)
            text = clean_text(raw)
            title = probable_title(raw, pdf.name) if raw else pdf.stem
            dois = sorted(set(match.group(0).rstrip(".") for match in DOI_RE.finditer(text)))
            years = sorted(set(match.group(0) for match in YEAR_RE.finditer(text)))
            strain_hits = sorted(set(match.group(1).strip(" .,:;") for match in STRAIN_RE.finditer(text)))[:6]
            lower = text.lower()
            biodeg_hits = sum(lower.count(term) for term in BIODEGRADATION_TERMS)
            molecular_hits = sum(lower.count(term) for term in MOLECULAR_TERMS)
            pathway_hits = sum(lower.count(term) for term in PATHWAY_TERMS)
            level = evidence_level(text) if text else "unreadable"
            snippet = best_snippet(text, pesticide, BIODEGRADATION_TERMS + MOLECULAR_TERMS + PATHWAY_TERMS)

            paper_rows.append(
                {
                    "pesticide": pesticide,
                    "pdf_file": pdf.name,
                    "probable_title": title,
                    "doi_candidates": "; ".join(dois[:4]),
                    "year_candidates": "; ".join(years[-4:]),
                    "candidate_microorganisms_or_strains": "; ".join(strain_hits),
                    "biodegradation_term_hits": biodeg_hits,
                    "molecular_term_hits": molecular_hits,
                    "pathway_term_hits": pathway_hits,
                    "automated_evidence_signal": level,
                    "review_decision": "",
                    "review_notes": error,
                }
            )
            if level != "low_signal" and text:
                evidence_rows.append(
                    {
                        "pesticide": pesticide,
                        "evidence_source": "PDF automated candidate; manual validation required",
                        "pdf_file": pdf.name,
                        "probable_title": title,
                        "doi": "; ".join(dois[:2]),
                        "year": years[-1] if years else "",
                        "microorganism": "; ".join(strain_hits),
                        "evidence_type": level,
                        "gene": "",
                        "enzyme_or_protein": "",
                        "substrate": pesticide,
                        "products_or_metabolites": "",
                        "experimental_method": "",
                        "database_use_recommendation": "Review against paper before integration",
                        "curator_notes": "",
                    }
                )
            if snippet:
                snippet_rows.append(
                    {
                        "pesticide": pesticide,
                        "pdf_file": pdf.name,
                        "probable_title": title,
                        "snippet": snippet,
                    }
                )

        existing_rows = evidence[evidence["Pesticide_clean"] == pesticide.casefold()].drop(columns=["Pesticide_clean"])
        pathway_template = pd.DataFrame(
            columns=[
                "pesticide",
                "pathway_name",
                "step_order",
                "substrate",
                "product",
                "reaction_label",
                "enzyme_or_gene",
                "microorganism",
                "evidence_strength",
                "arrow_style_for_viewer",
                "reference_doi",
                "source_pdf",
                "curator_validation_status",
                "notes",
            ]
        )
        readme = pd.DataFrame(
            [
                ["purpose", "Automated first-pass curation package for evidence-positive PesticideDB pathway review."],
                ["created", date.today().isoformat()],
                ["important", "Rows marked candidate are not final database records until manually validated from the paper."],
                ["recommended_workflow", "Review Paper Screening, confirm true biodegradation papers, then fill Evidence Candidates and Pathway Step Template."],
                ["source_folder", str(folder)],
                ["readable_pdf_count", len(real_pdfs)],
            ],
            columns=["field", "value"],
        )

        out_xlsx = out_dir / f"{safe_name(pesticide).lower()}_pathway_curation_review.xlsx"
        with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
            frame_safe(readme).to_excel(writer, sheet_name="README", index=False)
            frame_safe(existing_rows).to_excel(writer, sheet_name="Existing DB Evidence", index=False)
            frame_safe(paper_rows).to_excel(writer, sheet_name="Paper Screening", index=False)
            frame_safe(evidence_rows).to_excel(writer, sheet_name="Evidence Candidates", index=False)
            frame_safe(pathway_template).to_excel(writer, sheet_name="Pathway Step Template", index=False)
            frame_safe(snippet_rows).to_excel(writer, sheet_name="Keyword Snippets", index=False)
            style_workbook(writer)

        frame_safe(paper_rows).to_csv(out_dir / f"{safe_name(pesticide).lower()}_paper_screening.csv", index=False)
        frame_safe(evidence_rows).to_csv(out_dir / f"{safe_name(pesticide).lower()}_evidence_candidates.csv", index=False)
        pathway_template.to_csv(out_dir / f"{safe_name(pesticide).lower()}_pathway_steps_template.csv", index=False)

        manifest_rows.append(
            {
                "pesticide": pesticide,
                "output_folder": str(out_dir),
                "workbook": str(out_xlsx),
                "readable_pdfs_processed": len(real_pdfs),
                "candidate_evidence_rows": len(evidence_rows),
                "existing_database_rows": len(existing_rows),
            }
        )
        print(f"{pesticide}: pdfs={len(real_pdfs)} candidates={len(evidence_rows)}")

    manifest = pd.DataFrame(manifest_rows)
    manifest_dir = PROJECT_ROOT / "curation_outputs" / "evidence_pesticide_readable_curation_manifest_20260707"
    manifest_dir.mkdir(parents=True, exist_ok=True)
    manifest.to_csv(manifest_dir / "readable_folder_curation_manifest.csv", index=False)
    with pd.ExcelWriter(manifest_dir / "readable_folder_curation_manifest.xlsx", engine="openpyxl") as writer:
        manifest.to_excel(writer, sheet_name="Manifest", index=False)
        style_workbook(writer)
    print(f"manifest={manifest_dir / 'readable_folder_curation_manifest.xlsx'}")


if __name__ == "__main__":
    main()
