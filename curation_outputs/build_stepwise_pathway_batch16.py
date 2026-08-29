from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "stepwise_pathway_batch16_20260710"
MASTER_XLSX = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch16.xlsx"
MASTER_CSV = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch16.csv"


ROWS = [
    {
        "pesticide": "Fenvalerate",
        "pathway_name": "Citrobacter freundii CD-9 fenvalerate degradation",
        "microorganism": "Citrobacter freundii CD-9",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Fenvalerate",
        "product": "3-phenoxybenzaldehyde + butyl 1,2-phthalate",
        "reaction_label": "Ester-bond hydrolysis",
        "enzyme": "Intracellular crude enzyme activity",
        "gene": "",
        "evidence_type": "CRUDE_EXTRACT",
        "doi": "10.1186/s13568-020-01128-x",
        "reference_title": "Biodegradation and metabolic pathway of fenvalerate by Citrobacter freundii CD-9",
        "source_pdf": "Fenvalerate_10.1186_s13568-020-01128-x.pdf",
        "evidence_note": "The study reports that fenvalerate was first hydrolyzed by ester-bond cleavage to main intermediate products including 3-phenoxybenzaldehyde and butyl 1,2-phthalate; the degrading enzyme activity was mainly intracellular.",
    },
    {
        "pesticide": "Fenvalerate",
        "pathway_name": "Citrobacter freundii CD-9 fenvalerate degradation",
        "microorganism": "Citrobacter freundii CD-9",
        "completeness": "PARTIAL",
        "step_order": 2,
        "substrate": "Fenvalerate",
        "product": "3-chlorophenylacetic acid",
        "reaction_label": "Proposed ester-linkage hydrolysis branch",
        "enzyme": "Intracellular crude enzyme activity",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1186/s13568-020-01128-x",
        "reference_title": "Biodegradation and metabolic pathway of fenvalerate by Citrobacter freundii CD-9",
        "source_pdf": "Fenvalerate_10.1186_s13568-020-01128-x.pdf",
        "evidence_note": "The proposed pathway describes fenvalerate degradation to 3-chlorophenylacetic acid through ester-linkage hydrolysis; alpha-cyano-3-phenoxybenzyl alcohol is reported as a co-product of this branch rather than a separate downstream node.",
    },
    {
        "pesticide": "Fenvalerate",
        "pathway_name": "Citrobacter freundii CD-9 fenvalerate degradation",
        "microorganism": "Citrobacter freundii CD-9",
        "completeness": "PARTIAL",
        "step_order": 3,
        "substrate": "3-chlorophenylacetic acid",
        "product": "4-hydroxyphenylacetic acid",
        "reaction_label": "Proposed dehalogenation / hydroxylation branch",
        "enzyme": "Activity not assigned",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1186/s13568-020-01128-x",
        "reference_title": "Biodegradation and metabolic pathway of fenvalerate by Citrobacter freundii CD-9",
        "source_pdf": "Fenvalerate_10.1186_s13568-020-01128-x.pdf",
        "evidence_note": "The proposed route indicates conversion of 3-chlorophenylacetic acid to 4-hydroxyphenylacetic acid; the paper identifies multiple metabolites by HPLC/GC-MS.",
    },
    {
        "pesticide": "Deltamethrin",
        "pathway_name": "Acinetobacter junii and Klebsiella pneumoniae co-culture deltamethrin degradation",
        "microorganism": "Acinetobacter junii LH-1-1; Klebsiella pneumoniae BPBA052",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Deltamethrin",
        "product": "3-phenoxybenzaldehyde",
        "reaction_label": "Hydrolysis",
        "enzyme": "Co-culture hydrolytic activity not assigned",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1186/s13568-020-01043-1",
        "reference_title": "Bioremediation of deltamethrin and its metabolite 3-phenoxybenzoic acid by co-culture",
        "source_pdf": "Deltamethrin_10.1186_s13568-020-01043-1.pdf",
        "evidence_note": "GC-MS detected 3-phenoxybenzaldehyde during deltamethrin degradation; the proposed co-culture pathway starts with deltamethrin hydrolysis.",
    },
    {
        "pesticide": "Deltamethrin",
        "pathway_name": "Acinetobacter junii and Klebsiella pneumoniae co-culture deltamethrin degradation",
        "microorganism": "Acinetobacter junii LH-1-1; Klebsiella pneumoniae BPBA052",
        "completeness": "PARTIAL",
        "step_order": 2,
        "substrate": "Deltamethrin",
        "product": "1,2-benzenedicarboxylic butyl dacyl ester",
        "reaction_label": "Proposed downstream transformation",
        "enzyme": "Activity not assigned",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1186/s13568-020-01043-1",
        "reference_title": "Bioremediation of deltamethrin and its metabolite 3-phenoxybenzoic acid by co-culture",
        "source_pdf": "Deltamethrin_10.1186_s13568-020-01043-1.pdf",
        "evidence_note": "1,2-benzenedicarboxylic butyl dacyl ester was identified by GC-MS as a deltamethrin degradation metabolite in the co-culture system.",
    },
    {
        "pesticide": "Deltamethrin",
        "pathway_name": "Acinetobacter junii and Klebsiella pneumoniae co-culture deltamethrin degradation",
        "microorganism": "Acinetobacter junii LH-1-1; Klebsiella pneumoniae BPBA052",
        "completeness": "PARTIAL",
        "step_order": 3,
        "substrate": "Deltamethrin",
        "product": "Phenol",
        "reaction_label": "Proposed downstream transformation",
        "enzyme": "Activity not assigned",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1186/s13568-020-01043-1",
        "reference_title": "Bioremediation of deltamethrin and its metabolite 3-phenoxybenzoic acid by co-culture",
        "source_pdf": "Deltamethrin_10.1186_s13568-020-01043-1.pdf",
        "evidence_note": "Phenol was identified by GC-MS as a metabolite during deltamethrin degradation in the co-culture system.",
    },
]


SCREENING_DECISIONS = [
    {
        "pesticide": "Paraquat",
        "decision": "Not integrated",
        "reason": "The accessible paper reports anaerobic transformation/removal by Pseudomonas geniculata PQ01 but does not provide a clean named product/intermediate arrow.",
    },
    {
        "pesticide": "Pinoxaden",
        "decision": "Not integrated",
        "reason": "The accessible paper reports degradation and predicted genes, but no named microbial transformation products suitable for a pathway arrow.",
    },
    {
        "pesticide": "Thiacloprid",
        "decision": "Not integrated",
        "reason": "The accessible paper cites prior thiacloprid transformation literature but states that further studies are needed to explore metabolites in the tested system.",
    },
    {
        "pesticide": "Imazethapyr / Prochloraz / Fluxapyroxad / Pyrimethanil",
        "decision": "Not integrated",
        "reason": "Accessible PDFs support removal/degradation potential but do not provide named substrate-to-product microbial pathway steps for import.",
    },
]


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 72)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(ROWS)
    decisions = pd.DataFrame(SCREENING_DECISIONS)
    df.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch16_final.csv", index=False)
    decisions.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch16_screening_decisions.csv", index=False)
    df.to_csv(MASTER_CSV, index=False)
    with pd.ExcelWriter(OUT_DIR / "pesticidedb_stepwise_pathway_batch16_final.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 16", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 16", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    print(f"rows={len(df)}")
    print(f"pesticides={df['pesticide'].nunique()}")


if __name__ == "__main__":
    main()
