from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "stepwise_pathway_batch19_20260711"
MASTER_XLSX = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch19.xlsx"
MASTER_CSV = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch19.csv"


ROWS = [
    {
        "pesticide": "Methamidophos",
        "pathway_name": "Pseudomonas aeruginosa Is-6 methamidophos degradation",
        "microorganism": "Pseudomonas aeruginosa strain Is-6",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Methamidophos",
        "product": "O,S-dimethyl phosphorothioate (DMPT)",
        "reaction_label": "Hydrolysis of amino group",
        "enzyme": "Methamidophos phosphoamide hydrolase proposed",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1080/03601234.2013.836868",
        "reference_title": "Biodegradation of acephate and methamidophos by a soil bacterium Pseudomonas aeruginosa strain Is-6",
        "source_pdf": "Biodegradation of acephate and methamidophos by a soil bacterium Pseudomonas aeruginosa strain Is-6..pdf",
        "evidence_note": "The primary paper reports complete methamidophos degradation by Pseudomonas aeruginosa Is-6 and detects O,S-dimethyl phosphorothioate (DMPT) by HPLC/ESI-MS. The authors describe methamidophos degradation as hydrolysis of the amino group by a proposed methamidophos phosphoamide hydrolase, releasing ammonium ions and yielding DMPT.",
    },
]


SCREENING_DECISIONS = [
    {
        "pesticide": "Methamidophos",
        "decision": "Integrated",
        "reason": "A real PDF was available in the Acephate USB folder and contains a primary experimental methamidophos degradation step with detected DMPT.",
    },
    {
        "pesticide": "Dimethoate",
        "decision": "Not integrated in this batch",
        "reason": "The most relevant Dimethoate files remain HTML download placeholders. No real full text with named dimethoate microbial transformation products was available locally for this batch.",
    },
]


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 72)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(ROWS)
    decisions = pd.DataFrame(SCREENING_DECISIONS)
    df.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch19_final.csv", index=False)
    decisions.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch19_screening_decisions.csv", index=False)
    df.to_csv(MASTER_CSV, index=False)
    with pd.ExcelWriter(OUT_DIR / "pesticidedb_stepwise_pathway_batch19_final.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 19", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 19", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    print(f"rows={len(df)}")
    print(f"pesticides={df['pesticide'].nunique()}")


if __name__ == "__main__":
    main()
