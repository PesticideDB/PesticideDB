from __future__ import annotations

from pathlib import Path
import re

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "database_pathway_next_step_package_20260707"


def normalize(value: object) -> str:
    text = str(value or "").strip().lower().replace("&", "and")
    return re.sub(r"[^a-z0-9]+", "", text)


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells[:120])
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 62)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    with_evidence = pd.read_excel(PROJECT_ROOT / "pesticide_data.xlsx")
    no_evidence = pd.read_excel(PROJECT_ROOT / "no_evidence_pesticide.xlsx")
    no_col = "Pesticide" if "Pesticide" in no_evidence.columns else next(c for c in no_evidence.columns if c.lower() == "pesticide")

    inventory = pd.read_csv(PROJECT_ROOT / "curation_outputs" / "evidence_pesticide_pdf_inventory_20260707" / "evidence_pesticide_folder_inventory.csv")
    readable = pd.read_csv(PROJECT_ROOT / "curation_outputs" / "evidence_pesticide_pdf_inventory_20260707" / "evidence_pesticide_readable_pdf_queue.csv")
    redownload = pd.read_csv(PROJECT_ROOT / "curation_outputs" / "evidence_pesticide_pdf_inventory_20260707" / "evidence_pesticide_placeholder_redownload_queue.csv")
    top_review = pd.read_csv(PROJECT_ROOT / "curation_outputs" / "consolidated_pathway_review_queue_20260707" / "top_review_set_max5_per_pesticide.csv")

    with_names = sorted({str(v).strip() for v in with_evidence["Pesticide"].dropna() if str(v).strip()}, key=str.casefold)
    no_names = sorted({str(v).strip() for v in no_evidence[no_col].dropna() if str(v).strip()}, key=str.casefold)
    with_norm = {normalize(v): v for v in with_names}
    no_norm = {normalize(v): v for v in no_names}
    overlap = sorted(set(with_norm) & set(no_norm))

    audit = pd.DataFrame(
        [
            ["With-evidence unique pesticides used", len(with_norm)],
            ["No-evidence unique pesticides checked/excluded", len(no_norm)],
            ["Overlap between with-evidence and no-evidence lists", len(overlap)],
            ["Inventory rows generated from with-evidence list", len(inventory)],
            ["Evidence-positive folders with readable PDFs", int((inventory["real_pdf_count"] > 0).sum())],
            ["Evidence-positive folders blocked by placeholder PDFs", len(redownload)],
            ["Evidence-positive pesticides missing USB folder", int((inventory["folder_found"] == "No").sum())],
            ["Top review rows, max five per pesticide", len(top_review)],
            ["Database integration status", "Not integrated; files are for manual pathway/evidence validation first."],
        ],
        columns=["check", "result"],
    )

    excluded_no_evidence = pd.DataFrame({"excluded_no_evidence_pesticide": no_names})
    overlap_df = pd.DataFrame(
        [{"with_evidence_name": with_norm[k], "no_evidence_name": no_norm[k], "normalized_key": k} for k in overlap]
    )

    pathway_import_template = top_review.copy()
    for col in [
        "validated_for_database",
        "pathway_name",
        "step_order",
        "substrate",
        "product",
        "reaction_label",
        "enzyme_or_gene",
        "microorganism_validated",
        "evidence_strength",
        "arrow_style_for_viewer",
        "reference_doi_validated",
        "curator_notes",
    ]:
        pathway_import_template[col] = ""
    pathway_import_template["validated_for_database"] = "No"
    pathway_import_template["evidence_strength"] = "Needs manual validation"
    pathway_import_template["arrow_style_for_viewer"] = "solid=experimentally confirmed; dashed=inferred/partial; dotted=proposed"

    next_actions = pd.DataFrame(
        [
            ["1", "Review Top Review Max5 Each", "Open the top-review sheet and confirm which papers truly report microbial pesticide degradation."],
            ["2", "Fill pathway import template", "For confirmed papers, fill substrate, product, enzyme/gene, microorganism, DOI, and evidence strength."],
            ["3", "Re-download blocked PDFs", "Use the re-download queue for evidence-positive pesticides whose current files are HTML placeholders."],
            ["4", "Do not use no-evidence pesticides", "No-evidence pesticides are excluded from pathway curation until experimental evidence is added."],
            ["5", "Integrate only validated rows", "After manual validation, import only rows marked validated_for_database = Yes."],
        ],
        columns=["step", "task", "meaning"],
    )

    audit.to_csv(OUT_DIR / "with_without_evidence_audit.csv", index=False)
    pathway_import_template.to_csv(OUT_DIR / "pathway_database_import_review_template.csv", index=False)
    redownload.to_csv(OUT_DIR / "evidence_positive_redownload_needed.csv", index=False)

    with pd.ExcelWriter(OUT_DIR / "pesticidedb_database_pathway_next_step_package_20260707.xlsx", engine="openpyxl") as writer:
        audit.to_excel(writer, sheet_name="Evidence List Audit", index=False)
        next_actions.to_excel(writer, sheet_name="Next Actions", index=False)
        pathway_import_template.to_excel(writer, sheet_name="Pathway Import Review", index=False)
        top_review.to_excel(writer, sheet_name="Top Paper Review", index=False)
        readable.to_excel(writer, sheet_name="Readable PDF Folders", index=False)
        redownload.to_excel(writer, sheet_name="Re-download Needed", index=False)
        inventory.to_excel(writer, sheet_name="Full Evidence Inventory", index=False)
        excluded_no_evidence.to_excel(writer, sheet_name="Excluded No Evidence", index=False)
        overlap_df.to_excel(writer, sheet_name="Overlap Check", index=False)
        style_workbook(writer)

    print(f"output={OUT_DIR / 'pesticidedb_database_pathway_next_step_package_20260707.xlsx'}")
    print(f"with_evidence={len(with_norm)}")
    print(f"no_evidence={len(no_norm)}")
    print(f"overlap={len(overlap)}")
    print(f"top_review={len(top_review)}")


if __name__ == "__main__":
    main()
