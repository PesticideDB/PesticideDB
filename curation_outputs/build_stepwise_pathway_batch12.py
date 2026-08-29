from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "stepwise_pathway_batch12_20260708"
MASTER_XLSX = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch12.xlsx"
MASTER_CSV = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch12.csv"


ROWS = [
    {
        "pesticide": "Malathion",
        "pathway_name": "D1CarE5 carboxylesterase-mediated malathion hydrolysis",
        "microorganism": "Alicyclobacillus tengchongensis; recombinant Escherichia coli BL21(DE3)",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Malathion",
        "product": "Malathion monoacid and diacid derivatives",
        "reaction_label": "Carboxylesterase hydrolysis",
        "enzyme": "Malathion-hydrolyzing carboxylesterase D1CarE5",
        "gene": "D1CarE5",
        "evidence_type": "PURIFIED_ENZYME",
        "doi": "10.1007/s10529-013-1195-5",
        "reference_title": "Heterologous expression and characterization of a malathion-hydrolyzing carboxylesterase from a thermophilic bacterium, Alicyclobacillus tengchongensis",
        "source_pdf": "Malathion_10.1007_s10529-013-1195-5.pdf",
        "evidence_note": "The purified recombinant D1CarE5 enzyme hydrolyzed malathion; the paper describes carboxylesterase detoxification of malathion to monoacid and diacid derivatives.",
    },
    {
        "pesticide": "Chlorpyrifos-methyl",
        "pathway_name": "Bacillus megaterium and Pseudomonas syringae chlorpyrifos-methyl biodegradation",
        "microorganism": "Bacillus megaterium CM-Z19; Pseudomonas syringae CM-Z6",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Chlorpyrifos-methyl",
        "product": "3,5,6-trichloro-2-pyridinol",
        "reaction_label": "Phosphoester hydrolysis",
        "enzyme": "Phosphoesterase activity",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1590/0001-3765201920180694",
        "reference_title": "Comparative study on the biodegradation of chlorpyrifos-methyl by Bacillus megaterium CM-Z19 and Pseudomonas syringae CM-Z6",
        "source_pdf": "Chlorpyrifos-methyl_10.1590_0001-3765201920180694.pdf",
        "evidence_note": "The paper tracked chlorpyrifos-methyl degradation and TCP formation/degradation; no TCP was detected for strain CM-Z6 because generated TCP was rapidly degraded.",
    },
]


SCREENING_DECISIONS = [
    {
        "pesticide": "Paraquat",
        "decision": "Not integrated in this batch",
        "reason": "The verified PDF reports paraquat anaerobic transformation and formation of insoluble crystals, but the product is not represented as a named pathway compound suitable for a KEGG-like arrow.",
    },
    {
        "pesticide": "Dichlorvos",
        "decision": "Not integrated in this batch",
        "reason": "The verified PDF focuses on Trichoderma tolerance and multi-omics response under dichlorvos stress; it does not provide a named dichlorvos transformation product for a pathway arrow.",
    },
    {
        "pesticide": "Endosulfan",
        "decision": "Not integrated in this batch",
        "reason": "The verified PDF confirms fungal degradation of alpha-endosulfan, beta-endosulfan, and endosulfan sulfate, but extracted text does not provide a directed substrate-to-product step from the study.",
    },
    {
        "pesticide": "Chlorothalonil",
        "decision": "Not integrated in this batch",
        "reason": "The verified PDF reports chlorothalonil dissipation by Stenotrophomonas acidaminiphila BJ1, but does not identify the degradation product in the study text.",
    },
]


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 64)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(ROWS)
    decisions = pd.DataFrame(SCREENING_DECISIONS)
    df.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch12_final.csv", index=False)
    decisions.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch12_screening_decisions.csv", index=False)
    df.to_csv(MASTER_CSV, index=False)
    with pd.ExcelWriter(OUT_DIR / "pesticidedb_stepwise_pathway_batch12_final.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 12", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 12", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    print(f"rows={len(df)}")
    print(f"pesticides={df['pesticide'].nunique()}")


if __name__ == "__main__":
    main()
