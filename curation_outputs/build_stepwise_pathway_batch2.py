from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "stepwise_pathway_batch2_20260707"
MASTER_XLSX = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch2.xlsx"
MASTER_CSV = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch2.csv"


ROWS = [
    {
        "pesticide": "Teflubenzuron",
        "pathway_name": "Teflubenzuron transformation by microbial strains",
        "microorganism": "Bacillus brevis 625; Alcaligenes sp. 1431; Pseudomonas sp. 10W; Acinetobacter calcoaceticus 21",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Teflubenzuron",
        "product": "2,6-difluorobenzamide",
        "reaction_label": "Hydrolytic cleavage of phenylurea bridge",
        "enzyme": "Unassigned microbial hydrolysis activity",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1081/PFC-100106185",
        "reference_title": "Transformation of the insecticide teflubenzuron by microorganisms",
        "source_pdf": "Transformation of the insecticide teflubenzuron by microorganisms..pdf",
        "evidence_note": "TLC, HPLC, MS and 19F NMR identified 2,6-difluorobenzamide as a microbial transformation product.",
    },
    {
        "pesticide": "Teflubenzuron",
        "pathway_name": "Teflubenzuron transformation by microbial strains",
        "microorganism": "Bacillus brevis 625; Alcaligenes sp. 1431; Pseudomonas sp. 10W; Acinetobacter calcoaceticus 21",
        "completeness": "PARTIAL",
        "step_order": 2,
        "substrate": "2,6-difluorobenzamide",
        "product": "2,6-difluorobenzoic acid",
        "reaction_label": "Amide hydrolysis",
        "enzyme": "Unassigned microbial hydrolysis activity",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1081/PFC-100106185",
        "reference_title": "Transformation of the insecticide teflubenzuron by microorganisms",
        "source_pdf": "Transformation of the insecticide teflubenzuron by microorganisms..pdf",
        "evidence_note": "2,6-difluorobenzoic acid was identified and confirmed with a reference compound and 19F NMR.",
    },
    {
        "pesticide": "Teflubenzuron",
        "pathway_name": "Teflubenzuron transformation by microbial strains",
        "microorganism": "Bacillus brevis 625; Alcaligenes sp. 1431; Pseudomonas sp. 10W; Acinetobacter calcoaceticus 21",
        "completeness": "PARTIAL",
        "step_order": 3,
        "substrate": "Teflubenzuron",
        "product": "2,4-difluoro-3,5-dichloroaniline",
        "reaction_label": "Phenylurea-bridge cleavage",
        "enzyme": "Unassigned microbial transformation activity",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1081/PFC-100106185",
        "reference_title": "Transformation of the insecticide teflubenzuron by microorganisms",
        "source_pdf": "Transformation of the insecticide teflubenzuron by microorganisms..pdf",
        "evidence_note": "2,4-difluoro-3,5-dichloroaniline was identified by mass spectrometry as a minor product.",
    },
    {
        "pesticide": "Teflubenzuron",
        "pathway_name": "Teflubenzuron transformation by microbial strains",
        "microorganism": "Bacillus brevis 625; Alcaligenes sp. 1431; Pseudomonas sp. 10W; Acinetobacter calcoaceticus 21",
        "completeness": "PARTIAL",
        "step_order": 4,
        "substrate": "Teflubenzuron",
        "product": "1,3-bis(2,4-difluoro-3,5-dichlorophenyl)urea",
        "reaction_label": "Condensation/transformation product formation",
        "enzyme": "Unassigned microbial transformation activity",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1081/PFC-100106185",
        "reference_title": "Transformation of the insecticide teflubenzuron by microorganisms",
        "source_pdf": "Transformation of the insecticide teflubenzuron by microorganisms..pdf",
        "evidence_note": "The condensed product was identified by high-resolution MS.",
    },
    {
        "pesticide": "Vinclozolin",
        "pathway_name": "Vinclozolin chemical and biological transformation",
        "microorganism": "Mixed bacterial cultures MA and MB; Corynebacterium sp.",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Vinclozolin",
        "product": "2-[[(3,5-dichlorophenyl)carbamoyl]oxy]-2-methyl-3-butenoic acid",
        "reaction_label": "Oxazolidinedione-ring opening",
        "enzyme": "Unassigned transformation activity",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "PMID: 18309282",
        "reference_title": "Chemical and biological transformation of the fungicide vinclozolin",
        "source_pdf": "Chemical and biological transformation of the fungicide vinclozolin..pdf",
        "evidence_note": "The paper reports this open-form butenoic acid derivative as a vinclozolin transformation product.",
    },
    {
        "pesticide": "Vinclozolin",
        "pathway_name": "Vinclozolin chemical and biological transformation",
        "microorganism": "Mixed bacterial cultures MA and MB; Corynebacterium sp.",
        "completeness": "PARTIAL",
        "step_order": 2,
        "substrate": "2-[[(3,5-dichlorophenyl)carbamoyl]oxy]-2-methyl-3-butenoic acid",
        "product": "3,5-dichloroaniline",
        "reaction_label": "Anilide/aniline product formation",
        "enzyme": "Unassigned transformation activity",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "PMID: 18309282",
        "reference_title": "Chemical and biological transformation of the fungicide vinclozolin",
        "source_pdf": "Chemical and biological transformation of the fungicide vinclozolin..pdf",
        "evidence_note": "Mixed bacterial cultures converted vinclozolin through compound 2 toward 3,5-dichloroaniline.",
    },
    {
        "pesticide": "Vinclozolin",
        "pathway_name": "Vinclozolin chemical and biological transformation",
        "microorganism": "Mixed bacterial cultures MA and MB; Corynebacterium sp.",
        "completeness": "PARTIAL",
        "step_order": 3,
        "substrate": "Vinclozolin",
        "product": "3',5'-dichloro-2-hydroxy-2-methylbut-3-enanilide",
        "reaction_label": "Alternative ring-opening transformation",
        "enzyme": "Unassigned transformation activity",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "",
        "reference_title": "Chemical and biological transformation of the fungicide vinclozolin",
        "source_pdf": "Chemical and biological transformation of the fungicide vinclozolin..pdf",
        "evidence_note": "The paper reports a second pathway through the enanilide transformation product.",
    },
    {
        "pesticide": "Vinclozolin",
        "pathway_name": "Vinclozolin chemical and biological transformation",
        "microorganism": "Mixed bacterial cultures MA and MB; Corynebacterium sp.",
        "completeness": "PARTIAL",
        "step_order": 4,
        "substrate": "3',5'-dichloro-2-hydroxy-2-methylbut-3-enanilide",
        "product": "3,5-dichloroaniline",
        "reaction_label": "Aniline product formation",
        "enzyme": "Unassigned transformation activity",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "",
        "reference_title": "Chemical and biological transformation of the fungicide vinclozolin",
        "source_pdf": "Chemical and biological transformation of the fungicide vinclozolin..pdf",
        "evidence_note": "The paper reports conversion of the enanilide product toward 3,5-dichloroaniline.",
    },
    {
        "pesticide": "Vinclozolin",
        "pathway_name": "Vinclozolin biodegradation by Rhodococcus sp. T1-1",
        "microorganism": "Rhodococcus sp. T1-1",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Vinclozolin",
        "product": "3,5-dichloroaniline",
        "reaction_label": "Biodegradation to dichloroaniline metabolite",
        "enzyme": "Unassigned activity",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "",
        "reference_title": "Microbial biodegradation and toxicity of vinclozolin and its toxic metabolite 3,5-dichloroaniline",
        "source_pdf": "Microbial biodegradation and toxicity of vinclozolin and its toxic metabolite 3,5-dichloroaniline..pdf",
        "evidence_note": "3,5-dichloroaniline accumulated during vinclozolin degradation by Rhodococcus sp. T1-1.",
    },
    {
        "pesticide": "Vinclozolin",
        "pathway_name": "Vinclozolin biodegradation by Rhodococcus sp. T1-1",
        "microorganism": "Rhodococcus sp. T1-1",
        "completeness": "PARTIAL",
        "step_order": 2,
        "substrate": "3,5-dichloroaniline",
        "product": "Phenol",
        "reaction_label": "Dichloroaniline degradation",
        "enzyme": "Unassigned activity",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "",
        "reference_title": "Microbial biodegradation and toxicity of vinclozolin and its toxic metabolite 3,5-dichloroaniline",
        "source_pdf": "Microbial biodegradation and toxicity of vinclozolin and its toxic metabolite 3,5-dichloroaniline..pdf",
        "evidence_note": "GC/MS detected phenol from degraded 3,5-dichloroaniline.",
    },
]


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 64)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(ROWS)
    df.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch2_final.csv", index=False)
    df.to_csv(MASTER_CSV, index=False)
    with pd.ExcelWriter(OUT_DIR / "pesticidedb_stepwise_pathway_batch2_final.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 2", index=False)
        style_workbook(writer)
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 2", index=False)
        style_workbook(writer)
    print(f"rows={len(df)}")
    print(f"pesticides={df['pesticide'].nunique()}")
    print(f"project_master={MASTER_XLSX}")


if __name__ == "__main__":
    main()
