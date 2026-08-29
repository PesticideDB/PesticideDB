from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "stepwise_pathway_batch9_20260708"
MASTER_XLSX = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch9.xlsx"
MASTER_CSV = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch9.csv"


ISOPYRAZAM_SOURCE = "Fungicide isopyrazam degradative response toward extrinsically added fungal and bacterial strains..pdf"
ISOPYRAZAM_REFERENCE = "Fungicide isopyrazam degradative response toward extrinsically added fungal and bacterial strains"
ISOPYRAZAM_DOI = "10.1002/jobm.201900687"
ISOPYRAZAM_PATHWAY = "GC-MS-supported microbial degradation route of isopyrazam"
ISOPYRAZAM_MICROBES = (
    "Aspergillus flavus; Penicillium chrysogenum; Aspergillus niger; "
    "Aspergillus terreus; Aspergillus fumigatus; Xanthomonas axonopodis; "
    "Pseudomonas syringae"
)

SPIRODICLOFEN_SOURCE = "Enhanced degradation of spiro-insecticides and their leacher enol derivatives in soil by solarization and biosolarization techniques..pdf"
SPIRODICLOFEN_REFERENCE = "Enhanced degradation of spiro-insecticides and their leacher enol derivatives in soil by solarization and biosolarization techniques"
SPIRODICLOFEN_DOI = "10.1007/s11356-017-8589-1"


ROWS = [
    {
        "pesticide": "Isopyrazam",
        "pathway_name": ISOPYRAZAM_PATHWAY,
        "microorganism": ISOPYRAZAM_MICROBES,
        "completeness": "PROPOSED",
        "step_order": 1,
        "substrate": "Isopyrazam",
        "product": "3-difluoromethyl-1-methyl-1H-pyrazole-4-carboxylic acid (9-hydroxy-9-isopropyl-1,2,3,4-tetrahydro-1,4-methanonaphthalen-5-yl)-amide",
        "reaction_label": "Hydroxylation of isopropyl group",
        "enzyme": "Microbial hydroxylation activity; enzyme not genetically assigned",
        "gene": "",
        "evidence_type": "PROPOSED",
        "doi": ISOPYRAZAM_DOI,
        "reference_title": ISOPYRAZAM_REFERENCE,
        "source_pdf": ISOPYRAZAM_SOURCE,
        "evidence_note": "The paper reports GC-MS metabolite identification after incubation with fungal and bacterial strains and presents this as product B in the possible isopyrazam degradation route.",
    },
    {
        "pesticide": "Isopyrazam",
        "pathway_name": ISOPYRAZAM_PATHWAY,
        "microorganism": ISOPYRAZAM_MICROBES,
        "completeness": "PROPOSED",
        "step_order": 2,
        "substrate": "3-difluoromethyl-1-methyl-1H-pyrazole-4-carboxylic acid (9-hydroxy-9-isopropyl-1,2,3,4-tetrahydro-1,4-methanonaphthalen-5-yl)-amide",
        "product": "3-difluoromethyl-1-methyl-1H-pyrazole-4-carboxylic acid (9-isopropylidene-1,2,3,4-tetrahydro-1,4-methanonaphthalen-5-yl)-amide",
        "reaction_label": "Hydroxylation/dehydration of bicyclic substituent",
        "enzyme": "Microbial transformation activity; enzyme not genetically assigned",
        "gene": "",
        "evidence_type": "PROPOSED",
        "doi": ISOPYRAZAM_DOI,
        "reference_title": ISOPYRAZAM_REFERENCE,
        "source_pdf": ISOPYRAZAM_SOURCE,
        "evidence_note": "The paper presents this as product C in the possible mechanistic route and describes hydroxylation of the isopropyl group and bicyclic ring.",
    },
    {
        "pesticide": "Isopyrazam",
        "pathway_name": ISOPYRAZAM_PATHWAY,
        "microorganism": ISOPYRAZAM_MICROBES,
        "completeness": "PROPOSED",
        "step_order": 3,
        "substrate": "3-difluoromethyl-1-methyl-1H-pyrazole-4-carboxylic acid (9-isopropylidene-1,2,3,4-tetrahydro-1,4-methanonaphthalen-5-yl)-amide",
        "product": "3-difluoromethyl-1H-pyrazole-4-carboxylic acid",
        "reaction_label": "Amide-linkage breakdown and N-demethylation",
        "enzyme": "Microbial amide-cleavage/demethylation activity; enzyme not genetically assigned",
        "gene": "",
        "evidence_type": "PROPOSED",
        "doi": ISOPYRAZAM_DOI,
        "reference_title": ISOPYRAZAM_REFERENCE,
        "source_pdf": ISOPYRAZAM_SOURCE,
        "evidence_note": "The paper identifies pyrazole half-molecule acids and states that amide bond breakdown and demethylation are major processes in isopyrazam biotransformation.",
    },
    {
        "pesticide": "Isopyrazam",
        "pathway_name": ISOPYRAZAM_PATHWAY,
        "microorganism": ISOPYRAZAM_MICROBES,
        "completeness": "PROPOSED",
        "step_order": 4,
        "substrate": "3-difluoromethyl-1-methyl-1H-pyrazole-4-carboxylic acid (9-isopropylidene-1,2,3,4-tetrahydro-1,4-methanonaphthalen-5-yl)-amide",
        "product": "3-difluoromethyl-1-methyl-1H-pyrazole-4-amide",
        "reaction_label": "Amide-linkage breakdown",
        "enzyme": "Microbial amide-cleavage activity; enzyme not genetically assigned",
        "gene": "",
        "evidence_type": "PROPOSED",
        "doi": ISOPYRAZAM_DOI,
        "reference_title": ISOPYRAZAM_REFERENCE,
        "source_pdf": ISOPYRAZAM_SOURCE,
        "evidence_note": "The paper presents this pyrazole amide as product E in the proposed microbial degradation route.",
    },
    {
        "pesticide": "Spirodiclofen",
        "pathway_name": "Soil transformation of spirodiclofen to enol derivative",
        "microorganism": "Soil microcosm under solarization/biosolarization; microbial activity implicated",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Spirodiclofen",
        "product": "Spirodiclofen-enol",
        "reaction_label": "Hydrolysis to enol derivative",
        "enzyme": "Soil biochemical/hydrolytic transformation; enzyme not genetically assigned",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": SPIRODICLOFEN_DOI,
        "reference_title": SPIRODICLOFEN_REFERENCE,
        "source_pdf": SPIRODICLOFEN_SOURCE,
        "evidence_note": "The paper reports spirodiclofen-enol as a transformation product in soil/leachate experiments and states degradation was enhanced under solarization and biosolarization, partly associated with microbial activity changes.",
    },
]


SCREENING_DECISIONS = [
    {
        "pesticide": "Fenamiphos",
        "decision": "Not integrated in this batch",
        "reason": "The readable files in the Fenamiphos folder are fenamidone/fungicide residue papers, not fenamiphos pathway papers.",
    },
    {
        "pesticide": "Propamocarb",
        "decision": "Not integrated in this batch",
        "reason": "The microbial carbamate paper reports propamocarb was resistant to attack by the tested microbial species, so no product arrow was imported.",
    },
    {
        "pesticide": "Propargite",
        "decision": "Not integrated in this batch",
        "reason": "The Pseudomonas putida paper explicitly states that only propargite degradation was studied, not its metabolites.",
    },
    {
        "pesticide": "Pinoxaden",
        "decision": "Not integrated in this batch",
        "reason": "The Acinetobacter paper supports degradation and predicted genes but does not provide named transformation products suitable for a stepwise arrow.",
    },
    {
        "pesticide": "Fluxapyroxad",
        "decision": "Not integrated in this batch",
        "reason": "The yeast and soil/water-sediment papers quantify dissipation/degradation, but the readable sources did not provide a microbial product pathway for import.",
    },
    {
        "pesticide": "Penthiopyrad",
        "decision": "Not integrated in this batch",
        "reason": "The PAM evidence is mainly plant/field residue hydrolysis rather than a microbial pathway record, so it was left out of the microbial pathway layer.",
    },
    {
        "pesticide": "Acibenzolar-S-methyl",
        "decision": "Not integrated in this batch",
        "reason": "Readable papers were PGPR/residue/plant-quality studies without named microbial transformation products.",
    },
    {
        "pesticide": "Imazamox",
        "decision": "Not integrated in this batch",
        "reason": "The strongest named-product evidence in the local files was photodegradation or mixed sunlight/soil behavior rather than a clean microbial pathway record.",
    },
]


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 64)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(ROWS)
    decisions = pd.DataFrame(SCREENING_DECISIONS)
    df.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch9_final.csv", index=False)
    decisions.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch9_screening_decisions.csv", index=False)
    df.to_csv(MASTER_CSV, index=False)
    with pd.ExcelWriter(OUT_DIR / "pesticidedb_stepwise_pathway_batch9_final.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 9", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 9", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    print(f"rows={len(df)}")
    print(f"pesticides={df['pesticide'].nunique()}")


if __name__ == "__main__":
    main()
