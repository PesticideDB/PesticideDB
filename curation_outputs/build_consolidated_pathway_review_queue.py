from __future__ import annotations

from pathlib import Path
import re
import subprocess

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
MANIFEST = PROJECT_ROOT / "curation_outputs" / "evidence_pesticide_readable_curation_manifest_20260707" / "readable_folder_curation_manifest.csv"
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "consolidated_pathway_review_queue_20260707"
PDFTOTEXT = "/Users/nana/.cache/codex-runtimes/codex-primary-runtime/dependencies/native/poppler/poppler/bin/pdftotext"

DEGRADATION_TERMS = [
    "biodegradation",
    "biodegrade",
    "microbial degradation",
    "degradation",
    "degrade",
    "mineralization",
    "transformation",
    "biotransformation",
    "metabolite",
    "intermediate",
]

MICROBIAL_TERMS = [
    "bacteria",
    "bacterial",
    "fungus",
    "fungal",
    "fungi",
    "microbial",
    "microorganism",
    "strain",
    "isolate",
    "culture",
    "bacillus",
    "pseudomonas",
    "sphingomonas",
    "arthrobacter",
    "aspergillus",
    "trichoderma",
    "rhizobium",
    "stenotrophomonas",
]

MOLECULAR_CONTEXT_TERMS = [
    "gene",
    "enzyme",
    "protein",
    "monooxygenase",
    "dioxygenase",
    "hydrolase",
    "dehalogenase",
    "esterase",
    "amidase",
    "dehydrogenase",
]

ABIOTIC_FALSE_POSITIVE_TERMS = [
    "photocatalytic",
    "photolysis",
    "sonochemical",
    "adsorption",
    "advanced oxidation",
    "ozonation",
    "electrochemical",
    "hydrothermal",
    "protein degradation",
    "degradation of transcription factor",
    "tumor",
    "cancer",
]


def safe_name(name: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "_", name).strip("_").lower()


def clean_text(text: str) -> str:
    text = text.replace("\x00", " ")
    text = re.sub(r"[\x01-\x08\x0B-\x0C\x0E-\x1F]", " ", text)
    text = re.sub(r"-\n", "", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def extract_pdf_text(pdf_path: Path) -> str:
    try:
        result = subprocess.run(
            [PDFTOTEXT, "-layout", str(pdf_path), "-"],
            check=False,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=45,
        )
    except subprocess.TimeoutExpired:
        return ""
    if result.returncode != 0:
        return ""
    return clean_text(result.stdout)


def pesticide_regex(pesticide: str) -> re.Pattern[str]:
    parts = [re.escape(part) for part in re.split(r"[\s_-]+", pesticide.strip()) if part]
    if not parts:
        return re.compile(r"a^")
    return re.compile(r"\b" + r"[\s_-]*".join(parts) + r"\b", re.I)


def context_metrics(row: pd.Series) -> dict[str, int | str]:
    pesticide = str(row.get("pesticide", "")).strip()
    curation_folder = Path(str(row.get("curation_folder", "")))
    pdf_path = curation_folder.parent / str(row.get("pdf_file", ""))
    text = extract_pdf_text(pdf_path)
    lower = text.lower()
    title = str(row.get("probable_title", "")).lower()
    pattern = pesticide_regex(pesticide)
    matches = list(pattern.finditer(text))
    pesticide_mentions = len(matches)

    proximity_hits = 0
    microbial_near_pesticide = 0
    molecular_near_pesticide = 0
    best_context = ""
    for match in matches[:40]:
        start = max(0, match.start() - 500)
        end = min(len(text), match.end() + 500)
        window = text[start:end]
        wlower = window.lower()
        if any(term in wlower for term in DEGRADATION_TERMS):
            proximity_hits += 1
            if not best_context:
                best_context = window[:900]
        if any(term in wlower for term in MICROBIAL_TERMS):
            microbial_near_pesticide += 1
        if any(term in wlower for term in MOLECULAR_CONTEXT_TERMS):
            molecular_near_pesticide += 1

    abiotic_flags = sum(1 for term in ABIOTIC_FALSE_POSITIVE_TERMS if term in lower or term in title)
    title_has_pesticide = 1 if pattern.search(str(row.get("probable_title", ""))) else 0
    title_has_biodegradation = 1 if any(term in title for term in DEGRADATION_TERMS) else 0

    return {
        "pesticide_mentions": pesticide_mentions,
        "pesticide_degradation_contexts": proximity_hits,
        "microbial_contexts_near_pesticide": microbial_near_pesticide,
        "gene_enzyme_contexts_near_pesticide": molecular_near_pesticide,
        "abiotic_or_irrelevant_flags": abiotic_flags,
        "title_has_pesticide": title_has_pesticide,
        "title_has_degradation_term": title_has_biodegradation,
        "best_pesticide_context": clean_text(best_context)[:900],
    }


def score_row(row: pd.Series) -> tuple[int, str, str, str]:
    biodeg = int(row.get("biodegradation_term_hits", 0) or 0)
    molecular = int(row.get("molecular_term_hits", 0) or 0)
    pathway = int(row.get("pathway_term_hits", 0) or 0)
    signal = str(row.get("automated_evidence_signal", ""))
    pesticide_mentions = int(row.get("pesticide_mentions", 0) or 0)
    degradation_contexts = int(row.get("pesticide_degradation_contexts", 0) or 0)
    microbial_contexts = int(row.get("microbial_contexts_near_pesticide", 0) or 0)
    molecular_contexts = int(row.get("gene_enzyme_contexts_near_pesticide", 0) or 0)
    abiotic_flags = int(row.get("abiotic_or_irrelevant_flags", 0) or 0)
    title_has_pesticide = int(row.get("title_has_pesticide", 0) or 0)
    title_has_degradation = int(row.get("title_has_degradation_term", 0) or 0)

    score = 0
    score += min(pesticide_mentions, 8) * 4
    score += min(degradation_contexts, 8) * 9
    score += min(microbial_contexts, 8) * 8
    score += min(molecular_contexts, 8) * 8
    score += min(biodeg, 5) * 1
    score += min(molecular, 5) * 1
    score += min(pathway, 5) * 1
    score += title_has_pesticide * 10
    score += title_has_degradation * 8
    if "gene/enzyme_plus_pathway" in signal:
        score += 8
    elif "gene/enzyme" in signal:
        score += 6
    elif "pathway/metabolite" in signal:
        score += 5
    elif "whole-cell" in signal:
        score += 4
    score -= abiotic_flags * 18

    if degradation_contexts > 0 and microbial_contexts > 0 and score >= 70:
        priority = "High"
        action = "Review first: likely useful for pathway/protein evidence extraction."
        use = "Pesticide-specific microbial degradation candidate"
    elif degradation_contexts > 0 and score >= 35:
        priority = "Medium"
        action = "Review after high-priority papers; may contain useful evidence or context."
        use = "Pesticide-specific degradation candidate; microbial/protein evidence uncertain"
    else:
        priority = "Low"
        action = "Keep as background unless title/snippet clearly shows biodegradation evidence."
        use = "Background or likely non-microbial/non-specific"

    return score, priority, action, use


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells[:120])
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 64)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    manifest = pd.read_csv(MANIFEST)

    screening_frames = []
    evidence_frames = []
    snippet_frames = []

    for _, row in manifest.iterrows():
        pesticide = row["pesticide"]
        folder = Path(row["output_folder"])
        prefix = safe_name(pesticide)

        screening_path = folder / f"{prefix}_paper_screening.csv"
        evidence_path = folder / f"{prefix}_evidence_candidates.csv"

        if screening_path.exists():
            df = pd.read_csv(screening_path)
            df["curation_folder"] = str(folder)
            screening_frames.append(df)
        if evidence_path.exists():
            df = pd.read_csv(evidence_path)
            df["curation_folder"] = str(folder)
            evidence_frames.append(df)

        snippet_path = folder / f"{prefix}_keyword_snippets.csv"
        if snippet_path.exists():
            df = pd.read_csv(snippet_path)
            df["curation_folder"] = str(folder)
            snippet_frames.append(df)

    screening = pd.concat(screening_frames, ignore_index=True) if screening_frames else pd.DataFrame()
    evidence = pd.concat(evidence_frames, ignore_index=True) if evidence_frames else pd.DataFrame()

    if not screening.empty:
        metrics = screening.apply(context_metrics, axis=1, result_type="expand")
        screening = pd.concat([screening, metrics], axis=1)
        scored = screening.apply(score_row, axis=1, result_type="expand")
        screening["review_score"] = scored[0]
        screening["review_priority"] = scored[1]
        screening["recommended_review_action"] = scored[2]
        screening["likely_database_use"] = scored[3]
        screening = screening.sort_values(
            ["review_priority", "review_score", "pesticide", "pdf_file"],
            ascending=[True, False, True, True],
        )
        priority_order = {"High": 0, "Medium": 1, "Low": 2}
        screening["_priority_order"] = screening["review_priority"].map(priority_order).fillna(9)
        screening = screening.sort_values(["_priority_order", "review_score"], ascending=[True, False]).drop(columns=["_priority_order"])

    high_priority = screening[screening["review_priority"] == "High"].copy() if not screening.empty else pd.DataFrame()
    medium_priority = screening[screening["review_priority"] == "Medium"].copy() if not screening.empty else pd.DataFrame()
    top_review = (
        screening[screening["review_priority"].isin(["High", "Medium"])]
        .sort_values(["pesticide", "review_score"], ascending=[True, False])
        .groupby("pesticide", as_index=False, group_keys=False)
        .head(5)
        .sort_values(["pesticide", "review_score"], ascending=[True, False])
        if not screening.empty
        else pd.DataFrame()
    )

    summary = pd.DataFrame(
        [
            ["Readable pesticide folders processed", manifest.shape[0]],
            ["Readable PDFs screened", int(manifest["readable_pdfs_processed"].sum())],
            ["High-priority papers", len(high_priority)],
            ["Medium-priority papers", len(medium_priority)],
            ["Top review set, max 5 per pesticide", len(top_review)],
            ["Low-priority/background papers", int((screening["review_priority"] == "Low").sum()) if not screening.empty else 0],
            ["Candidate evidence rows", len(evidence)],
            ["Recommended next action", "Manually validate High-priority rows first, then fill products/intermediates in each pesticide pathway template."],
        ],
        columns=["metric", "value"],
    )

    screening.to_csv(OUT_DIR / "consolidated_paper_screening_ranked.csv", index=False)
    high_priority.to_csv(OUT_DIR / "high_priority_pathway_paper_review_queue.csv", index=False)
    top_review.to_csv(OUT_DIR / "top_review_set_max5_per_pesticide.csv", index=False)
    evidence.to_csv(OUT_DIR / "consolidated_evidence_candidates.csv", index=False)

    with pd.ExcelWriter(OUT_DIR / "pesticidedb_consolidated_pathway_review_queue_20260707.xlsx", engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        top_review.to_excel(writer, sheet_name="Top Review Max5 Each", index=False)
        high_priority.to_excel(writer, sheet_name="High Priority Review", index=False)
        medium_priority.to_excel(writer, sheet_name="Medium Priority Review", index=False)
        screening.to_excel(writer, sheet_name="All Ranked Screening", index=False)
        evidence.to_excel(writer, sheet_name="Evidence Candidates", index=False)
        manifest.to_excel(writer, sheet_name="Folder Manifest", index=False)
        style_workbook(writer)

    print(f"output={OUT_DIR / 'pesticidedb_consolidated_pathway_review_queue_20260707.xlsx'}")
    print(f"folders={manifest.shape[0]}")
    print(f"screened={len(screening)}")
    print(f"high={len(high_priority)}")
    print(f"medium={len(medium_priority)}")
    print(f"evidence_candidates={len(evidence)}")


if __name__ == "__main__":
    main()
