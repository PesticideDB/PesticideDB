from __future__ import annotations

from pathlib import Path
import re

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
SOURCE_DIR = PROJECT_ROOT / "curation_outputs" / "database_pathway_next_step_package_20260707"
SOURCE_FILE = SOURCE_DIR / "pathway_database_import_review_template.csv"
FINAL_DIR = PROJECT_ROOT / "curation_outputs" / "final_pathway_evidence_master_20260707"
MASTER_XLSX = PROJECT_ROOT / "PesticideDB_Pathway_Evidence_Master.xlsx"
MASTER_CSV = PROJECT_ROOT / "PesticideDB_Pathway_Evidence_Master.csv"


DROP_COLUMNS = {
    "review_decision",
    "review_notes",
    "review_score",
    "review_priority",
    "recommended_review_action",
    "likely_database_use",
    "validated_for_database",
    "curator_notes",
    "evidence_strength",
    "arrow_style_for_viewer",
    "pathway_name",
    "step_order",
    "substrate",
    "product",
    "reaction_label",
    "enzyme_or_gene",
    "microorganism_validated",
    "reference_doi_validated",
}

SIGNAL_MAP = {
    "candidate_gene/enzyme_plus_pathway": "Gene/enzyme and pathway/metabolite evidence",
    "candidate_gene/enzyme": "Gene/enzyme evidence",
    "candidate_pathway/metabolite": "Pathway/metabolite evidence",
    "candidate_whole-cell_or_transformation": "Whole-cell or transformation evidence",
    "low_signal": "Literature evidence",
}


def first_doi(value: object) -> str:
    text = "" if pd.isna(value) else str(value)
    match = re.search(r"\b10\.\d{4,9}/[-._;()/:A-Z0-9]+\b", text, re.I)
    return match.group(0).rstrip(".,;") if match else ""


def final_year(value: object) -> str:
    text = "" if pd.isna(value) else str(value)
    years = re.findall(r"\b(19[7-9]\d|20[0-2]\d)\b", text)
    return years[-1] if years else ""


def clean_text(value: object, limit: int | None = None) -> str:
    if pd.isna(value):
        return ""
    text = re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]", " ", str(value))
    text = re.sub(r"\s+", " ", text).strip()
    return text[:limit] if limit else text


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells[:120])
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 62)


def main() -> None:
    FINAL_DIR.mkdir(parents=True, exist_ok=True)
    source = pd.read_csv(SOURCE_FILE)
    final = pd.DataFrame()
    final["pesticide"] = source["pesticide"].map(clean_text)
    final["paper_title"] = source["probable_title"].map(clean_text)
    final["doi"] = source["doi_candidates"].map(first_doi)
    final["publication_year"] = source["year_candidates"].map(final_year)
    final["microorganism_or_strain"] = source["candidate_microorganisms_or_strains"].map(clean_text)
    final["evidence_category"] = source["automated_evidence_signal"].map(lambda v: SIGNAL_MAP.get(str(v), "Literature evidence"))
    final["source_pdf"] = source["pdf_file"].map(clean_text)
    final["pesticide_specific_context"] = source["best_pesticide_context"].map(lambda v: clean_text(v, 1200))
    final["biodegradation_term_hits"] = source["biodegradation_term_hits"]
    final["gene_enzyme_term_hits"] = source["molecular_term_hits"]
    final["pathway_metabolite_term_hits"] = source["pathway_term_hits"]
    final["pesticide_mentions"] = source["pesticide_mentions"]
    final["pesticide_degradation_contexts"] = source["pesticide_degradation_contexts"]
    final["microbial_contexts_near_pesticide"] = source["microbial_contexts_near_pesticide"]
    final["gene_enzyme_contexts_near_pesticide"] = source["gene_enzyme_contexts_near_pesticide"]
    final["curation_source_folder"] = source["curation_folder"].map(clean_text)

    final = final.drop_duplicates(subset=["pesticide", "paper_title", "doi", "source_pdf"]).sort_values(
        ["pesticide", "publication_year", "paper_title"],
        na_position="last",
    )

    original_copy = FINAL_DIR / "pathway_database_import_review_template_original.csv"
    clean_csv = FINAL_DIR / "pesticidedb_pathway_evidence_master_final.csv"
    clean_xlsx = FINAL_DIR / "pesticidedb_pathway_evidence_master_final.xlsx"
    source.to_csv(original_copy, index=False)
    final.to_csv(clean_csv, index=False)
    final.to_csv(MASTER_CSV, index=False)

    with pd.ExcelWriter(clean_xlsx, engine="openpyxl") as writer:
        final.to_excel(writer, sheet_name="Pathway Evidence Master", index=False)
        style_workbook(writer)
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        final.to_excel(writer, sheet_name="Pathway Evidence Master", index=False)
        style_workbook(writer)

    print(f"rows={len(final)}")
    print(f"final_xlsx={clean_xlsx}")
    print(f"master_xlsx={MASTER_XLSX}")


if __name__ == "__main__":
    main()
