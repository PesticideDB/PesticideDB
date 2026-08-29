from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "stepwise_pathway_batch8_20260708"
MASTER_XLSX = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch8.xlsx"
MASTER_CSV = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch8.csv"


GLUFOSINATE_SOURCE = "Initial steps in the degradation of phosphinothricin (glufosinate) by soil bacteria..pdf"
GLUFOSINATE_REFERENCE = "Initial steps in the degradation of phosphinothricin (glufosinate) by soil bacteria"
GLUFOSINATE_DOI = "10.1128/aem.55.3.711-716.1989"
QUINTOZENE_DH19_SOURCE = "Biodegradation of pentachloronitrobenzene by Arthrobacter nicotianae DH19..pdf"
QUINTOZENE_DH19_REFERENCE = "Biodegradation of pentachloronitrobenzene by Arthrobacter nicotianae DH19"
QUINTOZENE_DH19_DOI = "10.1111/lam.12476"
QUINTOZENE_QTH3_SOURCE = "Effective biodegradation of pentachloronitrobenzene by a novel strain Peudomonas putida QTH3 isolated from contaminated soil..pdf"
QUINTOZENE_QTH3_REFERENCE = "Effective biodegradation of pentachloronitrobenzene by a novel strain Pseudomonas putida QTH3 isolated from contaminated soil"
QUINTOZENE_QTH3_DOI = "10.1016/j.ecoenv.2019.109463"


ROWS = [
    {
        "pesticide": "Glufosinate-Ammonium",
        "pathway_name": "Initial bacterial transformations of phosphinothricin (glufosinate)",
        "microorganism": "Soil bacterial isolates; Rhodococcus sp.",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Glufosinate",
        "product": "2-oxo-4-[(hydroxy)(methyl)phosphinoyl]butyric acid",
        "reaction_label": "Transamination / oxidative deamination",
        "enzyme": "PPT transaminase / PPT oxidase; enzyme not genetically assigned",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": GLUFOSINATE_DOI,
        "reference_title": GLUFOSINATE_REFERENCE,
        "source_pdf": GLUFOSINATE_SOURCE,
        "evidence_note": "The paper reports bacterial transformation of phosphinothricin to the corresponding 2-oxo acid by transamination in multiple strains and oxidative deamination in Rhodococcus sp.",
    },
    {
        "pesticide": "Glufosinate-Ammonium",
        "pathway_name": "Initial bacterial transformations of phosphinothricin (glufosinate)",
        "microorganism": "Soil bacterial isolates",
        "completeness": "PARTIAL",
        "step_order": 2,
        "substrate": "Glufosinate",
        "product": "N-acetyl-PPT",
        "reaction_label": "N-acetylation",
        "enzyme": "PPT N-acetyltransferase; enzyme not genetically assigned",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": GLUFOSINATE_DOI,
        "reference_title": GLUFOSINATE_REFERENCE,
        "source_pdf": GLUFOSINATE_SOURCE,
        "evidence_note": "The paper reports acetyltransferase activity synthesizing N-acetyl-PPT from phosphinothricin and acetyl-CoA.",
    },
    {
        "pesticide": "Quintozene",
        "pathway_name": "Pentachloronitrobenzene biodegradation by Arthrobacter nicotianae DH19",
        "microorganism": "Arthrobacter nicotianae DH19",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Pentachloronitrobenzene (PCNB)",
        "product": "2,3,4,5,6-pentachloroaniline",
        "reaction_label": "Nitro reduction",
        "enzyme": "Nitroreductase; enzyme not genetically assigned",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": QUINTOZENE_DH19_DOI,
        "reference_title": QUINTOZENE_DH19_REFERENCE,
        "source_pdf": QUINTOZENE_DH19_SOURCE,
        "evidence_note": "GC-MS/MS identified pentachloroaniline among PCNB metabolites and the paper deduced nitroreduction as the initial transformation.",
    },
    {
        "pesticide": "Quintozene",
        "pathway_name": "Pentachloronitrobenzene biodegradation by Arthrobacter nicotianae DH19",
        "microorganism": "Arthrobacter nicotianae DH19",
        "completeness": "PARTIAL",
        "step_order": 2,
        "substrate": "2,3,4,5,6-pentachloroaniline",
        "product": "3,5-dichloroaniline",
        "reaction_label": "Stepwise dechlorination",
        "enzyme": "Dechlorination activity; enzyme not genetically assigned",
        "gene": "",
        "evidence_type": "PROPOSED",
        "doi": QUINTOZENE_DH19_DOI,
        "reference_title": QUINTOZENE_DH19_REFERENCE,
        "source_pdf": QUINTOZENE_DH19_SOURCE,
        "evidence_note": "3,5-dichloroaniline was identified as a metabolite and placed in the authors' proposed PCNB biodegradation pathway.",
    },
    {
        "pesticide": "Quintozene",
        "pathway_name": "Pentachloronitrobenzene biodegradation by Arthrobacter nicotianae DH19",
        "microorganism": "Arthrobacter nicotianae DH19",
        "completeness": "PARTIAL",
        "step_order": 3,
        "substrate": "3,5-dichloroaniline",
        "product": "Aniline",
        "reaction_label": "Dechlorination",
        "enzyme": "Dechlorination activity; enzyme not genetically assigned",
        "gene": "",
        "evidence_type": "PROPOSED",
        "doi": QUINTOZENE_DH19_DOI,
        "reference_title": QUINTOZENE_DH19_REFERENCE,
        "source_pdf": QUINTOZENE_DH19_SOURCE,
        "evidence_note": "Aniline was identified as a metabolite and placed downstream of dichloroaniline in the proposed DH19 pathway.",
    },
    {
        "pesticide": "Quintozene",
        "pathway_name": "Pentachloronitrobenzene biodegradation by Arthrobacter nicotianae DH19",
        "microorganism": "Arthrobacter nicotianae DH19",
        "completeness": "PARTIAL",
        "step_order": 4,
        "substrate": "Aniline",
        "product": "Catechol",
        "reaction_label": "Aromatic hydroxylation",
        "enzyme": "Aniline hydroxylation activity; enzyme not genetically assigned",
        "gene": "",
        "evidence_type": "PROPOSED",
        "doi": QUINTOZENE_DH19_DOI,
        "reference_title": QUINTOZENE_DH19_REFERENCE,
        "source_pdf": QUINTOZENE_DH19_SOURCE,
        "evidence_note": "Catechol was identified as a metabolite and the paper states that aniline is further degraded to catechol.",
    },
    {
        "pesticide": "Quintozene",
        "pathway_name": "Pentachloronitrobenzene biodegradation by Arthrobacter nicotianae DH19",
        "microorganism": "Arthrobacter nicotianae DH19",
        "completeness": "PARTIAL",
        "step_order": 5,
        "substrate": "Catechol",
        "product": "Adipic acid",
        "reaction_label": "Ring cleavage / downstream oxidation",
        "enzyme": "Catechol ring-cleavage activity; enzyme not genetically assigned",
        "gene": "",
        "evidence_type": "PROPOSED",
        "doi": QUINTOZENE_DH19_DOI,
        "reference_title": QUINTOZENE_DH19_REFERENCE,
        "source_pdf": QUINTOZENE_DH19_SOURCE,
        "evidence_note": "Adipic acid was identified as a metabolite and the paper places it downstream of catechol before further metabolism.",
    },
    {
        "pesticide": "Quintozene",
        "pathway_name": "Pentachloronitrobenzene biodegradation by Pseudomonas putida QTH3",
        "microorganism": "Pseudomonas putida QTH3",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Pentachloronitrobenzene (PCNB)",
        "product": "2,3,4,5,6-pentachloroaniline",
        "reaction_label": "Nitro reduction",
        "enzyme": "Intracellular PCNB transformation activity; enzyme not genetically assigned",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": QUINTOZENE_QTH3_DOI,
        "reference_title": QUINTOZENE_QTH3_REFERENCE,
        "source_pdf": QUINTOZENE_QTH3_SOURCE,
        "evidence_note": "The QTH3 study identified pentachloroaniline among PCNB metabolites and reported stronger intracellular than extracellular degradation activity.",
    },
    {
        "pesticide": "Quintozene",
        "pathway_name": "Pentachloronitrobenzene biodegradation by Pseudomonas putida QTH3",
        "microorganism": "Pseudomonas putida QTH3",
        "completeness": "PARTIAL",
        "step_order": 2,
        "substrate": "Pentachloronitrobenzene (PCNB)",
        "product": "Pentachlorothioanisole",
        "reaction_label": "Thioanisole formation",
        "enzyme": "Intracellular PCNB transformation activity; enzyme not genetically assigned",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": QUINTOZENE_QTH3_DOI,
        "reference_title": QUINTOZENE_QTH3_REFERENCE,
        "source_pdf": QUINTOZENE_QTH3_SOURCE,
        "evidence_note": "Pentachlorothioanisole was identified among PCNB metabolites in the QTH3 degradation study.",
    },
    {
        "pesticide": "Quintozene",
        "pathway_name": "Pentachloronitrobenzene biodegradation by Pseudomonas putida QTH3",
        "microorganism": "Pseudomonas putida QTH3",
        "completeness": "PARTIAL",
        "step_order": 3,
        "substrate": "2,3,4,5,6-pentachloroaniline",
        "product": "2,3,5,6-tetrachloroaniline",
        "reaction_label": "Dechlorination",
        "enzyme": "Intracellular dechlorination activity; enzyme not genetically assigned",
        "gene": "",
        "evidence_type": "PROPOSED",
        "doi": QUINTOZENE_QTH3_DOI,
        "reference_title": QUINTOZENE_QTH3_REFERENCE,
        "source_pdf": QUINTOZENE_QTH3_SOURCE,
        "evidence_note": "Tetrachloroaniline metabolites were identified and placed in the proposed QTH3 PCNB degradation pathway.",
    },
    {
        "pesticide": "Quintozene",
        "pathway_name": "Pentachloronitrobenzene biodegradation by Pseudomonas putida QTH3",
        "microorganism": "Pseudomonas putida QTH3",
        "completeness": "PARTIAL",
        "step_order": 4,
        "substrate": "2,3,4,5,6-pentachloroaniline",
        "product": "2,3,4,5-tetrachloroaniline",
        "reaction_label": "Dechlorination",
        "enzyme": "Intracellular dechlorination activity; enzyme not genetically assigned",
        "gene": "",
        "evidence_type": "PROPOSED",
        "doi": QUINTOZENE_QTH3_DOI,
        "reference_title": QUINTOZENE_QTH3_REFERENCE,
        "source_pdf": QUINTOZENE_QTH3_SOURCE,
        "evidence_note": "Tetrachloroaniline metabolites were identified and placed in the proposed QTH3 PCNB degradation pathway.",
    },
    {
        "pesticide": "Quintozene",
        "pathway_name": "Pentachloronitrobenzene biodegradation by Pseudomonas putida QTH3",
        "microorganism": "Pseudomonas putida QTH3",
        "completeness": "PARTIAL",
        "step_order": 5,
        "substrate": "2,3,5,6-tetrachloroaniline",
        "product": "Catechol",
        "reaction_label": "Aromatic dechlorination / hydroxylation",
        "enzyme": "Intracellular aromatic transformation activity; enzyme not genetically assigned",
        "gene": "",
        "evidence_type": "PROPOSED",
        "doi": QUINTOZENE_QTH3_DOI,
        "reference_title": QUINTOZENE_QTH3_REFERENCE,
        "source_pdf": QUINTOZENE_QTH3_SOURCE,
        "evidence_note": "Catechol was identified among QTH3 metabolites; this connection is retained as a proposed pathway step pending gene/enzyme assignment.",
    },
]


SCREENING_DECISIONS = [
    {
        "pesticide": "Acibenzolar-S-methyl",
        "decision": "Not integrated in this batch",
        "reason": "Readable local papers reviewed in this pass did not provide a clear microbial substrate-to-product pathway arrow suitable for the database sketch.",
    },
    {
        "pesticide": "Fenamiphos",
        "decision": "Not integrated in this batch",
        "reason": "No high-confidence readable local paper was available in this pass with named microbial transformation products and pathway order.",
    },
    {
        "pesticide": "Fluxapyroxad",
        "decision": "Not integrated in this batch",
        "reason": "Readable local papers mainly reported residue/dissipation or broad degradation without enough named microbial product-pathway evidence for import.",
    },
    {
        "pesticide": "Imazamox",
        "decision": "Not integrated in this batch",
        "reason": "Available by-product evidence found in this pass appeared mainly abiotic/photochemical, so it was not imported into the microbial pathway layer.",
    },
    {
        "pesticide": "Isopyrazam",
        "decision": "Not integrated in this batch",
        "reason": "Readable local papers did not yield a clear experimentally supported microbial substrate-to-product sequence.",
    },
    {
        "pesticide": "Penthiopyrad",
        "decision": "Not integrated in this batch",
        "reason": "Potential degradation papers need deeper review before importing pathway products; no conservative arrows were added in this batch.",
    },
    {
        "pesticide": "Pinoxaden",
        "decision": "Not integrated in this batch",
        "reason": "Previous screening found degradation evidence and predicted genes, but no named transformation products suitable for a stepwise pathway arrow.",
    },
    {
        "pesticide": "Propamocarb",
        "decision": "Not integrated in this batch",
        "reason": "Previous screening did not identify a clear microbial product pathway for import.",
    },
    {
        "pesticide": "Propargite",
        "decision": "Not integrated in this batch",
        "reason": "Previous screening found a study that explicitly did not characterize metabolites, so no pathway arrow was imported.",
    },
    {
        "pesticide": "Spirodiclofen",
        "decision": "Not integrated in this batch",
        "reason": "Readable local evidence was not yet specific enough for microbial product-pathway import.",
    },
    {
        "pesticide": "Kresoxim-Methyl",
        "decision": "Not integrated in this batch",
        "reason": "A strong StrH strobilurin paper exists locally, but Kresoxim-Methyl is not currently present as a Pesticide record in the database.",
    },
]


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 64)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(ROWS)
    decisions = pd.DataFrame(SCREENING_DECISIONS)
    df.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch8_final.csv", index=False)
    decisions.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch8_screening_decisions.csv", index=False)
    df.to_csv(MASTER_CSV, index=False)
    with pd.ExcelWriter(OUT_DIR / "pesticidedb_stepwise_pathway_batch8_final.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 8", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 8", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    print(f"rows={len(df)}")
    print(f"pesticides={df['pesticide'].nunique()}")


if __name__ == "__main__":
    main()
