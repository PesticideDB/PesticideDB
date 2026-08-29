from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "stepwise_pathway_batch13_20260708"
MASTER_XLSX = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch13.xlsx"
MASTER_CSV = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch13.csv"


ROWS = [
    {
        "pesticide": "Carbendazim",
        "pathway_name": "MheI-mediated carbendazim carbamate hydrolysis",
        "microorganism": "Mycobacterium sp. SD-4; Nocardioides sp. strain SG-4G",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Carbendazim",
        "product": "2-aminobenzimidazole",
        "reaction_label": "Carbamate hydrolysis",
        "enzyme": "Carbendazim hydrolase / serine hydrolase",
        "gene": "mheI",
        "evidence_type": "GENETIC",
        "doi": "10.1016/j.jhazmat.2017.02.007",
        "reference_title": "Carbendazim hydrolase evidence curated in PesticideDB protein records",
        "source_pdf": "Database curated protein/evidence records; source PDF not locally readable",
        "evidence_note": "PesticideDB protein records link carbendazim degradation to MheI/carbendazim hydrolase; this imports the conservative primary carbamate hydrolysis product.",
    },
    {
        "pesticide": "Carbaryl",
        "pathway_name": "Carbaryl hydrolase-mediated initial carbaryl hydrolysis",
        "microorganism": "Pseudomonas sp. C5pp; Pseudomonas sp. XWY-1; Rhizobium sp. X9",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Carbaryl",
        "product": "1-naphthol",
        "reaction_label": "Carbamate hydrolysis",
        "enzyme": "Carbaryl hydrolase",
        "gene": "mcbA / cehA",
        "evidence_type": "GENETIC",
        "doi": "10.1128/AEM.01866-18",
        "reference_title": "Carbaryl hydrolase and mcb/ceh gene evidence curated in PesticideDB records",
        "source_pdf": "Database curated evidence records; source PDF not locally readable",
        "evidence_note": "Database records link carbaryl degradation to carbaryl hydrolase genes including mcbA and cehA; this imports the conservative initial hydrolysis product 1-naphthol.",
    },
    {
        "pesticide": "Chlorothalonil",
        "pathway_name": "Chd hydrolytic dehalogenase chlorothalonil transformation",
        "microorganism": "Pseudomonas sp. CTN-3; Paracoccus sp. XF-3",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Chlorothalonil",
        "product": "4-hydroxy-2,5,6-trichloroisophthalonitrile",
        "reaction_label": "Hydrolytic dehalogenation",
        "enzyme": "Hydrolytic dehalogenase Chd",
        "gene": "chd",
        "evidence_type": "GENETIC",
        "doi": "10.1128/JB.01547-09",
        "reference_title": "Hydrolytic dehalogenase Chd evidence curated in PesticideDB protein records",
        "source_pdf": "Database curated protein/evidence records; source PDF not locally readable",
        "evidence_note": "PesticideDB protein records link chlorothalonil degradation to Chd hydrolytic dehalogenase; this imports the conservative primary hydrolytic dehalogenation product.",
    },
    {
        "pesticide": "Dichlorvos",
        "pathway_name": "Phosphoesterase-mediated dichlorvos hydrolysis",
        "microorganism": "Experimentally reported phosphoesterase system; organism not specified in local protein record",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Dichlorvos",
        "product": "Dimethyl phosphate + Dichloroacetaldehyde",
        "reaction_label": "Phosphoester hydrolysis",
        "enzyme": "Phosphoesterase",
        "gene": "opdA / mpd",
        "evidence_type": "GENETIC",
        "doi": "10.1016/j.chemosphere.2019.01.058",
        "reference_title": "Dichlorvos phosphoesterase opdA/mpd evidence curated in PesticideDB protein records",
        "source_pdf": "Database curated protein/evidence records; source PDF not locally readable",
        "evidence_note": "PesticideDB protein records link dichlorvos degradation to phosphoesterase genes opdA and mpd; this imports the conservative phosphoester hydrolysis products.",
    },
]


SCREENING_DECISIONS = [
    {
        "pesticide": "Dimethoate",
        "decision": "Not integrated in this batch",
        "reason": "Database protein evidence supports DmhA/dimethoate amidohydrolase, but exact product naming should be confirmed from the blocked source paper before adding a pathway arrow.",
    },
    {
        "pesticide": "Fipronil",
        "decision": "Not integrated in this batch",
        "reason": "Database records report degradation organisms and enzyme classes, but product direction differs among oxidation, reduction, hydrolysis, and photolytic routes; source-backed product arrows are needed.",
    },
    {
        "pesticide": "Endosulfan",
        "decision": "Not integrated in this batch",
        "reason": "Current accessible evidence supports degradation but does not provide a clean directed product step from the screened source.",
    },
    {
        "pesticide": "Paraquat",
        "decision": "Not integrated in this batch",
        "reason": "Current accessible evidence reports anaerobic transformation and insoluble products, but not a named pathway compound suitable for a reaction arrow.",
    },
]


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 64)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(ROWS)
    decisions = pd.DataFrame(SCREENING_DECISIONS)
    df.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch13_final.csv", index=False)
    decisions.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch13_screening_decisions.csv", index=False)
    df.to_csv(MASTER_CSV, index=False)
    with pd.ExcelWriter(OUT_DIR / "pesticidedb_stepwise_pathway_batch13_final.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 13", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 13", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    print(f"rows={len(df)}")
    print(f"pesticides={df['pesticide'].nunique()}")


if __name__ == "__main__":
    main()
