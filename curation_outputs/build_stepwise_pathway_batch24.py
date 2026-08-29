from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "stepwise_pathway_batch24_20260712"
MASTER_XLSX = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch24.xlsx"
MASTER_CSV = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch24.csv"


def row(
    pesticide,
    pathway_name,
    microorganism,
    step_order,
    substrate,
    product,
    reaction_label,
    enzyme,
    gene,
    evidence_type,
    doi,
    reference_title,
    note,
    completeness="PARTIAL",
):
    return {
        "pesticide": pesticide,
        "pathway_name": pathway_name,
        "microorganism": microorganism,
        "completeness": completeness,
        "step_order": step_order,
        "substrate": substrate,
        "product": product,
        "reaction_label": reaction_label,
        "enzyme": enzyme,
        "gene": gene,
        "evidence_type": evidence_type,
        "doi": doi,
        "reference_title": reference_title,
        "source_pdf": "Database reference from pesticide_data.xlsx / PBDB_master_with_ids.xlsx",
        "evidence_note": note,
    }


UNRESOLVED_PRODUCT = "Microbial transformation products not resolved in current database source"


ROWS = [
    row(
        "Etoxazole",
        "Paracoccus versutus Y4 etoxazole transformation evidence",
        "Paracoccus versutus Y4",
        1,
        "Etoxazole",
        "Etoxazole monooxygenase/dehydrogenase/hydrolase transformation products",
        "Monooxygenase, dehydrogenase, and hydrolase-associated transformation",
        "Monooxygenase; dehydrogenase; hydrolase",
        "",
        "WHOLE_CELL",
        "10.1016/j.jhazmat.2025.138448",
        "Database DOI record for etoxazole degradation by Paracoccus versutus Y4",
        "This DOI is present in pesticide_data.xlsx. The database supports enzyme-class evidence for etoxazole degradation; exact metabolites are not standardized in the current source table.",
    ),
    row(
        "Fenitrothion",
        "Burkholderia sp. FDS-1 fenitrothion hydrolase pathway",
        "Burkholderia sp. FDS-1",
        1,
        "Fenitrothion",
        "Fenitrothion organophosphorus hydrolysis products",
        "Organophosphorus hydrolase-mediated hydrolysis",
        "Organophosphorus hydrolase",
        "mpd",
        "GENETIC",
        "10.1007/s10532-005-7130-2",
        "Database DOI record for mpd-associated fenitrothion degradation",
        "This DOI is present in pesticide_data.xlsx. The database supports mpd hydrolase-associated degradation; exact product names are kept conservative until metabolites are standardized in the source data.",
    ),
    row(
        "Folpet",
        "Bacillus subtilis C5 folpet esterase evidence",
        "Bacillus subtilis C5",
        1,
        "Folpet",
        "Folpet esterase hydrolysis products",
        "Esterase-associated hydrolysis",
        "Esterase B1",
        "",
        "WHOLE_CELL",
        "10.1155/2014/863094",
        "Database DOI record for folpet degradation by Bacillus subtilis C5",
        "This DOI is present in pesticide_data.xlsx. The database supports esterase/hydrolase evidence for folpet degradation; exact product names are not standardized in the current source table.",
    ),
    row(
        "Imazethapyr",
        "Pseudomonas sp. IM-4 imazethapyr degradation evidence",
        "Pseudomonas sp. IM-4",
        1,
        "Imazethapyr",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.1007/s00284-009-9442-7",
        "Database DOI record for imazethapyr degradation by Pseudomonas sp. IM-4",
        "This DOI is present in pesticide_data.xlsx. The database supports microbial degradation; exact transformation products are not stored in the current source table.",
    ),
    row(
        "Prothioconazole",
        "Candida and bacterial prothioconazole degradation evidence",
        "Candida tropicalis strain W123; Enterobacter cloacae strain Y2625; Pseudomonas aeruginosa strain W313",
        1,
        "Prothioconazole",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.1016/j.ecoenv.2020.111203",
        "Database DOI record for prothioconazole degradation by microbial strains",
        "This DOI is present in pesticide_data.xlsx. The database supports microbial degradation; exact products are not standardized in the current source table.",
    ),
    row(
        "Cadusafos",
        "Cadusafos-degrading isolate CadI/CadII evidence",
        "Isolate CadI; Isolate CadII",
        1,
        "Cadusafos",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.1016/j.femsec.2005.01.012",
        "Database DOI record for cadusafos-degrading isolates",
        "This DOI is present in pesticide_data.xlsx. The database supports microbial cadusafos degradation; exact transformation products are not stored in the current source table.",
    ),
    row(
        "Chlorantraniliprole",
        "Pseudomonas sp. GW13 chlorantraniliprole amidase evidence",
        "Pseudomonas sp. GW13",
        1,
        "Chlorantraniliprole",
        "Chlorantraniliprole amidase transformation products",
        "Amidase-associated hydrolysis",
        "Amidase",
        "",
        "WHOLE_CELL",
        "10.3390/bioengineering6040106",
        "Database DOI record for chlorantraniliprole degradation by Pseudomonas sp. GW13",
        "This DOI is present in pesticide_data.xlsx. The database supports amidase-associated degradation; exact product labels are conservative until metabolites are standardized in the source table.",
    ),
    row(
        "Dicofol",
        "Trichoderma longbrachiatum dicofol cellulase-associated evidence",
        "Trichoderma longbrachiatum",
        1,
        "Dicofol",
        "Dicofol enzymatic transformation products",
        "Cellulase-associated enzymatic transformation",
        "Cellulase",
        "",
        "PURIFIED_ENZYME",
        "10.1016/j.jes.2014.12.023",
        "Database DOI record for enzymatic dicofol degradation by Trichoderma longbrachiatum",
        "This DOI is present in pesticide_data.xlsx. The database supports enzymatic dicofol degradation; exact metabolites are not standardized in the current source table.",
    ),
    row(
        "Ethion",
        "Azospirillum and Pseudomonas ethion degradation evidence",
        "Azospirillum species; Pseudomonas species",
        1,
        "Ethion",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.1016/j.femsle.2004.09.010",
        "Database DOI record for ethion degradation by Azospirillum and Pseudomonas species",
        "This DOI is present in pesticide_data.xlsx. The database supports microbial ethion degradation; exact products are not stored in the current source table.",
    ),
    row(
        "Azinphos-Methyl",
        "Pseudomonas fluorescens DSM 1976 azinphos-methyl degradation evidence",
        "Pseudomonas fluorescens DSM 1976",
        1,
        "Azinphos-Methyl",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.1021/jf00121a026",
        "Database DOI record for azinphos-methyl degradation by Pseudomonas fluorescens DSM 1976",
        "This DOI is present in pesticide_data.xlsx. The database supports microbial degradation; exact transformation products are not standardized in the current source table.",
    ),
]


SCREENING_DECISIONS = [
    {
        "pesticide": pesticide,
        "decision": "Integrated with conservative product label",
        "reason": "Reference DOI is present in pesticide_data.xlsx and/or PBDB_master_with_ids.xlsx. Product names were kept conservative where exact metabolites are not standardized in the current source data.",
    }
    for pesticide in sorted({r["pesticide"] for r in ROWS})
]


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 72)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(ROWS)
    decisions = pd.DataFrame(SCREENING_DECISIONS)
    df.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch24_final.csv", index=False)
    decisions.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch24_screening_decisions.csv", index=False)
    df.to_csv(MASTER_CSV, index=False)
    with pd.ExcelWriter(OUT_DIR / "pesticidedb_stepwise_pathway_batch24_final.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 24", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 24", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    print(f"rows={len(df)}")
    print(f"pesticides={df['pesticide'].nunique()}")


if __name__ == "__main__":
    main()
