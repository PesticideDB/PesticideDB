from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
BATCHES = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27]
MASTER_CSV = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master.csv"
MASTER_XLSX = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master.xlsx"


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 64)


def main() -> None:
    frames = []
    for batch in BATCHES:
        path = PROJECT_ROOT / f"PesticideDB_Stepwise_Pathway_Master_Batch{batch}.csv"
        if path.exists():
            df = pd.read_csv(path)
            df.insert(0, "batch", batch)
            frames.append(df)
    combined = pd.concat(frames, ignore_index=True)
    combined.to_csv(MASTER_CSV, index=False)
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        combined.to_excel(writer, sheet_name="Stepwise Pathway Master", index=False)
        style_workbook(writer)
    print(f"rows={len(combined)}")
    print(f"pesticides={combined['pesticide'].nunique()}")


if __name__ == "__main__":
    main()
