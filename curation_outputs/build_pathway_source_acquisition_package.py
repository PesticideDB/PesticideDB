from __future__ import annotations

from collections import Counter
from pathlib import Path

import os
import sys
import django
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "pathway_source_acquisition_package_20260708"
QUEUE = PROJECT_ROOT / "curation_outputs" / "pathway_next_acquisition_queue_20260708" / "pesticidedb_pathway_next_acquisition_queue.csv"
MISSING_CURRENT = PROJECT_ROOT / "PesticideDB_Missing_Stepwise_Pathway_Information.csv"


os.environ.setdefault("DJANGO_SETTINGS_MODULE", "PepDatabase.settings")
sys.path.insert(0, str(PROJECT_ROOT))
django.setup()

from base.models import Pesticide, ProteinRecord  # noqa: E402


HIGH_VALUE = {
    "DDT": 1,
    "Acetamiprid": 1,
    "Glyphosate": 1,
    "Paraquat": 1,
    "Imidacloprid": 1,
    "Chlorpyrifos-methyl": 1,
    "Carbaryl": 1,
    "Fipronil": 1,
    "Malathion": 1,
    "Dichlorvos": 1,
    "Carbendazim": 1,
    "Endosulfan": 1,
    "Dimethoate": 1,
    "Chlorothalonil": 1,
    "Aldrin and Dieldrin": 2,
    "Cypermethrin": 2,
    "Dithiocarbamate": 2,
    "Carbosulfan": 3,
    "Propoxur": 3,
}


def compact(values: list[str], limit: int = 8) -> str:
    cleaned = []
    for value in values:
        text = str(value or "").strip()
        if not text or text.lower() == "nan":
            continue
        if text not in cleaned:
            cleaned.append(text)
    suffix = "" if len(cleaned) <= limit else f" | +{len(cleaned) - limit} more"
    return " | ".join(cleaned[:limit]) + suffix


def doi_like(value: str) -> str:
    text = str(value or "").strip()
    if text.lower().startswith("no doi"):
        return ""
    if text.startswith("10."):
        return text
    return ""


def database_leads(pesticide: str) -> dict[str, str | int]:
    records = list(
        Pesticide.objects
        .filter(pesticide__iexact=pesticide)
        .values("microorganism", "gene", "enzyme", "reference", "doi", "publication_year")
    )
    proteins = list(
        ProteinRecord.objects
        .filter(pesticide__iexact=pesticide)
        .values("pesticidedb_protein_id", "reported_protein_name", "gene_name", "microorganism", "doi")
    )
    doi_values = []
    for row in records:
        doi_values.append(doi_like(row.get("doi")))
        doi_values.append(doi_like(row.get("reference")))
    for row in proteins:
        doi_values.append(doi_like(row.get("doi")))

    genes = [row.get("gene") for row in records] + [row.get("gene_name") for row in proteins]
    enzymes = [row.get("enzyme") for row in records] + [row.get("reported_protein_name") for row in proteins]
    microbes = [row.get("microorganism") for row in records] + [row.get("microorganism") for row in proteins]
    years = [row.get("publication_year") for row in records if row.get("publication_year")]

    return {
        "database_record_count": len(records),
        "protein_record_count": len(proteins),
        "database_doi_leads": compact(doi_values, 10),
        "database_gene_leads": compact(genes, 10),
        "database_enzyme_leads": compact(enzymes, 10),
        "database_microbe_leads": compact(microbes, 8),
        "year_range": f"{min(years)}-{max(years)}" if years else "",
    }


def rank(row: pd.Series) -> int:
    pesticide = row["pesticide"]
    if pesticide in HIGH_VALUE:
        return HIGH_VALUE[pesticide]
    if int(row["protein_record_count"] or 0) >= 3:
        return 2
    if int(row["database_record_count"] or 0) >= 10:
        return 2
    if row["priority_group"] == "Placeholder files - redownload PDFs by DOI/title":
        return 3
    return 4


def action(row: pd.Series) -> str:
    if row["priority_group"] == "Placeholder files - redownload PDFs by DOI/title":
        return "Redownload real PDFs for DOI/title leads, then rerun readability audit and pathway screening."
    if row["priority_group"] == "Screened readable files - needs different product paper":
        return "Search for a different paper that reports named microbial products/intermediates; current readable files are not enough for pathway arrows."
    if row["database_doi_leads"]:
        return "Use database DOI leads first; obtain full text and extract substrate, product/intermediate, reaction, enzyme/gene, organism, DOI, and evidence wording."
    return "Run fresh literature search for experimentally reported microbial degradation products before pathway import."


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 76)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    queue = pd.read_csv(QUEUE).fillna("")
    if MISSING_CURRENT.exists():
        missing = pd.read_csv(MISSING_CURRENT).fillna("")
        missing_pesticides = {str(value).strip() for value in missing["pesticide"] if str(value).strip()}
        queue = queue[queue["pesticide"].isin(missing_pesticides)].copy()
    lead_rows = []
    for _, row in queue.iterrows():
        leads = database_leads(row["pesticide"])
        lead_rows.append({**row.to_dict(), **leads})
    enriched = pd.DataFrame(lead_rows)
    if enriched.empty:
        enriched = pd.DataFrame(
            columns=[
                "priority_group",
                "acquisition_priority",
                "pesticide",
                "database_record_count",
                "protein_record_count",
                "database_doi_leads",
                "database_gene_leads",
                "database_enzyme_leads",
                "database_microbe_leads",
                "year_range",
                "recommended_source_action",
            ]
        )
    else:
        enriched.insert(1, "acquisition_priority", enriched.apply(rank, axis=1))
        enriched["recommended_source_action"] = enriched.apply(action, axis=1)
        enriched = enriched.sort_values(
            [
                "acquisition_priority",
                "protein_record_count",
                "database_record_count",
                "html_placeholder_count",
                "pesticide",
            ],
            ascending=[True, False, False, False, True],
        )

    if enriched.empty:
        summary = pd.DataFrame(columns=["acquisition_priority", "priority_group", "pesticide_count"])
    else:
        summary = (
            enriched.groupby(["acquisition_priority", "priority_group"], dropna=False)
            .size()
            .reset_index(name="pesticide_count")
            .sort_values(["acquisition_priority", "priority_group"])
        )
    doi_counter = Counter()
    for _, row in enriched.iterrows():
        for doi in str(row["database_doi_leads"]).split(" | "):
            doi = doi.strip()
            if doi:
                doi_counter[doi] += 1
    doi_rows = pd.DataFrame(
        [{"doi": doi, "pesticide_count": count} for doi, count in doi_counter.most_common()]
    )

    csv_path = OUT_DIR / "pesticidedb_pathway_source_acquisition_priority.csv"
    xlsx_path = OUT_DIR / "pesticidedb_pathway_source_acquisition_priority.xlsx"
    enriched.to_csv(csv_path, index=False)
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        enriched.to_excel(writer, sheet_name="Priority Queue", index=False)
        doi_rows.to_excel(writer, sheet_name="Database DOI Leads", index=False)
        style_workbook(writer)
    print(csv_path)
    print(f"rows={len(enriched)}")
    print(summary.to_string(index=False))


if __name__ == "__main__":
    main()
