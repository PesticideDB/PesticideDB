from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "stepwise_pathway_batch14_20260709"
MASTER_XLSX = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch14.xlsx"
MASTER_CSV = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch14.csv"


ROWS = [
    {
        "pesticide": "Flonicamid",
        "pathway_name": "Ensifer adhaerens flonicamid nitrile hydration",
        "microorganism": "Ensifer adhaerens CGMCC 6315",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Flonicamid",
        "product": "N-(4-trifluoromethylnicotinoyl) glycinamide",
        "reaction_label": "Nitrile hydration",
        "enzyme": "Nitrile hydratase PnhA / CnhA",
        "gene": "pnhA / cnhA",
        "evidence_type": "PURIFIED_ENZYME",
        "doi": "10.1186/s12934-021-01620-4",
        "reference_title": "Biodegradation of flonicamid by Ensifer adhaerens CGMCC 6315 and enzymatic characterization of the nitrile hydratases involved",
        "source_pdf": "Flonicamid_10.1186_s12934-021-01620-4.pdf",
        "evidence_note": "The paper reports transformation of flonicamid to TFNG-AM via a hydration pathway mediated by nitrile hydratases PnhA and CnhA.",
    },
    {
        "pesticide": "Dicamba",
        "pathway_name": "Dmt06 tetrahydrofolate-dependent dicamba demethylation",
        "microorganism": "Dicamba-degrading microbial consortium; recombinant Escherichia coli BL21 expression",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Dicamba",
        "product": "3,6-dichlorosalicylate",
        "reaction_label": "Tetrahydrofolate-dependent demethylation",
        "enzyme": "Dicamba demethylase Dmt06",
        "gene": "dmt06",
        "evidence_type": "PURIFIED_ENZYME",
        "doi": "10.3389/fmicb.2022.978577",
        "reference_title": "Cloning of a novel tetrahydrofolate-dependent dicamba demethylase gene from dicamba-degrading consortium and characterization of the gene product",
        "source_pdf": "Dicamba_10.3389_fmicb.2022.978577.pdf",
        "evidence_note": "Purified Dmt06 catalyzed methyl transfer from dicamba to THF, generating 3,6-DCSA and 5-methyl-THF.",
    },
    {
        "pesticide": "Iprodione",
        "pathway_name": "Azospirillum sp. A1-3 iprodione degradation",
        "microorganism": "Azospirillum sp. A1-3",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Iprodione",
        "product": "N-(3,5-dichlorophenyl)-2,4-dioxoimidazolidine",
        "reaction_label": "N1 amide bond hydrolysis",
        "enzyme": "Iprodione hydrolase",
        "gene": "ipaH",
        "evidence_type": "GENETIC",
        "doi": "10.3389/fmicb.2022.1057030",
        "reference_title": "Degradation of iprodione by a novel strain Azospirillum sp. A1-3 isolated from Tibet",
        "source_pdf": "Iprodione_10.3389_fmicb.2022.1057030.pdf",
        "evidence_note": "GC-MS/MS identified N-(3,5-dichlorophenyl)-2,4-dioxoimidazolidine as the first iprodione degradation product; ipaH was detected in strain A1-3.",
    },
    {
        "pesticide": "Iprodione",
        "pathway_name": "Azospirillum sp. A1-3 iprodione degradation",
        "microorganism": "Azospirillum sp. A1-3",
        "completeness": "PARTIAL",
        "step_order": 2,
        "substrate": "N-(3,5-dichlorophenyl)-2,4-dioxoimidazolidine",
        "product": "(3,5-dichlorophenylurea) acetic acid",
        "reaction_label": "Hydantoin ring cleavage",
        "enzyme": "Hydantoin ring-cleavage activity",
        "gene": "ddaH-like activity",
        "evidence_type": "METABOLITE",
        "doi": "10.3389/fmicb.2022.1057030",
        "reference_title": "Degradation of iprodione by a novel strain Azospirillum sp. A1-3 isolated from Tibet",
        "source_pdf": "Iprodione_10.3389_fmicb.2022.1057030.pdf",
        "evidence_note": "The study reports conversion of N-(3,5-dichlorophenyl)-2,4-dioxoimidazolidine to (3,5-dichlorophenylurea) acetic acid; downstream 3,5-dichloroaniline was not formed by this strain.",
    },
    {
        "pesticide": "Cyromazine",
        "pathway_name": "Melamine-degrading bacteria cyromazine transformation",
        "microorganism": "Arthrobacter sp. MCO; Arthrobacter sp. CSP; Nocardioides sp. ATD6",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Cyromazine",
        "product": "N-cyclopropylammeline",
        "reaction_label": "Microbial deamination / triazine transformation",
        "enzyme": "Enzyme not identified",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1584/jpestics.D15-044",
        "reference_title": "Biodegradation of cyromazine by melamine-degrading bacteria",
        "source_pdf": "Cyromazine_10.1584_jpestics.D15-044.pdf",
        "evidence_note": "UPLC-MS/MS identified N-cyclopropylammeline as the cyromazine metabolite during degradation by melamine-degrading bacteria.",
    },
    {
        "pesticide": "Oxamyl",
        "pathway_name": "Pseudomonas oxamyl carbamate hydrolysis",
        "microorganism": "Pseudomonas spp. oxamyl-degrading isolates",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Oxamyl",
        "product": "Oxamyl oxime + Methylamine",
        "reaction_label": "Carbamate hydrolysis",
        "enzyme": "CehA-like carbamate hydrolase",
        "gene": "cehA-like gene",
        "evidence_type": "GENETIC",
        "doi": "10.3389/fmicb.2016.00616",
        "reference_title": "Isolation of oxamyl-degrading Pseudomonas strains carrying a CehA-like carbamate-hydrolase gene",
        "source_pdf": "Oxamyl_10.3389_fmicb.2016.00616.pdf",
        "evidence_note": "The isolates hydrolyzed oxamyl to oxamyl oxime, which was not further transformed, and utilized methylamine as carbon and nitrogen source; cehA-like transcription supported involvement in hydrolysis.",
    },
    {
        "pesticide": "Lindane",
        "pathway_name": "Sequential anaerobic lindane dechlorination by mixed microbial cultures",
        "microorganism": "Sequential mixed anaerobic microbial cultures I, II, and III",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Lindane",
        "product": "Monochlorobenzene + Benzene",
        "reaction_label": "Reductive dechlorination",
        "enzyme": "Dehalogenase activity not assigned",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1101/2020.10.25.354597",
        "reference_title": "Biotransformation of lindane to non-toxic end products by sequential treatment with mixed anaerobic microbial cultures",
        "source_pdf": "Lindane_10.1101_2020.10.25.354597.pdf",
        "evidence_note": "The open preprint reports dechlorination of lindane to monochlorobenzene and benzene by culture I in a sequential anaerobic treatment approach.",
    },
]


SCREENING_DECISIONS = [
    {
        "pesticide": "Bentazone",
        "decision": "Not integrated in this batch",
        "reason": "The downloaded paper identifies many transformation products and mechanisms, but a compact, named stepwise route needs manual table extraction before import.",
    },
    {
        "pesticide": "Fenvalerate / Permethrin / Deltamethrin / Bifenthrin",
        "decision": "Not integrated in this batch",
        "reason": "Downloaded pyrethroid papers contain useful degradation/pathway evidence, but the exact pesticide-specific first products need manual separation to avoid merging similar pyrethroid routes.",
    },
    {
        "pesticide": "Paraquat / Endosulfan",
        "decision": "Not integrated in this batch",
        "reason": "The accessible PDFs were already screened; they support transformation/degradation but do not provide a clean named reaction arrow beyond prior decisions.",
    },
]


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 64)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(ROWS)
    decisions = pd.DataFrame(SCREENING_DECISIONS)
    df.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch14_final.csv", index=False)
    decisions.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch14_screening_decisions.csv", index=False)
    df.to_csv(MASTER_CSV, index=False)
    with pd.ExcelWriter(OUT_DIR / "pesticidedb_stepwise_pathway_batch14_final.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 14", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 14", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    print(f"rows={len(df)}")
    print(f"pesticides={df['pesticide'].nunique()}")


if __name__ == "__main__":
    main()
