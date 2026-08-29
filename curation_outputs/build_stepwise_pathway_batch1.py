from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "stepwise_pathway_batch1_20260707"
MASTER_XLSX = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch1.xlsx"
MASTER_CSV = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch1.csv"


ROWS = [
    {
        "pesticide": "Pymetrozine",
        "pathway_name": "Pymetrozine degradation by Pseudomonas sp. BYT-1",
        "microorganism": "Pseudomonas sp. BYT-1",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Pymetrozine",
        "product": "4-amino-6-methyl-4,5-dihydro-2H-[1,2,4]triazin-3-one (AMDT)",
        "reaction_label": "Oxidative hydrolysis of C=N double bond",
        "enzyme": "Cell-free extract activity; enzyme not assigned in this paper",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1021/acs.jafc.8b06155",
        "reference_title": "Isolation and Characterization of the Pymetrozine-Degrading Strain Pseudomonas sp. BYT-1",
        "source_pdf": "Isolation and Characterization of the Pymetrozine-Degrading Strain Pseudomonas sp. BYT-1..pdf",
        "evidence_note": "HPLC/MS/MS identified AMDT and nicotinic acid; AMDT is shown as the direct cleavage product and nicotinic acid is represented through the nicotinaldehyde intermediate route.",
    },
    {
        "pesticide": "Pymetrozine",
        "pathway_name": "Pymetrozine degradation by Pseudomonas sp. BYT-1",
        "microorganism": "Pseudomonas sp. BYT-1",
        "completeness": "PARTIAL",
        "step_order": 2,
        "substrate": "Nicotinic acid",
        "product": "6-hydroxynicotinic acid",
        "reaction_label": "Hydroxylation",
        "enzyme": "Unassigned hydroxylation activity",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1021/acs.jafc.8b06155",
        "reference_title": "Isolation and Characterization of the Pymetrozine-Degrading Strain Pseudomonas sp. BYT-1",
        "source_pdf": "Isolation and Characterization of the Pymetrozine-Degrading Strain Pseudomonas sp. BYT-1..pdf",
        "evidence_note": "6-hydroxynicotinic acid was detected as an intermediate during nicotinic acid degradation.",
    },
    {
        "pesticide": "Pymetrozine",
        "pathway_name": "Pymetrozine PyzH hydrolase reaction",
        "microorganism": "Pseudomonas sp. BYT-1",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Pymetrozine",
        "product": "Nicotinaldehyde",
        "reaction_label": "C=N double-bond hydrolysis",
        "enzyme": "Pymetrozine hydrolase PyzH",
        "gene": "pyzH",
        "evidence_type": "PURIFIED_ENZYME",
        "doi": "10.1111/1462-2920.15557",
        "reference_title": "A novel hydrolase PyzH catalyses the cleavage of C=N double bond for pymetrozine degradation in Pseudomonas sp. BYT-1",
        "source_pdf": "A novel hydrolase PyzH catalyses the cleavage of C=N double bond for pymetrozine degradation in Pseudomonas sp. BYT-1..pdf",
        "evidence_note": "pyzH disruption/complementation and purified PyzH enzyme assays support the initial hydrolysis step; AMDT is the co-product of this cleavage and nicotinaldehyde is the pyridine-side intermediate.",
    },
    {
        "pesticide": "Pymetrozine",
        "pathway_name": "Pymetrozine PyzH hydrolase reaction",
        "microorganism": "Pseudomonas sp. BYT-1",
        "completeness": "PARTIAL",
        "step_order": 2,
        "substrate": "Nicotinaldehyde",
        "product": "Nicotinic acid",
        "reaction_label": "Aldehyde oxidation",
        "enzyme": "Putative dehydrogenase",
        "gene": "",
        "evidence_type": "PROPOSED",
        "doi": "10.1111/1462-2920.15557",
        "reference_title": "A novel hydrolase PyzH catalyses the cleavage of C=N double bond for pymetrozine degradation in Pseudomonas sp. BYT-1",
        "source_pdf": "A novel hydrolase PyzH catalyses the cleavage of C=N double bond for pymetrozine degradation in Pseudomonas sp. BYT-1..pdf",
        "evidence_note": "The paper explains that nicotinaldehyde can be rapidly converted to nicotinic acid in strain BYT-1; the enzyme is not assigned.",
    },
    {
        "pesticide": "Triazophos",
        "pathway_name": "Triazophos degradation by Diaphorobacter sp. TPD-1",
        "microorganism": "Diaphorobacter sp. TPD-1",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Triazophos",
        "product": "1-phenyl-3-hydroxy-1,2,4-triazole (PHT) + O,O-diethyl phosphorothioic acid",
        "reaction_label": "P-O ester bond hydrolysis",
        "enzyme": "Unassigned hydrolase activity",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1007/s00284-010-9746-7",
        "reference_title": "Identification of the Biochemical Degradation Pathway of Triazophos and its Intermediate in Diaphorobacter sp. TPD-1",
        "source_pdf": "Identification of the biochemical degradation pathway of triazophos and its intermediate in Diaphorobacter sp. TPD-1.pdf",
        "evidence_note": "The paper proposes PHT and O,O-diethyl phosphorothioic acid from P-O ester-bond hydrolysis.",
    },
    {
        "pesticide": "Triazophos",
        "pathway_name": "Triazophos degradation by Diaphorobacter sp. TPD-1",
        "microorganism": "Diaphorobacter sp. TPD-1",
        "completeness": "PARTIAL",
        "step_order": 2,
        "substrate": "1-phenyl-3-hydroxy-1,2,4-triazole (PHT)",
        "product": "(E)-1-formyl-2-phenyldiazene",
        "reaction_label": "Triazole-ring cleavage",
        "enzyme": "Unassigned activity",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1007/s00284-010-9746-7",
        "reference_title": "Identification of the Biochemical Degradation Pathway of Triazophos and its Intermediate in Diaphorobacter sp. TPD-1",
        "source_pdf": "Identification of the biochemical degradation pathway of triazophos and its intermediate in Diaphorobacter sp. TPD-1.pdf",
        "evidence_note": "MS/MS identified (E)-1-formyl-2-phenyldiazene as a PHT degradation metabolite.",
    },
    {
        "pesticide": "Triazophos",
        "pathway_name": "Triazophos degradation by Diaphorobacter sp. TPD-1",
        "microorganism": "Diaphorobacter sp. TPD-1",
        "completeness": "PARTIAL",
        "step_order": 3,
        "substrate": "(E)-1-formyl-2-phenyldiazene",
        "product": "2-phenylhydrazinecarboxylic acid",
        "reaction_label": "Hydration",
        "enzyme": "Unassigned activity",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1007/s00284-010-9746-7",
        "reference_title": "Identification of the Biochemical Degradation Pathway of Triazophos and its Intermediate in Diaphorobacter sp. TPD-1",
        "source_pdf": "Identification of the biochemical degradation pathway of triazophos and its intermediate in Diaphorobacter sp. TPD-1.pdf",
        "evidence_note": "The proposed pathway converts (E)-1-formyl-2-phenyldiazene to 2-phenylhydrazinecarboxylic acid by addition of water.",
    },
    {
        "pesticide": "Triazophos",
        "pathway_name": "Triazophos degradation by Diaphorobacter sp. TPD-1",
        "microorganism": "Diaphorobacter sp. TPD-1",
        "completeness": "PARTIAL",
        "step_order": 4,
        "substrate": "2-phenylhydrazinecarboxylic acid",
        "product": "Phenylhydrazine",
        "reaction_label": "Decarboxylation",
        "enzyme": "Unassigned activity",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1007/s00284-010-9746-7",
        "reference_title": "Identification of the Biochemical Degradation Pathway of Triazophos and its Intermediate in Diaphorobacter sp. TPD-1",
        "source_pdf": "Identification of the biochemical degradation pathway of triazophos and its intermediate in Diaphorobacter sp. TPD-1.pdf",
        "evidence_note": "MS/MS identified phenylhydrazine; the paper proposes decarboxylation of 2-phenylhydrazinecarboxylic acid.",
    },
    {
        "pesticide": "Triazophos",
        "pathway_name": "Triazophos hydrolysis by Burkholderia sp. SZL-1",
        "microorganism": "Burkholderia sp. SZL-1",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Triazophos",
        "product": "1-phenyl-3-hydroxy-1,2,4-triazole (PHT)",
        "reaction_label": "P-O bond hydrolysis",
        "enzyme": "Triazophos hydrolase TrhA",
        "gene": "trhA",
        "evidence_type": "PURIFIED_ENZYME",
        "doi": "10.1093/femsle/fnw108",
        "reference_title": "Cloning, expression and mutation of a triazophos hydrolase gene from Burkholderia sp. SZL-1",
        "source_pdf": "Cloning, expression and mutation of a triazophos hydrolase gene from Burkholderia sp. SZL-1..pdf",
        "evidence_note": "trhA was cloned, expressed, purified, and classified as a hydrolase responsible for triazophos hydrolysis to PHT.",
    },
    {
        "pesticide": "Thiamethoxam",
        "pathway_name": "Thiamethoxam transformation by Ensifer adhaerens TMX-23",
        "microorganism": "Ensifer adhaerens TMX-23",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Thiamethoxam",
        "product": "Nitrosoimino thiamethoxam",
        "reaction_label": "Nitro reduction",
        "enzyme": "Unassigned nitro-reduction activity",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1007/s00253-012-4638-3",
        "reference_title": "Biodegradation of the neonicotinoid insecticide thiamethoxam by the nitrogen-fixing and plant-growth-promoting rhizobacterium Ensifer adhaerens strain TMX-23",
        "source_pdf": "Biodegradation of the neonicotinoid insecticide thiamethoxam by the nitrogen-fixing and plant-growth-promoting rhizobacterium Ensifer adhaerens strain TMX-23..pdf",
        "evidence_note": "HPLC/MS analysis identified nitrosoimino thiamethoxam as a metabolite.",
    },
    {
        "pesticide": "Thiamethoxam",
        "pathway_name": "Thiamethoxam transformation by Ensifer adhaerens TMX-23",
        "microorganism": "Ensifer adhaerens TMX-23",
        "completeness": "PARTIAL",
        "step_order": 2,
        "substrate": "Thiamethoxam",
        "product": "Urea thiamethoxam",
        "reaction_label": "Nitroimino-group transformation",
        "enzyme": "Unassigned transformation activity",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1007/s00253-012-4638-3",
        "reference_title": "Biodegradation of the neonicotinoid insecticide thiamethoxam by the nitrogen-fixing and plant-growth-promoting rhizobacterium Ensifer adhaerens strain TMX-23",
        "source_pdf": "Biodegradation of the neonicotinoid insecticide thiamethoxam by the nitrogen-fixing and plant-growth-promoting rhizobacterium Ensifer adhaerens strain TMX-23..pdf",
        "evidence_note": "HPLC/MS analysis identified urea thiamethoxam as a metabolite; paper states nitro reduction is the major pathway.",
    },
    {
        "pesticide": "Abamectin",
        "pathway_name": "Abamectin B1a degradation by Stenotrophomonas maltophilia ZJB-14120",
        "microorganism": "Stenotrophomonas maltophilia ZJB-14120",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Abamectin B1a",
        "product": "Metabolite A (m/z 886.8; C24-hydroxylated abamectin B1a)",
        "reaction_label": "Hydroxylation of C24 methyl group",
        "enzyme": "Unassigned activity",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1016/j.resmic.2015.04.002",
        "reference_title": "Degradation of abamectin by newly isolated Stenotrophomonas maltophilia ZJB-14120 and characterization of its abamectin-tolerance mechanism",
        "source_pdf": "Degradation of abamectin by newly isolated Stenotrophomonas maltophilia ZJB-14120 and characterization of its abamectin-tolerance mechanism..pdf",
        "evidence_note": "HPLC-ESI-MS identified metabolite A and the paper proposes hydroxylation of the C24 methyl group.",
    },
    {
        "pesticide": "Abamectin",
        "pathway_name": "Abamectin B1a degradation by Stenotrophomonas maltophilia ZJB-14120",
        "microorganism": "Stenotrophomonas maltophilia ZJB-14120",
        "completeness": "PARTIAL",
        "step_order": 2,
        "substrate": "Metabolite A (m/z 886.8)",
        "product": "Metabolite B (m/z 709.0)",
        "reaction_label": "C-O and C-C bond cleavage",
        "enzyme": "Unassigned activity",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1016/j.resmic.2015.04.002",
        "reference_title": "Degradation of abamectin by newly isolated Stenotrophomonas maltophilia ZJB-14120 and characterization of its abamectin-tolerance mechanism",
        "source_pdf": "Degradation of abamectin by newly isolated Stenotrophomonas maltophilia ZJB-14120 and characterization of its abamectin-tolerance mechanism..pdf",
        "evidence_note": "The paper proposes degradation of metabolite A to metabolite B by C17-C21 C-O bond breakage and C19-C20 cleavage.",
    },
    {
        "pesticide": "Abamectin",
        "pathway_name": "Abamectin B1a degradation by Stenotrophomonas maltophilia ZJB-14120",
        "microorganism": "Stenotrophomonas maltophilia ZJB-14120",
        "completeness": "PARTIAL",
        "step_order": 3,
        "substrate": "Metabolite B (m/z 709.0)",
        "product": "Metabolite C (m/z 485.0)",
        "reaction_label": "Double-bond breakage and lactone-ring opening",
        "enzyme": "Unassigned activity",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1016/j.resmic.2015.04.002",
        "reference_title": "Degradation of abamectin by newly isolated Stenotrophomonas maltophilia ZJB-14120 and characterization of its abamectin-tolerance mechanism",
        "source_pdf": "Degradation of abamectin by newly isolated Stenotrophomonas maltophilia ZJB-14120 and characterization of its abamectin-tolerance mechanism..pdf",
        "evidence_note": "The paper proposes conversion of metabolite B to metabolite C by breaking C8-C9, cleaving C18-C19, and opening the lactone ring.",
    },
    {
        "pesticide": "Abamectin",
        "pathway_name": "Abamectin B1a degradation by Stenotrophomonas maltophilia ZJB-14120",
        "microorganism": "Stenotrophomonas maltophilia ZJB-14120",
        "completeness": "PARTIAL",
        "step_order": 4,
        "substrate": "Metabolite C (m/z 485.0)",
        "product": "Metabolite D (m/z 292.7)",
        "reaction_label": "C-O bond cleavage and demethylation",
        "enzyme": "Unassigned activity",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1016/j.resmic.2015.04.002",
        "reference_title": "Degradation of abamectin by newly isolated Stenotrophomonas maltophilia ZJB-14120 and characterization of its abamectin-tolerance mechanism",
        "source_pdf": "Degradation of abamectin by newly isolated Stenotrophomonas maltophilia ZJB-14120 and characterization of its abamectin-tolerance mechanism..pdf",
        "evidence_note": "The paper proposes conversion of metabolite C to metabolite D by C10-C13 C-O bond breakage and demethylation.",
    },
]


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 64)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(ROWS)
    df.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch1_final.csv", index=False)
    df.to_csv(MASTER_CSV, index=False)
    with pd.ExcelWriter(OUT_DIR / "pesticidedb_stepwise_pathway_batch1_final.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 1", index=False)
        style_workbook(writer)
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 1", index=False)
        style_workbook(writer)
    print(f"rows={len(df)}")
    print(f"pesticides={df['pesticide'].nunique()}")
    print(f"project_master={MASTER_XLSX}")


if __name__ == "__main__":
    main()
