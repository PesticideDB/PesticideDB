from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "stepwise_pathway_batch20_20260711"
MASTER_XLSX = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch20.xlsx"
MASTER_CSV = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch20.csv"


def row(
    pesticide,
    pathway_name,
    microorganism,
    step_order,
    substrate,
    product,
    reaction_label,
    enzyme,
    gene,
    evidence_type,
    doi,
    reference_title,
    note,
    completeness="PARTIAL",
):
    return {
        "pesticide": pesticide,
        "pathway_name": pathway_name,
        "microorganism": microorganism,
        "completeness": completeness,
        "step_order": step_order,
        "substrate": substrate,
        "product": product,
        "reaction_label": reaction_label,
        "enzyme": enzyme,
        "gene": gene,
        "evidence_type": evidence_type,
        "doi": doi,
        "reference_title": reference_title,
        "source_pdf": "Database reference from pesticide_data.xlsx / PBDB_master_with_ids.xlsx",
        "evidence_note": note,
    }


ROWS = [
    row(
        "Buprofezin",
        "Rhodococcus qingshengii YL-1 buprofezin catabolic pathway",
        "Rhodococcus qingshengii YL-1",
        1,
        "Buprofezin",
        "2-tert-butylimino-3-isopropyl-1,3,5-thiadiazinan-4-one (2-BI)",
        "Dihydroxylation, aromatic-ring cleavage, and amide-bond cleavage",
        "Rieske nonheme iron oxygenase; dehydrogenase; aromatic ring-cleavage enzyme",
        "RHO; bfzB; bfzC",
        "GENETIC",
        "10.1128/AEM.00868-17",
        "Molecular mechanism and genetic determinants of buprofezin degradation",
        "This DOI is present in pesticide_data.xlsx. The pathway links buprofezin transformation by Rhodococcus qingshengii YL-1 with the bfz gene cluster and detected upstream catabolic product 2-BI.",
    ),
    row(
        "Acetochlor",
        "Bacillus sp. ACD-9 acetochlor degradation pathway",
        "Bacillus sp. ACD-9",
        1,
        "Acetochlor",
        "2-chloro-N-(2-methyl-6-ethylphenyl) acetamide (CMEPA)",
        "Chloroacetamide transformation",
        "",
        "",
        "METABOLITE",
        "10.1007/s13205-020-2056-2",
        "Degradation of acetochlor and beneficial effect of phosphate-solubilizing Bacillus sp. ACD-9 on maize seedlings",
        "This DOI is present in pesticide_data.xlsx. The source reports LC-MS identification of CMEPA as the probable acetochlor degradation product.",
    ),
    row(
        "Diphenylamine",
        "Burkholderia and Ralstonia diphenylamine dioxygenation pathway",
        "Burkholderia sp. JS667; Ralstonia sp. JS668",
        1,
        "Diphenylamine",
        "Aniline + catechol",
        "Dioxygenation and spontaneous rearomatization",
        "Multicomponent ring-hydroxylating dioxygenase",
        "dpaAa; tdnC",
        "GENETIC",
        "10.1128/AEM.02198-08",
        "Microbial metabolism of diphenylamine",
        "This DOI is present in pesticide_data.xlsx. The source links diphenylamine conversion to aniline and catechol with ring-hydroxylating dioxygenase genes.",
    ),
    row(
        "Aldrin and Dieldrin",
        "Mucor racemosus DDF dieldrin degradation pathway",
        "Mucor racemosus strain DDF",
        1,
        "Dieldrin",
        "Aldrin trans-diol",
        "Fungal transformation",
        "",
        "",
        "METABOLITE",
        "10.1584/jpestics.J18-03",
        "Biodegradability and biodegradation pathways of chlorinated cyclodiene insecticides by soil fungi",
        "This DOI is present in pesticide_data.xlsx. The source reports dieldrin degradation by Mucor racemosus strain DDF with production of aldrin trans-diol.",
    ),
    row(
        "Aldrin and Dieldrin",
        "Mucor racemosus DDF dieldrin degradation pathway",
        "Mucor racemosus strain DDF",
        2,
        "Aldrin trans-diol",
        "Aldrin trans-diol exo- and endo-phosphates",
        "Phosphate metabolite formation",
        "",
        "",
        "METABOLITE",
        "10.1584/jpestics.J18-03",
        "Biodegradability and biodegradation pathways of chlorinated cyclodiene insecticides by soil fungi",
        "This DOI is present in pesticide_data.xlsx. The source reports further conversion of aldrin trans-diol to exo- and endo-phosphate metabolites.",
    ),
    row(
        "Parathion-Methyl",
        "Bacillus strain C5 parathion-methyl hydrolysis pathway",
        "Acinetobacter radioresistens USTB-04",
        1,
        "Parathion-Methyl",
        "p-nitrophenol + dimethyl thiophosphate",
        "Phosphotriester hydrolysis",
        "Hydrolase",
        "",
        "WHOLE_CELL",
        "10.1016/S1001-0742(07)60205-8",
        "Identification of a marine Bacillus strain C5 and parathion-methyl degradation characteristics",
        "This DOI is present in pesticide_data.xlsx. The row captures the first hydrolysis step generally reported for parathion-methyl degradation.",
    ),
    row(
        "Pendimethalin",
        "Bacillus subtilis Y3 pendimethalin nitroreduction pathway",
        "Bacillus subtilis Y3",
        1,
        "Pendimethalin",
        "Reduced pendimethalin metabolite",
        "Nitro-group reduction",
        "Pendimethalin nitroreductase",
        "pnr",
        "GENETIC",
        "10.1128/AEM.01771-16",
        "Identification of a pendimethalin nitroreductase from Bacillus subtilis Y3",
        "This DOI is present in pesticide_data.xlsx. The database reference supports the pnr nitroreductase step; the product label is kept conservative because the current database source does not store a more specific metabolite name.",
    ),
    row(
        "Dicamba",
        "Pseudomonas maltophilia DI-6 dicamba O-demethylation pathway",
        "Pseudomonas maltophilia DI-6",
        1,
        "Dicamba",
        "3,6-dichlorosalicylic acid",
        "O-demethylation",
        "Dicamba monooxygenase",
        "ddmA; ddmB; ddmC",
        "GENETIC",
        "10.1128/aem.63.4.1623-1626.1997",
        "Dicamba O-demethylase genes from Pseudomonas maltophilia DI-6",
        "This DOI is present in pesticide_data.xlsx. Dicamba O-demethylase converts dicamba to 3,6-dichlorosalicylic acid.",
    ),
    row(
        "Flonicamid",
        "Alcaligenes faecalis CGMCC 17553 flonicamid nitrile hydrolysis pathway",
        "Alcaligenes faecalis CGMCC 17553",
        1,
        "Flonicamid",
        "N-(4-trifluoromethylnicotinoyl)glycinamide (TFNG-AM)",
        "Nitrile hydration / amide formation",
        "Nitrile hydratase / amidase system",
        "NitA; NitB; NitC; NitD; NitE",
        "GENETIC",
        "10.1021/acs.jafc.9b04245",
        "Flonicamid biodegradation genes in Alcaligenes faecalis CGMCC 17553",
        "This DOI is present in pesticide_data.xlsx and PBDB_master_with_ids.xlsx. The step is represented as the first database-supported nitrile-hydrolysis transformation.",
    ),
    row(
        "Iprodione",
        "Paenarthrobacter sp. YJN-5 iprodione hydrolase pathway",
        "Paenarthrobacter sp. YJN-5",
        1,
        "Iprodione",
        "Iprodione hydantoin-ring hydrolysis product",
        "Amide/hydantoin-ring hydrolysis",
        "IpaH amidase",
        "ipaH",
        "GENETIC",
        "10.1128/AEM.01150-18",
        "Identification of iprodione catabolic amidase IpaH in Paenarthrobacter sp. YJN-5",
        "This DOI is present in pesticide_data.xlsx and PBDB_master_with_ids.xlsx. The product name is kept conservative because the current database source stores the protein/gene evidence but not a standardized metabolite name.",
    ),
    row(
        "Sulfoxaflor",
        "Aminobacter sp. CGMCC 1.17253 sulfoxaflor nitrile hydration pathway",
        "Aminobacter sp. CGMCC 1.17253",
        1,
        "Sulfoxaflor",
        "Sulfoxaflor amide",
        "Nitrile hydration",
        "Nitrile hydratase",
        "nhaseA; nhaseB",
        "GENETIC",
        "10.1021/acs.jafc.9b06668",
        "Nitrile hydratase-mediated sulfoxaflor degradation by Aminobacter sp. CGMCC 1.17253",
        "This DOI is present in pesticide_data.xlsx and PBDB_master_with_ids.xlsx. The row represents the database-backed nitrile hydratase transformation of sulfoxaflor.",
    ),
]


SCREENING_DECISIONS = [
    {"pesticide": pesticide, "decision": "Integrated", "reason": "Reference DOI is present in pesticide_data.xlsx and/or PBDB_master_with_ids.xlsx."}
    for pesticide in sorted({r["pesticide"] for r in ROWS})
] + [
    {
        "pesticide": "Difenoconazole",
        "decision": "Not integrated",
        "reason": "The DOI used in the earlier draft is not present in pesticide_data.xlsx or PBDB_master_with_ids.xlsx.",
    },
    {
        "pesticide": "Paraquat",
        "decision": "Not integrated",
        "reason": "The Lipomyces starkeyi source DOI is not present in pesticide_data.xlsx or PBDB_master_with_ids.xlsx.",
    },
    {
        "pesticide": "Imazamox",
        "decision": "Not integrated",
        "reason": "The draft row came from a PDF not represented by a DOI/reference in pesticide_data.xlsx or PBDB_master_with_ids.xlsx.",
    },
    {
        "pesticide": "Propanil",
        "decision": "Not integrated",
        "reason": "The AmiH52 propanil DOI is not present in pesticide_data.xlsx or PBDB_master_with_ids.xlsx.",
    },
]


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 72)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(ROWS)
    decisions = pd.DataFrame(SCREENING_DECISIONS)
    df.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch20_final.csv", index=False)
    decisions.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch20_screening_decisions.csv", index=False)
    df.to_csv(MASTER_CSV, index=False)
    with pd.ExcelWriter(OUT_DIR / "pesticidedb_stepwise_pathway_batch20_final.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 20", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 20", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    print(f"rows={len(df)}")
    print(f"pesticides={df['pesticide'].nunique()}")


if __name__ == "__main__":
    main()
