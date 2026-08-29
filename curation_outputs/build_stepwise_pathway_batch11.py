from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "stepwise_pathway_batch11_20260708"
MASTER_XLSX = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch11.xlsx"
MASTER_CSV = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch11.csv"


ACETAMIPRID_DOI = "10.1371/journal.pone.0082603"
DDT_DOI = "10.4014/jmb.1701.01073"
GLYPHOSATE_DOI = "10.2323/jgam.58.263"
IMIDACLOPRID_DOI = "10.1186/s13568-019-0942-y"


ROWS = [
    {
        "pesticide": "Acetamiprid",
        "pathway_name": "Ochrobactrum sp. D-12 acetamiprid biodegradation",
        "microorganism": "Ochrobactrum sp. D-12",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Acetamiprid",
        "product": "N-methyl-(6-chloro-3-pyridyl)methylamine",
        "reaction_label": "Microbial transformation to chloropyridinyl methylamine",
        "enzyme": "Enzyme not identified",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": ACETAMIPRID_DOI,
        "reference_title": "Microbial Degradation of Acetamiprid by Ochrobactrum sp. D-12 Isolated from Contaminated Soil",
        "source_pdf": "Acetamiprid_10.1371_journal.pone.0082603.pdf",
        "evidence_note": "LC-MS identified N-methyl-(6-chloro-3-pyridyl)methylamine as a metabolic intermediate during acetamiprid degradation by strain D-12.",
    },
    {
        "pesticide": "DDT",
        "pathway_name": "Pleurotus ostreatus and Pseudomonas aeruginosa DDT transformation products",
        "microorganism": "Pleurotus ostreatus; Pseudomonas aeruginosa",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "DDT",
        "product": "DDD",
        "reaction_label": "Reductive dechlorination product detected",
        "enzyme": "Enzyme not identified",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": DDT_DOI,
        "reference_title": "Optimization of DDT degradation by Pleurotus ostreatus using biosurfactant-producing bacteria",
        "source_pdf": "DDT_10.4014_jmb.1701.01073.pdf",
        "evidence_note": "DDD was detected as a metabolic product from DDT degradation in mixed fungal-bacterial culture.",
    },
    {
        "pesticide": "DDT",
        "pathway_name": "Pleurotus ostreatus and Pseudomonas aeruginosa DDT transformation products",
        "microorganism": "Pleurotus ostreatus; Pseudomonas aeruginosa",
        "completeness": "PARTIAL",
        "step_order": 2,
        "substrate": "DDT",
        "product": "DDE",
        "reaction_label": "Dehydrochlorination product detected",
        "enzyme": "Enzyme not identified",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": DDT_DOI,
        "reference_title": "Optimization of DDT degradation by Pleurotus ostreatus using biosurfactant-producing bacteria",
        "source_pdf": "DDT_10.4014_jmb.1701.01073.pdf",
        "evidence_note": "DDE was detected as a metabolic product from DDT degradation in mixed fungal-bacterial culture.",
    },
    {
        "pesticide": "DDT",
        "pathway_name": "Pleurotus ostreatus and Pseudomonas aeruginosa DDT transformation products",
        "microorganism": "Pleurotus ostreatus; Pseudomonas aeruginosa",
        "completeness": "PARTIAL",
        "step_order": 3,
        "substrate": "DDT",
        "product": "DDMU",
        "reaction_label": "Dechlorination product detected",
        "enzyme": "Enzyme not identified",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": DDT_DOI,
        "reference_title": "Optimization of DDT degradation by Pleurotus ostreatus using biosurfactant-producing bacteria",
        "source_pdf": "DDT_10.4014_jmb.1701.01073.pdf",
        "evidence_note": "DDMU was detected as a metabolic product from DDT degradation in mixed fungal-bacterial culture.",
    },
    {
        "pesticide": "Glyphosate",
        "pathway_name": "Bacillus cereus CB4 glyphosate degradation routes",
        "microorganism": "Bacillus cereus CB4",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Glyphosate",
        "product": "AMPA + Glyoxylate",
        "reaction_label": "Glyphosate oxidoreductase route",
        "enzyme": "Glyphosate oxidoreductase activity",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": GLYPHOSATE_DOI,
        "reference_title": "Characterization of a glyphosate-degrading Bacillus cereus CB4 strain",
        "source_pdf": "Glyphosate_10.2323_jgam.58.263.pdf",
        "evidence_note": "The paper reports glyphosate degradation to AMPA and glyoxylate by glyphosate oxidoreductase activity.",
    },
    {
        "pesticide": "Glyphosate",
        "pathway_name": "Bacillus cereus CB4 glyphosate degradation routes",
        "microorganism": "Bacillus cereus CB4",
        "completeness": "PARTIAL",
        "step_order": 2,
        "substrate": "Glyphosate",
        "product": "Sarcosine",
        "reaction_label": "C-P bond cleavage",
        "enzyme": "C-P lyase activity",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": GLYPHOSATE_DOI,
        "reference_title": "Characterization of a glyphosate-degrading Bacillus cereus CB4 strain",
        "source_pdf": "Glyphosate_10.2323_jgam.58.263.pdf",
        "evidence_note": "The paper reports a concurrent C-P lyase route producing sarcosine during glyphosate degradation.",
    },
    {
        "pesticide": "Glyphosate",
        "pathway_name": "Bacillus cereus CB4 glyphosate degradation routes",
        "microorganism": "Bacillus cereus CB4",
        "completeness": "PARTIAL",
        "step_order": 3,
        "substrate": "Sarcosine",
        "product": "Glycine + Formaldehyde",
        "reaction_label": "Sarcosine downstream transformation",
        "enzyme": "C-P lyase-associated pathway activity",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": GLYPHOSATE_DOI,
        "reference_title": "Characterization of a glyphosate-degrading Bacillus cereus CB4 strain",
        "source_pdf": "Glyphosate_10.2323_jgam.58.263.pdf",
        "evidence_note": "The paper reports sarcosine, glycine, and formaldehyde as products in the C-P lyase-associated route.",
    },
    {
        "pesticide": "Imidacloprid",
        "pathway_name": "Hymenobacter latericoloratus CGMCC 16346 imidacloprid transformation",
        "microorganism": "Hymenobacter latericoloratus CGMCC 16346",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Imidacloprid",
        "product": "5-hydroxy imidacloprid",
        "reaction_label": "Hydroxylation",
        "enzyme": "Enzyme not identified",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": IMIDACLOPRID_DOI,
        "reference_title": "Biodegradation of imidacloprid by Hymenobacter latericoloratus CGMCC 16346",
        "source_pdf": "Imidacloprid_10.1186_s13568-019-0942-y.pdf",
        "evidence_note": "LC-MS and standards identified 5-hydroxy imidacloprid as a metabolite from imidacloprid transformation.",
    },
    {
        "pesticide": "Imidacloprid",
        "pathway_name": "Hymenobacter latericoloratus CGMCC 16346 imidacloprid transformation",
        "microorganism": "Hymenobacter latericoloratus CGMCC 16346",
        "completeness": "PARTIAL",
        "step_order": 2,
        "substrate": "5-hydroxy imidacloprid",
        "product": "Olefin imidacloprid",
        "reaction_label": "Dehydration / olefin formation",
        "enzyme": "Enzyme not identified",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": IMIDACLOPRID_DOI,
        "reference_title": "Biodegradation of imidacloprid by Hymenobacter latericoloratus CGMCC 16346",
        "source_pdf": "Imidacloprid_10.1186_s13568-019-0942-y.pdf",
        "evidence_note": "The paper describes imidacloprid transformation via hydroxylation to 5-hydroxy imidacloprid and the olefin metabolite.",
    },
]


SCREENING_DECISIONS = [
    {
        "pesticide": "Chlorothalonil",
        "decision": "Not integrated in this batch",
        "reason": "The downloaded open-access PDF did not yield a clear substrate-product transformation in text extraction. A pathway arrow should wait for a product table, figure, or readable paper text.",
    },
    {
        "pesticide": "Endosulfan",
        "decision": "Not integrated in this batch",
        "reason": "The paper confirms fungal degradation of alpha-endosulfan, beta-endosulfan, and endosulfan sulfate, but the extracted text did not provide a clear directed compound-to-compound step.",
    },
    {
        "pesticide": "Malathion",
        "decision": "Not integrated in this batch",
        "reason": "The downloaded papers provide enzyme or degradation context, but the extracted text did not identify a named malathion product suitable for a conservative pathway arrow.",
    },
    {
        "pesticide": "Dichlorvos",
        "decision": "Not integrated in this batch",
        "reason": "The downloaded open-access source did not provide an extractable named degradation product in the screening terms.",
    },
    {
        "pesticide": "Paraquat",
        "decision": "Not integrated in this batch",
        "reason": "The paper reports paraquat anaerobic transformation and insoluble crystals, but the product identity was not resolved as a named pathway compound.",
    },
]


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 64)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(ROWS)
    decisions = pd.DataFrame(SCREENING_DECISIONS)
    df.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch11_final.csv", index=False)
    decisions.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch11_screening_decisions.csv", index=False)
    df.to_csv(MASTER_CSV, index=False)
    with pd.ExcelWriter(OUT_DIR / "pesticidedb_stepwise_pathway_batch11_final.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 11", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 11", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    print(f"rows={len(df)}")
    print(f"pesticides={df['pesticide'].nunique()}")


if __name__ == "__main__":
    main()
