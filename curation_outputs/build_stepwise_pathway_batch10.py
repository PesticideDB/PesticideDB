from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "stepwise_pathway_batch10_20260708"
MASTER_XLSX = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch10.xlsx"
MASTER_CSV = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch10.csv"


TWO_FOUR_D_PATHWAY = "Curated tfd/pca pathway for 2,4-D degradation"
TWO_FOUR_D_REFERENCE = "Engineered E. coli expression of 2,4-D degradation pathway genes tfdA-F and pcaIJF"
TWO_FOUR_D_DOI = "10.1016/j.jhazmat.2023.131099"

MCPA_PATHWAY = "Initial tfdA-catalyzed MCPA cleavage"
MCPA_REFERENCE = "Genetic and microbial evidence for tfdA-mediated MCPA degradation"
MCPA_DOI = "10.1128/jb.184.15.4054-4064.2002"

CARBOFURAN_PATHWAY = "Carbofuran hydrolase-mediated carbamate cleavage"
CARBOFURAN_REFERENCE = "Cloning and expression of a carbofuran hydrolase gene from Achromobacter sp. WM111"
CARBOFURAN_DOI = "10.1128/jb.171.7.4038-4044.1989"


ROWS = [
    {
        "pesticide": "2,4-D",
        "pathway_name": TWO_FOUR_D_PATHWAY,
        "microorganism": "Engineered Escherichia coli strain BL-3164; Cupriavidus pinatubonensis JMP134 pathway genes",
        "completeness": "COMPLETE",
        "step_order": 1,
        "substrate": "2,4-D",
        "product": "2,4-dichlorophenol",
        "reaction_label": "alpha-ketoglutarate-dependent dioxygenation / ether cleavage",
        "enzyme": "2,4-dichlorophenoxyacetate dioxygenase",
        "gene": "tfdA",
        "evidence_type": "GENETIC",
        "doi": TWO_FOUR_D_DOI,
        "reference_title": TWO_FOUR_D_REFERENCE,
        "source_pdf": "Database curated protein/evidence records; source PDF not local",
        "evidence_note": "PesticideDB protein records include tfdA as the first 2,4-D degradation enzyme in the engineered tfd/pca pathway.",
    },
    {
        "pesticide": "2,4-D",
        "pathway_name": TWO_FOUR_D_PATHWAY,
        "microorganism": "Cupriavidus pinatubonensis JMP134; engineered Escherichia coli strain BL-3164",
        "completeness": "COMPLETE",
        "step_order": 2,
        "substrate": "2,4-dichlorophenol",
        "product": "3,5-dichlorocatechol",
        "reaction_label": "Hydroxylation",
        "enzyme": "2,4-dichlorophenol 6-monooxygenase",
        "gene": "tfdB",
        "evidence_type": "GENETIC",
        "doi": TWO_FOUR_D_DOI,
        "reference_title": TWO_FOUR_D_REFERENCE,
        "source_pdf": "Database curated protein/evidence records; source PDF not local",
        "evidence_note": "PesticideDB protein records include tfdB/2,4-dichlorophenol 6-monooxygenase for the chlorophenol-to-chlorocatechol step.",
    },
    {
        "pesticide": "2,4-D",
        "pathway_name": TWO_FOUR_D_PATHWAY,
        "microorganism": "Engineered Escherichia coli strain BL-3164",
        "completeness": "COMPLETE",
        "step_order": 3,
        "substrate": "3,5-dichlorocatechol",
        "product": "2,4-dichloro-cis,cis-muconate",
        "reaction_label": "Ring cleavage",
        "enzyme": "Chlorocatechol 1,2-dioxygenase",
        "gene": "tfdC",
        "evidence_type": "GENETIC",
        "doi": TWO_FOUR_D_DOI,
        "reference_title": TWO_FOUR_D_REFERENCE,
        "source_pdf": "Database curated protein/evidence records; source PDF not local",
        "evidence_note": "PesticideDB protein records include tfdC/chlorocatechol 1,2-dioxygenase for chlorocatechol ortho-cleavage.",
    },
    {
        "pesticide": "2,4-D",
        "pathway_name": TWO_FOUR_D_PATHWAY,
        "microorganism": "Engineered Escherichia coli strain BL-3164",
        "completeness": "COMPLETE",
        "step_order": 4,
        "substrate": "2,4-dichloro-cis,cis-muconate",
        "product": "Dienelactone",
        "reaction_label": "Cycloisomerization",
        "enzyme": "Chloromuconate cycloisomerase",
        "gene": "tfdD",
        "evidence_type": "GENETIC",
        "doi": TWO_FOUR_D_DOI,
        "reference_title": TWO_FOUR_D_REFERENCE,
        "source_pdf": "Database curated protein/evidence records; source PDF not local",
        "evidence_note": "PesticideDB protein records include tfdD/chloromuconate cycloisomerase downstream of chlorocatechol ring cleavage.",
    },
    {
        "pesticide": "2,4-D",
        "pathway_name": TWO_FOUR_D_PATHWAY,
        "microorganism": "Engineered Escherichia coli strain BL-3164",
        "completeness": "COMPLETE",
        "step_order": 5,
        "substrate": "Dienelactone",
        "product": "Maleylacetate",
        "reaction_label": "Hydrolysis",
        "enzyme": "Dienelactone hydrolase",
        "gene": "tfdE",
        "evidence_type": "GENETIC",
        "doi": TWO_FOUR_D_DOI,
        "reference_title": TWO_FOUR_D_REFERENCE,
        "source_pdf": "Database curated protein/evidence records; source PDF not local",
        "evidence_note": "PesticideDB protein records include tfdE/dienelactone hydrolase in the lower chlorocatechol pathway.",
    },
    {
        "pesticide": "2,4-D",
        "pathway_name": TWO_FOUR_D_PATHWAY,
        "microorganism": "Engineered Escherichia coli strain BL-3164",
        "completeness": "COMPLETE",
        "step_order": 6,
        "substrate": "Maleylacetate",
        "product": "3-oxoadipate",
        "reaction_label": "Reduction",
        "enzyme": "Maleylacetate reductase",
        "gene": "tfdF",
        "evidence_type": "GENETIC",
        "doi": TWO_FOUR_D_DOI,
        "reference_title": TWO_FOUR_D_REFERENCE,
        "source_pdf": "Database curated protein/evidence records; source PDF not local",
        "evidence_note": "PesticideDB protein records include tfdF/maleylacetate reductase connecting the chlorocatechol route to beta-ketoadipate metabolism.",
    },
    {
        "pesticide": "2,4-D",
        "pathway_name": TWO_FOUR_D_PATHWAY,
        "microorganism": "Engineered Escherichia coli strain BL-3164",
        "completeness": "COMPLETE",
        "step_order": 7,
        "substrate": "3-oxoadipate",
        "product": "3-oxoadipyl-CoA",
        "reaction_label": "CoA transfer",
        "enzyme": "Succinyl-CoA:3-oxoadipate CoA transferase",
        "gene": "pcaI/pcaJ",
        "evidence_type": "GENETIC",
        "doi": TWO_FOUR_D_DOI,
        "reference_title": TWO_FOUR_D_REFERENCE,
        "source_pdf": "Database curated protein/evidence records; source PDF not local",
        "evidence_note": "PesticideDB protein records include pcaI and pcaJ as beta-ketoadipate pathway CoA-transferase subunits.",
    },
    {
        "pesticide": "2,4-D",
        "pathway_name": TWO_FOUR_D_PATHWAY,
        "microorganism": "Engineered Escherichia coli strain BL-3164",
        "completeness": "COMPLETE",
        "step_order": 8,
        "substrate": "3-oxoadipyl-CoA",
        "product": "Acetyl-CoA + Succinyl-CoA",
        "reaction_label": "Thiolysis",
        "enzyme": "Beta-ketoadipyl-CoA thiolase",
        "gene": "pcaF",
        "evidence_type": "GENETIC",
        "doi": TWO_FOUR_D_DOI,
        "reference_title": TWO_FOUR_D_REFERENCE,
        "source_pdf": "Database curated protein/evidence records; source PDF not local",
        "evidence_note": "PesticideDB protein records include pcaF as the thiolase producing central-carbon-metabolism products.",
    },
    {
        "pesticide": "MCPA",
        "pathway_name": MCPA_PATHWAY,
        "microorganism": "Ralstonia eutropha JMP134; Comamonas acidovorans MC1; Cupriavidus spp.",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "MCPA",
        "product": "4-chloro-2-methylphenol",
        "reaction_label": "alpha-ketoglutarate-dependent ether cleavage",
        "enzyme": "Phenoxyalkanoic acid dioxygenase",
        "gene": "tfdA",
        "evidence_type": "GENETIC",
        "doi": MCPA_DOI,
        "reference_title": MCPA_REFERENCE,
        "source_pdf": "Database curated evidence records; source PDF not local",
        "evidence_note": "Database records link MCPA degradation to tfdA/tfd-like phenoxy acid degradation genes; this batch imports only the conservative initial cleavage product.",
    },
    {
        "pesticide": "Carbofuran",
        "pathway_name": CARBOFURAN_PATHWAY,
        "microorganism": "Achromobacter sp. WM111",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Carbofuran",
        "product": "Carbofuran phenol",
        "reaction_label": "Carbamate hydrolysis",
        "enzyme": "Carbofuran hydrolase",
        "gene": "mcd",
        "evidence_type": "GENETIC",
        "doi": CARBOFURAN_DOI,
        "reference_title": CARBOFURAN_REFERENCE,
        "source_pdf": "Database curated evidence records; source PDF not local",
        "evidence_note": "Database records link Achromobacter sp. WM111 carbofuran degradation to mcd/carbofuran hydrolase; only the primary hydrolysis product is imported here.",
    },
]


SCREENING_DECISIONS = [
    {
        "pesticide": "DDT",
        "decision": "Not integrated in this batch",
        "reason": "The database has extensive DDT evidence, but a full DDT route has multiple aerobic/anaerobic branches. Import should use a source-backed product table or pathway figure to avoid mixing branch-specific metabolites.",
    },
    {
        "pesticide": "Acetamiprid",
        "decision": "Not integrated in this batch",
        "reason": "The database has strong nitrile hydratase/amidase protein evidence, but exact product names need to be extracted from the original papers before a pathway sketch is imported.",
    },
    {
        "pesticide": "Carbosulfan",
        "decision": "Not integrated in this batch",
        "reason": "Only one database record was found and product/intermediate information was insufficient for a conservative stepwise arrow.",
    },
    {
        "pesticide": "Propoxur",
        "decision": "Not integrated in this batch",
        "reason": "Only one database record was found and product/intermediate information was insufficient for a conservative stepwise arrow.",
    },
]


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 64)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(ROWS)
    decisions = pd.DataFrame(SCREENING_DECISIONS)
    df.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch10_final.csv", index=False)
    decisions.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch10_screening_decisions.csv", index=False)
    df.to_csv(MASTER_CSV, index=False)
    with pd.ExcelWriter(OUT_DIR / "pesticidedb_stepwise_pathway_batch10_final.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 10", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 10", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    print(f"rows={len(df)}")
    print(f"pesticides={df['pesticide'].nunique()}")


if __name__ == "__main__":
    main()
