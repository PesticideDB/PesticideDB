from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "stepwise_pathway_batch17_20260710"
MASTER_XLSX = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch17.xlsx"
MASTER_CSV = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch17.csv"


ROWS = [
    {
        "pesticide": "Propiconazole",
        "pathway_name": "Burkholderia sp. BBK_9 propiconazole bioconversion",
        "microorganism": "Burkholderia sp. BBK_9",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Propiconazole",
        "product": "1-(2,4-dichlorophenyl)-2-(1H-1,2,4-triazol-1-yl) ethanone",
        "reaction_label": "Bioconversion / dioxolane-ring degradation",
        "enzyme": "Enzyme not assigned",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1007/s13205-016-0429-3",
        "reference_title": "Biodegradation of propiconazole by newly isolated Burkholderia sp. strain BBK_9",
        "source_pdf": "Propiconazole_10.1007_s13205-016-0429-3.pdf",
        "evidence_note": "The paper reports GC-MS/fragmentation-based formation of this metabolite during propiconazole bioconversion by Burkholderia sp. BBK_9 and proposes a degradation pathway. Chromosomal degradation genes were inferred; plasmid curing indicated plasmid-borne genes were not required.",
    },
    {
        "pesticide": "Propiconazole",
        "pathway_name": "Burkholderia sp. BBK_9 propiconazole bioconversion",
        "microorganism": "Burkholderia sp. BBK_9",
        "completeness": "PARTIAL",
        "step_order": 2,
        "substrate": "Propiconazole",
        "product": "1-[2-(4-chlorophenyl)ethyl]-1H-1,2,4-triazole",
        "reaction_label": "Bioconversion / dechlorinated triazole product formation",
        "enzyme": "Enzyme not assigned",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1007/s13205-016-0429-3",
        "reference_title": "Biodegradation of propiconazole by newly isolated Burkholderia sp. strain BBK_9",
        "source_pdf": "Propiconazole_10.1007_s13205-016-0429-3.pdf",
        "evidence_note": "The paper names this compound as one of three metabolites formed during propiconazole bioconversion by strain BBK_9. Chromosomal degradation genes were inferred; plasmid curing indicated plasmid-borne genes were not required.",
    },
    {
        "pesticide": "Propiconazole",
        "pathway_name": "Burkholderia sp. BBK_9 propiconazole bioconversion",
        "microorganism": "Burkholderia sp. BBK_9",
        "completeness": "PARTIAL",
        "step_order": 3,
        "substrate": "Propiconazole",
        "product": "1-ethyl-1H-1,2,4-triazole",
        "reaction_label": "Bioconversion / triazole product formation",
        "enzyme": "Enzyme not assigned",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1007/s13205-016-0429-3",
        "reference_title": "Biodegradation of propiconazole by newly isolated Burkholderia sp. strain BBK_9",
        "source_pdf": "Propiconazole_10.1007_s13205-016-0429-3.pdf",
        "evidence_note": "The paper names 1-ethyl-1H-1,2,4-triazole as a propiconazole bioconversion metabolite and notes that plasmid curing did not abolish degradation. Chromosomal degradation genes were inferred; plasmid curing indicated plasmid-borne genes were not required.",
    },
    {
        "pesticide": "Bentazone",
        "pathway_name": "Trametes versicolor bentazone transformation products",
        "microorganism": "Trametes versicolor",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Bentazone",
        "product": "TP256 (hydroxyl-bentazone)",
        "reaction_label": "Aromatic hydroxylation",
        "enzyme": "Cytochrome P450-associated activity",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1016/j.jhazmat.2020.124476",
        "reference_title": "Remediation of bentazone contaminated water by Trametes versicolor: characterization, identification of transformation products, and implementation in a trickle-bed reactor under non-sterile conditions",
        "source_pdf": "Bentazone_10.1016_j.jhazmat.2020.124476.pdf",
        "evidence_note": "UPLC-HRMS detected TP256 as hydroxyl-bentazone. The structure is tentative because reference standards were not used; the paper assigns confidence level 3.",
    },
    {
        "pesticide": "Bentazone",
        "pathway_name": "Trametes versicolor bentazone transformation products",
        "microorganism": "Trametes versicolor",
        "completeness": "PARTIAL",
        "step_order": 2,
        "substrate": "Bentazone",
        "product": "TP268 (oxidized N-methyl-bentazone form)",
        "reaction_label": "Oxidation / methylation-associated transformation",
        "enzyme": "Cytochrome P450-associated activity",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1016/j.jhazmat.2020.124476",
        "reference_title": "Remediation of bentazone contaminated water by Trametes versicolor: characterization, identification of transformation products, and implementation in a trickle-bed reactor under non-sterile conditions",
        "source_pdf": "Bentazone_10.1016_j.jhazmat.2020.124476.pdf",
        "evidence_note": "TP268 was identified as an oxidized form of N-methyl-bentazone; the paper reports tentative HRMS confidence-level assignment.",
    },
    {
        "pesticide": "Bentazone",
        "pathway_name": "Trametes versicolor bentazone transformation products",
        "microorganism": "Trametes versicolor",
        "completeness": "PARTIAL",
        "step_order": 3,
        "substrate": "TP268 (oxidized N-methyl-bentazone form)",
        "product": "TP284a (further oxidized N-methyl-bentazone form)",
        "reaction_label": "Further oxidation",
        "enzyme": "Cytochrome P450-associated activity",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1016/j.jhazmat.2020.124476",
        "reference_title": "Remediation of bentazone contaminated water by Trametes versicolor: characterization, identification of transformation products, and implementation in a trickle-bed reactor under non-sterile conditions",
        "source_pdf": "Bentazone_10.1016_j.jhazmat.2020.124476.pdf",
        "evidence_note": "TP284a appeared later than TP268, and the paper suggests TP284a may result from oxidation of TP268.",
    },
    {
        "pesticide": "Bentazone",
        "pathway_name": "Trametes versicolor bentazone transformation products",
        "microorganism": "Trametes versicolor",
        "completeness": "PARTIAL",
        "step_order": 4,
        "substrate": "Bentazone",
        "product": "TP284b (carboxylated bentazone form)",
        "reaction_label": "Carboxylation",
        "enzyme": "Enzyme not assigned",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1016/j.jhazmat.2020.124476",
        "reference_title": "Remediation of bentazone contaminated water by Trametes versicolor: characterization, identification of transformation products, and implementation in a trickle-bed reactor under non-sterile conditions",
        "source_pdf": "Bentazone_10.1016_j.jhazmat.2020.124476.pdf",
        "evidence_note": "TP284b is proposed as a carboxylated bentazone form, likely at the isopropyl moiety, based on HRMS fragmentation.",
    },
    {
        "pesticide": "Bentazone",
        "pathway_name": "Trametes versicolor bentazone transformation products",
        "microorganism": "Trametes versicolor",
        "completeness": "PARTIAL",
        "step_order": 5,
        "substrate": "Bentazone",
        "product": "TP285 (hydroxylated N-nitroso bentazone form)",
        "reaction_label": "Hydroxylation and N-nitrosation",
        "enzyme": "Enzyme not assigned",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1016/j.jhazmat.2020.124476",
        "reference_title": "Remediation of bentazone contaminated water by Trametes versicolor: characterization, identification of transformation products, and implementation in a trickle-bed reactor under non-sterile conditions",
        "source_pdf": "Bentazone_10.1016_j.jhazmat.2020.124476.pdf",
        "evidence_note": "The paper tentatively assigns TP285 as produced by hydroxylation and N-nitrosation of the parent compound's secondary amine.",
    },
]


SCREENING_DECISIONS = [
    {
        "pesticide": "Flubendiamide",
        "decision": "Not integrated",
        "reason": "Accessible papers report fungal/bacterial degradation and laccase involvement but do not provide named transformation products for a substrate-to-product arrow.",
    },
    {
        "pesticide": "Difenoconazole / Tebuconazole",
        "decision": "Not integrated",
        "reason": "Accessible papers support triazole fungicide degradation or dissipation but do not name pesticide-specific microbial products/intermediates in the tested system.",
    },
    {
        "pesticide": "Propiconazole yeast paper",
        "decision": "Not integrated",
        "reason": "The yeast paper reports propiconazole removal by Rhodotorula glutinis but does not identify transformation products; the Burkholderia paper was used instead.",
    },
]


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 72)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(ROWS)
    decisions = pd.DataFrame(SCREENING_DECISIONS)
    df.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch17_final.csv", index=False)
    decisions.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch17_screening_decisions.csv", index=False)
    df.to_csv(MASTER_CSV, index=False)
    with pd.ExcelWriter(OUT_DIR / "pesticidedb_stepwise_pathway_batch17_final.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 17", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 17", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    print(f"rows={len(df)}")
    print(f"pesticides={df['pesticide'].nunique()}")


if __name__ == "__main__":
    main()
