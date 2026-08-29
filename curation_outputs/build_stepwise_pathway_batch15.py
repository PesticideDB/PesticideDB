from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "stepwise_pathway_batch15_20260709"
MASTER_XLSX = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch15.xlsx"
MASTER_CSV = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch15.csv"


ROWS = [
    {
        "pesticide": "Aldrin and Dieldrin",
        "pathway_name": "Mucor racemosus aerobic dieldrin transformation",
        "microorganism": "Mucor racemosus strain DDF",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Dieldrin",
        "product": "Aldrin trans-diol",
        "reaction_label": "Aerobic fungal transformation",
        "enzyme": "Enzyme not assigned",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1584/jpestics.J18-03",
        "reference_title": "Biodegradability and biodegradation pathways of chlorinated cyclodiene insecticides by soil fungi",
        "source_pdf": "Aldrin_and_Dieldrin_10.1584_jpestics.J18-03.pdf",
        "evidence_note": "The paper reports >90% dieldrin degradation by Mucor racemosus strain DDF and production of aldrin trans-diol.",
    },
    {
        "pesticide": "Aldrin and Dieldrin",
        "pathway_name": "Mucor racemosus aerobic dieldrin transformation",
        "microorganism": "Mucor racemosus strain DDF",
        "completeness": "PARTIAL",
        "step_order": 2,
        "substrate": "Aldrin trans-diol",
        "product": "Aldrin trans-diol exo- and endo-phosphates",
        "reaction_label": "Phosphate metabolite formation",
        "enzyme": "Enzyme not assigned",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1584/jpestics.J18-03",
        "reference_title": "Biodegradability and biodegradation pathways of chlorinated cyclodiene insecticides by soil fungi",
        "source_pdf": "Aldrin_and_Dieldrin_10.1584_jpestics.J18-03.pdf",
        "evidence_note": "Strain DDF reduced aldrin trans-diol while producing metabolites determined as aldrin trans-diol exo- and endo-phosphates.",
    },
    {
        "pesticide": "Endosulfan",
        "pathway_name": "Mortierella aerobic endosulfan transformation",
        "microorganism": "Mortierella sp. strains W8 and Cm1-45",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Endosulfan",
        "product": "Endosulfan diol",
        "reaction_label": "Hydrolytic transformation",
        "enzyme": "Enzyme not assigned",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1584/jpestics.J18-03",
        "reference_title": "Biodegradability and biodegradation pathways of chlorinated cyclodiene insecticides by soil fungi",
        "source_pdf": "Aldrin_and_Dieldrin_10.1584_jpestics.J18-03.pdf",
        "evidence_note": "Mortierella sp. W8 and Cm1-45 generated endosulfan diol as the first reported transformation product of alpha- and beta-endosulfan.",
    },
    {
        "pesticide": "Endosulfan",
        "pathway_name": "Mortierella aerobic endosulfan transformation",
        "microorganism": "Mortierella sp. strains W8 and Cm1-45",
        "completeness": "PARTIAL",
        "step_order": 2,
        "substrate": "Endosulfan diol",
        "product": "Endosulfan lactone",
        "reaction_label": "Further fungal conversion",
        "enzyme": "Enzyme not assigned",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1584/jpestics.J18-03",
        "reference_title": "Biodegradability and biodegradation pathways of chlorinated cyclodiene insecticides by soil fungi",
        "source_pdf": "Aldrin_and_Dieldrin_10.1584_jpestics.J18-03.pdf",
        "evidence_note": "The paper reports further conversion of endosulfan diol to endosulfan lactone during fungal endosulfan degradation.",
    },
    {
        "pesticide": "Permethrin",
        "pathway_name": "Acinetobacter baumannii ZH-14 permethrin hydrolysis pathway",
        "microorganism": "Acinetobacter baumannii ZH-14",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Permethrin",
        "product": "3-Phenoxybenzenemethanol",
        "reaction_label": "Ester hydrolysis",
        "enzyme": "Hydrolytic activity not assigned",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.3389/fmicb.2018.00098",
        "reference_title": "Kinetics and novel degradation pathway of permethrin in Acinetobacter baumannii ZH-14",
        "source_pdf": "Permethrin_10.3389_fmicb.2018.00098.pdf",
        "evidence_note": "HPLC/GC-MS identified 3-phenoxybenzenemethanol as a major intermediate, and the proposed pathway starts with ester-linkage hydrolysis; 2,2-dimethyl-3-(2,2-dichlorovinyl) cyclopropanecarboxylic acid is reported as the corresponding acid co-product rather than the continuing pathway node.",
    },
    {
        "pesticide": "Permethrin",
        "pathway_name": "Acinetobacter baumannii ZH-14 permethrin hydrolysis pathway",
        "microorganism": "Acinetobacter baumannii ZH-14",
        "completeness": "PARTIAL",
        "step_order": 2,
        "substrate": "3-Phenoxybenzenemethanol",
        "product": "3-Phenoxybenzaldehyde",
        "reaction_label": "Redox transformation",
        "enzyme": "Oxidoreductive activity not assigned",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.3389/fmicb.2018.00098",
        "reference_title": "Kinetics and novel degradation pathway of permethrin in Acinetobacter baumannii ZH-14",
        "source_pdf": "Permethrin_10.3389_fmicb.2018.00098.pdf",
        "evidence_note": "The paper identifies 3-phenoxybenzaldehyde as a major intermediate and proposes redox conversion from 3-phenoxybenzenemethanol.",
    },
    {
        "pesticide": "Permethrin",
        "pathway_name": "Acinetobacter baumannii ZH-14 permethrin hydrolysis pathway",
        "microorganism": "Acinetobacter baumannii ZH-14",
        "completeness": "PARTIAL",
        "step_order": 3,
        "substrate": "3-Phenoxybenzaldehyde",
        "product": "1,2-Benzenedicarboxylic acid",
        "reaction_label": "Diaryl cleavage",
        "enzyme": "Cleavage activity not assigned",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.3389/fmicb.2018.00098",
        "reference_title": "Kinetics and novel degradation pathway of permethrin in Acinetobacter baumannii ZH-14",
        "source_pdf": "Permethrin_10.3389_fmicb.2018.00098.pdf",
        "evidence_note": "The proposed permethrin pathway links 3-phenoxybenzaldehyde to 1,2-benzenedicarboxylic acid by diaryl cleavage.",
    },
    {
        "pesticide": "Bifenthrin",
        "pathway_name": "Candida pelliculosa ZS-02 bifenthrin hydrolysis pathway",
        "microorganism": "Candida pelliculosa strain ZS-02",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Bifenthrin",
        "product": "Cyclopropanecarboxylic acid + 2-methyl-3-biphenylyl methanol",
        "reaction_label": "Carboxylester hydrolysis",
        "enzyme": "Hydrolytic activity not assigned",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1371/journal.pone.0030862",
        "reference_title": "Microbial detoxification of bifenthrin by a novel yeast and its potential for contaminated soils treatment",
        "source_pdf": "Bifenthrin_10.1371_journal.pone.0030862.pdf",
        "evidence_note": "GC-MS based pathway analysis reports hydrolysis of bifenthrin to cyclopropanecarboxylic acid and 2-methyl-3-biphenylyl methanol.",
    },
    {
        "pesticide": "Bifenthrin",
        "pathway_name": "Candida pelliculosa ZS-02 bifenthrin hydrolysis pathway",
        "microorganism": "Candida pelliculosa strain ZS-02",
        "completeness": "PARTIAL",
        "step_order": 2,
        "substrate": "2-methyl-3-biphenylyl methanol",
        "product": "4-trifluoromethoxyphenol + 2-chloro-6-fluorobenzyl alcohol + 3,5-dimethoxyphenol",
        "reaction_label": "Biphenyl cleavage",
        "enzyme": "Cleavage activity not assigned",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1371/journal.pone.0030862",
        "reference_title": "Microbial detoxification of bifenthrin by a novel yeast and its potential for contaminated soils treatment",
        "source_pdf": "Bifenthrin_10.1371_journal.pone.0030862.pdf",
        "evidence_note": "The paper reports further transformation of 2-methyl-3-biphenylyl methanol by biphenyl cleavage to three named products.",
    },
]


SCREENING_DECISIONS = [
    {
        "pesticide": "Fenvalerate",
        "decision": "Not integrated in this batch",
        "reason": "Open-access papers identify multiple metabolites, but the extract does not provide a single unambiguous ordered substrate-to-product chain suitable for automatic pathway import.",
    },
    {
        "pesticide": "Bentazone",
        "decision": "Not integrated in this batch",
        "reason": "The paper reports 19 transformation products; the table/figure needs manual extraction before adding exact pathway arrows.",
    },
    {
        "pesticide": "Pyrimethanil / Fluxapyroxad / Paraquat / Pinoxaden",
        "decision": "Not integrated in this batch",
        "reason": "Accessible text supports degradation or genomic potential but does not give a clean named microbial substrate-to-product step.",
    },
]


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 64)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(ROWS)
    decisions = pd.DataFrame(SCREENING_DECISIONS)
    df.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch15_final.csv", index=False)
    decisions.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch15_screening_decisions.csv", index=False)
    df.to_csv(MASTER_CSV, index=False)
    with pd.ExcelWriter(OUT_DIR / "pesticidedb_stepwise_pathway_batch15_final.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 15", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 15", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    print(f"rows={len(df)}")
    print(f"pesticides={df['pesticide'].nunique()}")


if __name__ == "__main__":
    main()
