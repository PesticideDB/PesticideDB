from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "stepwise_pathway_batch7_20260707"
MASTER_XLSX = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch7.xlsx"
MASTER_CSV = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch7.csv"


SOURCE_PDF = "Detoxification Esterase StrH Initiates Strobilurin Fungicide Degradation in iHyphomicrobiumi sp. Strain DY-1..pdf"
REFERENCE_TITLE = "Detoxification Esterase StrH Initiates Strobilurin Fungicide Degradation in Hyphomicrobium sp. Strain DY-1"
DOI = "10.1128/AEM.00103-21"
PATHWAY_NAME = "Strobilurin de-esterification by Hyphomicrobium sp. strain DY-1 esterase StrH"


ROWS = [
    {
        "pesticide": "Trifloxystrobin",
        "pathway_name": PATHWAY_NAME,
        "microorganism": "Hyphomicrobium sp. strain DY-1",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Trifloxystrobin",
        "product": "Trifloxystrobin acid",
        "reaction_label": "De-esterification",
        "enzyme": "Detoxification esterase StrH",
        "gene": "strH",
        "evidence_type": "GENETIC",
        "doi": DOI,
        "reference_title": REFERENCE_TITLE,
        "source_pdf": SOURCE_PDF,
        "evidence_note": "MS/MS identified trifloxystrobin acid as the product of trifloxystrobin hydrolysis in strain DY-1; cloned strH and recombinant StrH transform trifloxystrobin to trifloxystrobin acid.",
    },
    {
        "pesticide": "Pyraclostrobin",
        "pathway_name": PATHWAY_NAME,
        "microorganism": "Hyphomicrobium sp. strain DY-1; recombinant StrH",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Pyraclostrobin",
        "product": "Pyraclostrobin acid",
        "reaction_label": "De-esterification",
        "enzyme": "Detoxification esterase StrH",
        "gene": "strH",
        "evidence_type": "GENETIC",
        "doi": DOI,
        "reference_title": REFERENCE_TITLE,
        "source_pdf": SOURCE_PDF,
        "evidence_note": "The paper reports purified/recombinant StrH catalytic activity toward pyraclostrobin and states that StrH de-esterifies strobilurin fungicides to the corresponding parent acids.",
    },
    {
        "pesticide": "Azoxystrobin",
        "pathway_name": PATHWAY_NAME,
        "microorganism": "Hyphomicrobium sp. strain DY-1; recombinant StrH",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Azoxystrobin",
        "product": "Azoxystrobin acid",
        "reaction_label": "De-esterification",
        "enzyme": "Detoxification esterase StrH",
        "gene": "strH",
        "evidence_type": "GENETIC",
        "doi": DOI,
        "reference_title": REFERENCE_TITLE,
        "source_pdf": SOURCE_PDF,
        "evidence_note": "The paper reports measurable StrH catalytic activity toward azoxystrobin and states that StrH de-esterifies strobilurin fungicides to the corresponding parent acids; activity is much lower than for trifloxystrobin.",
    },
]


SCREENING_DECISIONS = [
    {
        "pesticide": "Carbaryl",
        "decision": "Not integrated in this batch",
        "reason": "The high-priority local PDF files were HTML placeholders/unreadable, so product arrows were not imported from them.",
    },
    {
        "pesticide": "Chlorothalonil",
        "decision": "Not integrated in this batch",
        "reason": "The high-priority local hydrolytic dehalogenase PDF was an HTML placeholder/unreadable.",
    },
    {
        "pesticide": "Dicamba",
        "decision": "Not integrated in this batch",
        "reason": "The high-priority local methyltransferase PDFs were HTML placeholders/unreadable.",
    },
    {
        "pesticide": "Dichlobenil",
        "decision": "Not integrated in this batch",
        "reason": "The high-priority local biodegradation PDFs were HTML placeholders/unreadable.",
    },
    {
        "pesticide": "Cyromazine",
        "decision": "Not integrated in this batch",
        "reason": "The high-priority local melamine-degrading bacteria PDFs were HTML placeholders/unreadable.",
    },
    {
        "pesticide": "Propiconazole",
        "decision": "Not integrated in this batch",
        "reason": "The high-priority local Burkholderia sp. BBK-9 PDF was an HTML placeholder/unreadable.",
    },
    {
        "pesticide": "Carbendazim",
        "decision": "Not integrated in this batch",
        "reason": "The high-priority local carbendazim degrader PDFs were HTML placeholders/unreadable.",
    },
]


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 64)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(ROWS)
    decisions = pd.DataFrame(SCREENING_DECISIONS)
    df.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch7_final.csv", index=False)
    decisions.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch7_screening_decisions.csv", index=False)
    df.to_csv(MASTER_CSV, index=False)
    with pd.ExcelWriter(OUT_DIR / "pesticidedb_stepwise_pathway_batch7_final.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 7", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 7", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    print(f"rows={len(df)}")
    print(f"pesticides={df['pesticide'].nunique()}")


if __name__ == "__main__":
    main()
