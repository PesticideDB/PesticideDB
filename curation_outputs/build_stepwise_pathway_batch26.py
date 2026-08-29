from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "stepwise_pathway_batch26_20260712"
MASTER_XLSX = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch26.xlsx"
MASTER_CSV = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch26.csv"
UNRESOLVED_PRODUCT = "Microbial transformation products not resolved in current database source"


def row(pesticide, microorganism, product, reaction_label, enzyme, gene, evidence_type, doi, title, note):
    return {
        "pesticide": pesticide,
        "pathway_name": f"{microorganism} {pesticide} degradation evidence",
        "microorganism": microorganism,
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": pesticide,
        "product": product,
        "reaction_label": reaction_label,
        "enzyme": enzyme,
        "gene": gene,
        "evidence_type": evidence_type,
        "doi": doi,
        "reference_title": title,
        "source_pdf": "Database reference from pesticide_data.xlsx / PBDB_master_with_ids.xlsx",
        "evidence_note": note,
    }


ROWS = [
    row(
        "Indoxacarb",
        "Bacillus cereus",
        "Indoxacarb carboxylesterase hydrolysis products",
        "Carboxylesterase-associated hydrolysis",
        "Carboxylesterase",
        "",
        "WHOLE_CELL",
        "10.1016/j.bjm.2016.01.012",
        "Database DOI record for indoxacarb degradation by Bacillus cereus",
        "This DOI is present in pesticide_data.xlsx. The database supports carboxylesterase-associated degradation; exact product names are not standardized in the current source table.",
    ),
    row(
        "Pyrimethanil",
        "Serratia sarumanii SBS19",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.1007/s10532-025-10144-2",
        "Database DOI record for pyrimethanil degradation by Serratia sarumanii",
        "This DOI is present in pesticide_data.xlsx. The database supports microbial degradation; exact transformation products are not standardized in the current source table.",
    ),
    row(
        "Thiabendazole",
        "Sphingomonas phylotype (B13)",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.1007/s00253-017-8128-5",
        "Database DOI record for thiabendazole degradation by Sphingomonas phylotype B13",
        "This DOI is present in pesticide_data.xlsx. The database supports microbial degradation; exact transformation products are not standardized in the current source table.",
    ),
    row(
        "Carbosulfan",
        "Bacillus licheniformis strain B-1",
        "Carbosulfan phosphodiesterase-associated products",
        "Phosphodiesterase-associated transformation",
        "Phosphodiesterase",
        "",
        "WHOLE_CELL",
        "10.1007/s10532-020-09899-7",
        "Database DOI record for carbosulfan degradation by Bacillus licheniformis strain B-1",
        "This DOI is present in pesticide_data.xlsx. The database supports phosphodiesterase-associated degradation; exact metabolites are not standardized in the current source table.",
    ),
    row(
        "Propoxur",
        "Arthrobacter sp.",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.1080/03601239209372800",
        "Database DOI record for propoxur degradation by Arthrobacter sp.",
        "This DOI is present in pesticide_data.xlsx. The database supports microbial degradation; exact transformation products are not standardized in the current source table.",
    ),
    row(
        "Acibenzolar-S-methyl",
        "Bacillus subtilis GB03; Bacillus subtilis FZB24; Bacillus amyloliquefaciens IN937a; Bacillus pumilus SE34",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.1007/s10532-011-9509-6",
        "Database DOI record for acibenzolar-S-methyl degradation by plant growth-promoting Bacillus strains",
        "This DOI is present in pesticide_data.xlsx. The database supports microbial degradation but the reviewed source data does not provide a curated product arrow, so the pathway is represented conservatively.",
    ),
    row(
        "Penthiopyrad",
        "Bacillus subtilis PCM 486; Trichoderma harzianum KKP 534",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.3390/molecules25061421",
        "Database DOI record for penthiopyrad degradation by Bacillus subtilis and Trichoderma harzianum",
        "This DOI is present in pesticide_data.xlsx. The database supports microbial degradation; exact microbial transformation products are not standardized in the current source table.",
    ),
    row(
        "Propargite",
        "Pseudomonas putida SPR 13; Pseudomonas putida SPR 8",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.1016/j.jhazmat.2009.09.050",
        "Database DOI record for propargite degradation by Pseudomonas putida",
        "This DOI is present in pesticide_data.xlsx. The database supports microbial degradation; the source is represented conservatively because product identification was not standardized in the current database table.",
    ),
    row(
        "Fluxapyroxad",
        "Saccharomyces cerevisiae",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.1038/s41598-020-78177-6",
        "Database DOI record for fluxapyroxad degradation by Saccharomyces cerevisiae",
        "This DOI is present in pesticide_data.xlsx. The database supports biological degradation; exact transformation products are not standardized in the current source table.",
    ),
    row(
        "Propamocarb",
        "Paenibacillus polymyxa",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.1016/j.ecoenv.2018.10.093",
        "Database DOI record for propamocarb degradation by Paenibacillus polymyxa",
        "This DOI is present in pesticide_data.xlsx. The database supports microbial degradation; exact transformation products are not standardized in the current source table.",
    ),
]


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 72)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(ROWS)
    decisions = pd.DataFrame(
        [
            {
                "pesticide": pesticide,
                "decision": "Integrated with conservative product label",
                "reason": "Reference DOI is present in pesticide_data.xlsx. Product names were kept conservative where exact metabolites are not standardized in the current source data.",
            }
            for pesticide in sorted({r["pesticide"] for r in ROWS})
        ]
    )
    df.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch26_final.csv", index=False)
    decisions.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch26_screening_decisions.csv", index=False)
    df.to_csv(MASTER_CSV, index=False)
    with pd.ExcelWriter(OUT_DIR / "pesticidedb_stepwise_pathway_batch26_final.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 26", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 26", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    print(f"rows={len(df)}")
    print(f"pesticides={df['pesticide'].nunique()}")


if __name__ == "__main__":
    main()
