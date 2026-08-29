from __future__ import annotations

import csv
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "stepwise_pathway_batch3_20260707"
MASTER_XLSX = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch3.xlsx"
MASTER_CSV = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch3.csv"

SOURCE_FILES = [
    PROJECT_ROOT / "curation_outputs" / "2_Phenylphenol_pathway_curation_20260707" / "2_phenylphenol_pathway_steps.csv",
    PROJECT_ROOT / "curation_outputs" / "Acephate_pathway_curation_20260707" / "acephate_pathway_steps.csv",
]

EXCLUDED_STEP_IDS = {
    "OPP-S04",  # Generic downstream central metabolism, not a chemical product arrow.
    "OPP-S05",  # Side route starts from 2,2'-dihydroxybiphenyl, not the selected pesticide.
    "OPP-S06",  # Side route starts from 2,2'-dihydroxybiphenyl, not the selected pesticide.
    "OPP-S07",  # Product not identified.
    "ACE-S08",  # Product is a vague fragment class, not a defined pathway compound.
}

EVIDENCE_MAP = {
    "Solid": "GENETIC",
    "Dashed": "METABOLITE",
    "Dotted": "PROPOSED",
}


def evidence_type(row: dict[str, str]) -> str:
    arrow = row.get("arrow_style_recommendation", "").strip()
    mapped = EVIDENCE_MAP.get(arrow)
    if mapped:
        return mapped
    evidence = row.get("evidence_type", "").casefold()
    if "enzyme" in evidence or "gene" in evidence:
        return "GENETIC"
    if "detected" in evidence or "metabolite" in evidence:
        return "METABOLITE"
    return "PROPOSED"


def clean(value: str | None) -> str:
    return (value or "").strip()


def include_row(row: dict[str, str]) -> bool:
    if row["step_id"] in EXCLUDED_STEP_IDS:
        return False
    product = row.get("product", "").casefold()
    if "not identified" in product or "fragments" in product or product == "central metabolism":
        return False
    return True


def read_rows() -> list[dict[str, str]]:
    rows = []
    for source in SOURCE_FILES:
        with source.open(newline="", encoding="utf-8-sig") as handle:
            for row in csv.DictReader(handle):
                if not include_row(row):
                    continue
                organism = clean(row["organism_or_system"])
                variant = clean(row["pathway_variant"])
                pathway_name = f"{clean(row['pesticide'])} degradation - {variant}"
                if variant == "Insect gut isolates":
                    pathway_name = f"{pathway_name} - {organism}"
                rows.append({
                    "pesticide": clean(row["pesticide"]),
                    "pathway_name": pathway_name,
                    "microorganism": organism,
                    "completeness": "PARTIAL",
                    "step_order": int(row["step_order"]),
                    "substrate": clean(row["substrate"]),
                    "product": clean(row["product"]),
                    "reaction_label": clean(row["enzyme"]) or clean(row["evidence_type"]),
                    "enzyme": clean(row["enzyme"]),
                    "gene": clean(row["gene"]),
                    "evidence_type": evidence_type(row),
                    "doi": clean(row["source_doi"]),
                    "reference_title": (
                        f"{clean(row['pesticide'])} pathway evidence - {clean(row['pathway_variant'])}"
                    ),
                    "source_pdf": f"Curated package row {clean(row['step_id'])}",
                    "evidence_note": clean(row["notes"]),
                })
    return rows


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 64)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(read_rows())
    df.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch3_final.csv", index=False)
    df.to_csv(MASTER_CSV, index=False)
    with pd.ExcelWriter(OUT_DIR / "pesticidedb_stepwise_pathway_batch3_final.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 3", index=False)
        style_workbook(writer)
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 3", index=False)
        style_workbook(writer)
    print(f"rows={len(df)}")
    print(f"pesticides={df['pesticide'].nunique()}")
    print(f"project_master={MASTER_XLSX}")


if __name__ == "__main__":
    main()
