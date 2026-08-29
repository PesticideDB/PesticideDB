from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "stepwise_pathway_batch23_20260712"
MASTER_XLSX = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch23.xlsx"
MASTER_CSV = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch23.csv"


def row(
    pesticide,
    pathway_name,
    microorganism,
    step_order,
    substrate,
    product,
    reaction_label,
    enzyme,
    gene,
    evidence_type,
    doi,
    reference_title,
    note,
    completeness="PARTIAL",
):
    return {
        "pesticide": pesticide,
        "pathway_name": pathway_name,
        "microorganism": microorganism,
        "completeness": completeness,
        "step_order": step_order,
        "substrate": substrate,
        "product": product,
        "reaction_label": reaction_label,
        "enzyme": enzyme,
        "gene": gene,
        "evidence_type": evidence_type,
        "doi": doi,
        "reference_title": reference_title,
        "source_pdf": "Database reference from pesticide_data.xlsx / PBDB_master_with_ids.xlsx",
        "evidence_note": note,
    }


UNRESOLVED_PRODUCT = "Microbial transformation products not resolved in current database source"


ROWS = [
    row(
        "Pirimiphos-Methyl",
        "Lactobacillus plantarum pirimiphos-methyl degradation evidence",
        "Lactobacillus plantarum",
        1,
        "Pirimiphos-Methyl",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.1111/lam.12128",
        "Database DOI record for pirimiphos-methyl degradation by Lactobacillus plantarum",
        "This DOI is present in pesticide_data.xlsx. The database supports microbial degradation; exact transformation products are not standardized in the current source table.",
    ),
    row(
        "Fenpropathrin",
        "Sphingobium sp. JQL4-5 fenpropathrin hydrolase evidence",
        "Sphingobium sp. JQL4-5(JQL4-5-mpd)",
        1,
        "Fenpropathrin",
        "Pyrethroid ester-cleavage products",
        "Hydrolase-associated pyrethroid degradation",
        "Methyl parathion hydrolase",
        "mpd",
        "GENETIC",
        "10.1016/j.jenvman.2010.06.010",
        "Database DOI record for mpd-associated fenpropathrin degradation",
        "This DOI is present in pesticide_data.xlsx. The database supports mpd hydrolase-associated degradation; exact product names are kept broad until metabolites are standardized in the source data.",
    ),
    row(
        "Prochloraz",
        "Bacillus cereus WD-2 prochloraz degradation evidence",
        "Bacillus cereus strain WD-2",
        1,
        "Prochloraz",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.1371/journal.pone.0220975",
        "Database DOI record for prochloraz degradation by Bacillus cereus WD-2",
        "This DOI is present in pesticide_data.xlsx. The database supports microbial prochloraz degradation; exact products are not stored in the current source table.",
    ),
    row(
        "Ethoprophos",
        "Sphingomonas and Flavobacterium ethoprophos degradation evidence",
        "Sphingomonas paucimobilis; Flavobacterium sp.",
        1,
        "Ethoprophos",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.1016/j.femsec.2005.01.012",
        "Database DOI record for ethoprophos degradation in soil bacteria",
        "This DOI is present in pesticide_data.xlsx. The database supports microbial degradation; exact transformation products are not standardized in the current source table.",
    ),
    row(
        "Flubendiamide",
        "Botryosphaeria and fungal flubendiamide laccase evidence",
        "Botryosphaeria rhodina; Trichoderma koningiopsis; Neurospora sp.",
        1,
        "Flubendiamide",
        "Oxidative transformation products not resolved in current database source",
        "Laccase/peroxidase-associated transformation",
        "Laccase",
        "",
        "WHOLE_CELL",
        "10.1007/s13213-019-01536-w",
        "Database DOI record for fungal flubendiamide degradation",
        "This DOI is present in pesticide_data.xlsx. The database supports fungal enzymatic/degradation evidence; exact products are not standardized in the current source table.",
    ),
    row(
        "Mesotrione",
        "Amycolatopsis nivea La24 mesotrione nitroreductase evidence",
        "Amycolatopsis nivea La24",
        1,
        "Mesotrione",
        "Mesotrione reduction products",
        "Nitroreductase-associated transformation",
        "Nitroreductase",
        "",
        "WHOLE_CELL",
        "10.1016/j.jhazmat.2024.134951",
        "Database DOI record for mesotrione degradation by Amycolatopsis nivea La24",
        "This DOI is present in pesticide_data.xlsx. The database supports nitroreductase-associated degradation; exact product labels are conservative until metabolites are standardized in the source table.",
    ),
    row(
        "Boscalid",
        "Lysinibacillus and Bacillus boscalid degradation evidence",
        "Lysinibacillus boronitolerans; Peribacillus muralis; Bacillus simplex",
        1,
        "Boscalid",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.1016/j.envpol.2022.120484",
        "Database DOI record for boscalid degradation by Bacillus-related isolates",
        "This DOI is present in pesticide_data.xlsx. The database supports microbial degradation but does not store named transformation products.",
    ),
    row(
        "Chlorpropham",
        "Bacillus licheniformis NKC-1 chlorpropham hydrolase pathway",
        "Bacillus licheniformis NKC-1",
        1,
        "Chlorpropham",
        "3-chloroaniline and downstream chlorocatechol products",
        "Hydrolysis followed by dioxygenase-associated aromatic transformation",
        "Chlorpropham hydrolase; 3-chloroaniline dioxygenase; 4-chlorocatechol 1,2-dioxygenase",
        "",
        "WHOLE_CELL",
        "10.1007/s11274-018-2494-8",
        "Database DOI record for chlorpropham degradation by Bacillus licheniformis NKC-1",
        "This DOI is present in pesticide_data.xlsx. The database supports hydrolase and dioxygenase evidence for chlorpropham degradation; the product label summarizes the supported enzyme sequence conservatively.",
    ),
    row(
        "Clothianidin",
        "Pseudomonas clothianidin degradation evidence",
        "Pseudomonas oleovorans SA2; Pseudomonas stutzeri strain SMK",
        1,
        "Clothianidin",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.1021/acsomega.3c09749",
        "Database DOI record for clothianidin degradation by Pseudomonas oleovorans SA2",
        "This DOI is present in pesticide_data.xlsx. The database supports microbial degradation; exact product names are not standardized in the current source table.",
    ),
    row(
        "Dinotefuran",
        "Burkholderia sp. C3 dinotefuran hydrolase evidence",
        "Burkholderia sp. C3",
        1,
        "Dinotefuran",
        "Dinotefuran hydrolysis products",
        "Carbaryl hydrolase-associated transformation",
        "Carbaryl hydrolase",
        "cahA",
        "GENETIC",
        "10.1007/s10532-013-9629-2",
        "Database DOI record for Burkholderia sp. C3 hydrolase-associated dinotefuran degradation",
        "This DOI is present in pesticide_data.xlsx. The database supports cahA hydrolase-associated degradation; exact product names are kept conservative until standardized in the source table.",
    ),
]


SCREENING_DECISIONS = [
    {
        "pesticide": pesticide,
        "decision": "Integrated with conservative product label",
        "reason": "Reference DOI is present in pesticide_data.xlsx and/or PBDB_master_with_ids.xlsx. Product names were kept conservative where exact metabolites are not standardized in the current source data.",
    }
    for pesticide in sorted({r["pesticide"] for r in ROWS})
]


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 72)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(ROWS)
    decisions = pd.DataFrame(SCREENING_DECISIONS)
    df.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch23_final.csv", index=False)
    decisions.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch23_screening_decisions.csv", index=False)
    df.to_csv(MASTER_CSV, index=False)
    with pd.ExcelWriter(OUT_DIR / "pesticidedb_stepwise_pathway_batch23_final.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 23", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 23", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    print(f"rows={len(df)}")
    print(f"pesticides={df['pesticide'].nunique()}")


if __name__ == "__main__":
    main()
