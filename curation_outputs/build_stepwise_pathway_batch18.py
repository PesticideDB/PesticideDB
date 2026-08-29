from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "stepwise_pathway_batch18_20260711"
MASTER_XLSX = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch18.xlsx"
MASTER_CSV = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch18.csv"


ROWS = [
    {
        "pesticide": "Fipronil",
        "pathway_name": "Enterobacter chengduensis G2.8 fipronil metabolite formation",
        "microorganism": "Enterobacter chengduensis strain G2.8",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Fipronil",
        "product": "Fipronil-sulfone",
        "reaction_label": "Oxidation to fipronil-sulfone",
        "enzyme": "Enzyme not assigned; oxidation inferred from GC-MS metabolite profile",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.3390/life13091935",
        "reference_title": "Fipronil Degradation in Soil by Enterobacter chengduensis Strain G2.8: Metabolic Perspective",
        "source_pdf": "PMC10532730 XML via NCBI EFetch",
        "evidence_note": "The open-access article reports GC-MS monitoring of fipronil, fipronil-sulfone, and fipronil-sulfide during incubation with Enterobacter chengduensis G2.8. Fipronil-sulfone production was attributed to the fipronil oxidation pathway and was later degraded during the 14-day assay.",
    },
    {
        "pesticide": "Fipronil",
        "pathway_name": "Enterobacter chengduensis G2.8 fipronil metabolite formation",
        "microorganism": "Enterobacter chengduensis strain G2.8",
        "completeness": "PARTIAL",
        "step_order": 2,
        "substrate": "Fipronil",
        "product": "Fipronil-sulfide",
        "reaction_label": "Reduction to fipronil-sulfide",
        "enzyme": "Enzyme not assigned; reduction inferred from GC-MS metabolite profile",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.3390/life13091935",
        "reference_title": "Fipronil Degradation in Soil by Enterobacter chengduensis Strain G2.8: Metabolic Perspective",
        "source_pdf": "PMC10532730 XML via NCBI EFetch",
        "evidence_note": "The paper reports fipronil-sulfide as a measured metabolite formed during fipronil degradation by Enterobacter chengduensis G2.8. The authors describe fipronil-sulfide as the reduction product and report subsequent metabolite decrease during the assay.",
    },
]


SCREENING_DECISIONS = [
    {
        "pesticide": "Dimethoate",
        "decision": "Not integrated in this batch",
        "reason": "The most relevant local Dimethoate files, including Exiguobacterium and Bacillus licheniformis titles, were HTML download placeholders rather than readable article PDFs. Keep Dimethoate in the source-acquisition queue until a real full text is available.",
    },
    {
        "pesticide": "Fipronil",
        "decision": "Integrated",
        "reason": "The NCBI E-utilities XML for PMC10532730 was accessible and reported measured fipronil-sulfone and fipronil-sulfide during degradation by Enterobacter chengduensis G2.8.",
    },
]


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 72)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(ROWS)
    decisions = pd.DataFrame(SCREENING_DECISIONS)
    df.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch18_final.csv", index=False)
    decisions.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch18_screening_decisions.csv", index=False)
    df.to_csv(MASTER_CSV, index=False)
    with pd.ExcelWriter(OUT_DIR / "pesticidedb_stepwise_pathway_batch18_final.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 18", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 18", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    print(f"rows={len(df)}")
    print(f"pesticides={df['pesticide'].nunique()}")


if __name__ == "__main__":
    main()
