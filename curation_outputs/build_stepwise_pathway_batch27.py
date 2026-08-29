from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "stepwise_pathway_batch27_20260712"
MASTER_XLSX = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch27.xlsx"
MASTER_CSV = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch27.csv"
UNRESOLVED_PRODUCT = "Microbial transformation products not resolved in current database source"


def row(pesticide, microorganism, reaction_label, gene, doi, title, note):
    return {
        "pesticide": pesticide,
        "pathway_name": f"{microorganism} {pesticide} degradation evidence",
        "microorganism": microorganism,
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": pesticide,
        "product": UNRESOLVED_PRODUCT,
        "reaction_label": reaction_label,
        "enzyme": "",
        "gene": gene,
        "evidence_type": "WHOLE_CELL" if not gene else "GENETIC",
        "doi": doi,
        "reference_title": title,
        "source_pdf": "Database reference from pesticide_data.xlsx / PBDB_master_with_ids.xlsx",
        "evidence_note": note,
    }


ROWS = [
    row(
        "Difenoconazole",
        "Lysinibacillus sp. BTKU3",
        "Whole-cell degradation",
        "",
        "10.1016/j.chemosphere.2021.131694",
        "Database DOI record for difenoconazole degradation by Lysinibacillus sp. BTKU3",
        "This DOI is present in pesticide_data.xlsx. The database supports microbial difenoconazole degradation; exact transformation products are not standardized in the current source table.",
    ),
    row(
        "Imazamox",
        "Streptomycetaceae strains JX02 and JX06",
        "Whole-cell degradation",
        "",
        "10.1080/03601234.2022.2064673",
        "Database DOI record for imazamox degradation by Streptomycetaceae strains",
        "This DOI is present in pesticide_data.xlsx. The database supports microbial imazamox degradation; exact transformation products are not standardized in the current source table.",
    ),
    row(
        "Paraquat",
        "Pseudomonas geniculata strain PQ01",
        "Whole-cell degradation",
        "",
        "10.3389/fmicb.2020.02003",
        "Database DOI record for paraquat degradation by Pseudomonas geniculata strain PQ01",
        "This DOI is present in pesticide_data.xlsx. The database supports microbial paraquat degradation; exact transformation products are not standardized in the current source table.",
    ),
]


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 72)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(ROWS)
    decisions = pd.DataFrame(
        [
            {
                "pesticide": pesticide,
                "decision": "Integrated with conservative product label",
                "reason": "Reference DOI is present in pesticide_data.xlsx. Product names were kept conservative because exact metabolites are not standardized in the current source data.",
            }
            for pesticide in sorted({r["pesticide"] for r in ROWS})
        ]
    )
    df.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch27_final.csv", index=False)
    decisions.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch27_screening_decisions.csv", index=False)
    df.to_csv(MASTER_CSV, index=False)
    with pd.ExcelWriter(OUT_DIR / "pesticidedb_stepwise_pathway_batch27_final.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 27", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 27", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    print(f"rows={len(df)}")
    print(f"pesticides={df['pesticide'].nunique()}")


if __name__ == "__main__":
    main()
