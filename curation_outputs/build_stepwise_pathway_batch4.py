from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "stepwise_pathway_batch4_20260707"
MASTER_XLSX = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch4.xlsx"
MASTER_CSV = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch4.csv"


ROWS = [
    {
        "pesticide": "Thiophanate-Methyl",
        "pathway_name": "Thiophanate-Methyl biodegradation by Enterobacter sp. TDS-1 and Bacillus sp. TDS-2",
        "microorganism": "Enterobacter sp. TDS-1; Bacillus sp. TDS-2",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Thiophanate-methyl",
        "product": "Carbendazim (MBC)",
        "reaction_label": "Conversion to primary benzimidazole metabolite",
        "enzyme": "Enzyme not identified",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1007/s10532-010-9430-4",
        "reference_title": "Biodegradation kinetics of the benzimidazole fungicide thiophanate-methyl by bacteria isolated from loamy sand soil",
        "source_pdf": "Biodegradation kinetics of the benzimidazole fungicide thiophanate-methyl by bacteria isolated from loamy sand soil..pdf",
        "evidence_note": "The paper describes the common initial conversion of thiophanate-methyl to carbendazim and studies TM degradation by Enterobacter sp. TDS-1 and Bacillus sp. TDS-2.",
    },
    {
        "pesticide": "Thiophanate-Methyl",
        "pathway_name": "Thiophanate-Methyl biodegradation by Enterobacter sp. TDS-1 and Bacillus sp. TDS-2",
        "microorganism": "Enterobacter sp. TDS-1; Bacillus sp. TDS-2",
        "completeness": "PARTIAL",
        "step_order": 2,
        "substrate": "Carbendazim (MBC)",
        "product": "2-aminobenzimidazole (2-AB)",
        "reaction_label": "Carbendazim transformation to 2-aminobenzimidazole",
        "enzyme": "Enzyme not identified",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1007/s10532-010-9430-4",
        "reference_title": "Biodegradation kinetics of the benzimidazole fungicide thiophanate-methyl by bacteria isolated from loamy sand soil",
        "source_pdf": "Biodegradation kinetics of the benzimidazole fungicide thiophanate-methyl by bacteria isolated from loamy sand soil..pdf",
        "evidence_note": "Analysis of degradation products in soil indicated transformation of carbendazim (MBC) to 2-aminobenzimidazole (2-AB).",
    },
    {
        "pesticide": "Thiophanate-Methyl",
        "pathway_name": "Thiophanate-Methyl biodegradation during oyster mushroom cultivation",
        "microorganism": "Pleurotus ostreatus var. florida cultivation system",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Thiophanate-methyl",
        "product": "Carbendazim",
        "reaction_label": "Biodegradation to primary metabolite",
        "enzyme": "Enzyme not identified",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1007/s11356-019-07050-5",
        "reference_title": "Substrate sterilization with thiophanate-methyl and its biodegradation to carbendazim in oyster mushroom",
        "source_pdf": "Substrate sterilization with thiophanate-methyl and its biodegradation to carbendazim in oyster mushroom (Pleurotus ostreatus var. florida)..pdf",
        "evidence_note": "Residue analysis detected thiophanate-methyl and its primary metabolite carbendazim during oyster mushroom cultivation.",
    },
]


SCREENING_DECISIONS = [
    {
        "pesticide": "Mandipropamid",
        "screening_decision": "Not integrated in Batch 4",
        "reason": "Current text extraction confirms five degradation products and CGA 380778, but product identities/pathway sequence require figure/table verification before drawing arrows.",
        "recommended_next_action": "Render and manually inspect pathway/product figures from the water-sediment microcosm paper before integration.",
    },
    {
        "pesticide": "Pinoxaden",
        "screening_decision": "Not integrated in Batch 4",
        "reason": "Current local candidate reports Acinetobacter degradation and predicted genes, but named chemical products/intermediates were not located in the text screening.",
        "recommended_next_action": "Review full paper figures/tables or provide product/intermediate evidence before creating pathway arrows.",
    },
    {
        "pesticide": "Imazamox",
        "screening_decision": "Not integrated in Batch 4",
        "reason": "Strong product evidence in screened papers is primarily photolysis/soil fate rather than microbial degradation pathway evidence.",
        "recommended_next_action": "Integrate only if microbial product/intermediate evidence is confirmed or if the database scope is expanded to abiotic photolysis pathways.",
    },
]


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 64)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(ROWS)
    decisions = pd.DataFrame(SCREENING_DECISIONS)
    df.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch4_final.csv", index=False)
    decisions.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch4_screening_decisions.csv", index=False)
    df.to_csv(MASTER_CSV, index=False)
    with pd.ExcelWriter(OUT_DIR / "pesticidedb_stepwise_pathway_batch4_final.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 4", index=False)
        decisions.to_excel(writer, sheet_name="Screened Not Integrated", index=False)
        style_workbook(writer)
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 4", index=False)
        decisions.to_excel(writer, sheet_name="Screened Not Integrated", index=False)
        style_workbook(writer)
    print(f"rows={len(df)}")
    print(f"pesticides={df['pesticide'].nunique()}")
    print(f"screened_not_integrated={len(decisions)}")


if __name__ == "__main__":
    main()
