from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "stepwise_pathway_batch22_20260712"
MASTER_XLSX = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch22.xlsx"
MASTER_CSV = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch22.csv"


def row(
    pesticide,
    pathway_name,
    microorganism,
    step_order,
    substrate,
    product,
    reaction_label,
    enzyme,
    gene,
    evidence_type,
    doi,
    reference_title,
    note,
    completeness="PARTIAL",
):
    return {
        "pesticide": pesticide,
        "pathway_name": pathway_name,
        "microorganism": microorganism,
        "completeness": completeness,
        "step_order": step_order,
        "substrate": substrate,
        "product": product,
        "reaction_label": reaction_label,
        "enzyme": enzyme,
        "gene": gene,
        "evidence_type": evidence_type,
        "doi": doi,
        "reference_title": reference_title,
        "source_pdf": "Database reference from pesticide_data.xlsx / PBDB_master_with_ids.xlsx",
        "evidence_note": note,
    }


UNRESOLVED_PRODUCT = "Microbial transformation products not resolved in current database source"


ROWS = [
    row(
        "Esfenvalerate",
        "Bacillus sp. CBMAI2052 esfenvalerate carboxylesterase evidence",
        "Bacillus sp. CBMAI2052",
        1,
        "Esfenvalerate",
        "Pyrethroid ester-cleavage products",
        "Carboxylesterase-associated hydrolysis",
        "Carboxylesterases",
        "",
        "WHOLE_CELL",
        "10.21577/0103-5053.20200051",
        "Database DOI record for esfenvalerate degradation by Brazilian marine-derived bacteria",
        "This DOI is present in pesticide_data.xlsx. The database supports bacterial esfenvalerate degradation with carboxylesterase evidence; exact metabolite names are not standardized in the current source table.",
    ),
    row(
        "Metalaxyl",
        "Candida tropicalis metalaxyl degradation evidence",
        "Candida tropicalis",
        1,
        "Metalaxyl",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.1007/s00284-020-02121-0",
        "Database DOI record for metalaxyl degradation by Candida tropicalis",
        "This DOI is present in pesticide_data.xlsx. The current database source supports microbial degradation but does not store named transformation products.",
    ),
    row(
        "Dithiocarbamate",
        "Pseudomonas dithiocarbamate degradation evidence",
        "Pseudomonas otitidis strain TD-8; Pseudomonas stutzeri strain TD-18",
        1,
        "Dithiocarbamate",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.1007/s11356-022-22731-4",
        "Database DOI record for dithiocarbamate degradation by Pseudomonas strains",
        "This DOI is present in pesticide_data.xlsx. The database supports whole-cell degradation; no exact transformation product is stored in the current source table.",
    ),
    row(
        "Cypermethrin",
        "Lysinibacillus cresolivuorans HIS7 cypermethrin degradation evidence",
        "Lysinibacillus cresolivuorans strain HIS7",
        1,
        "Cypermethrin",
        "Pyrethroid degradation products not resolved in current database source",
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.3390/plants10091903",
        "Database DOI record for cypermethrin degradation by Lysinibacillus cresolivuorans HIS7",
        "This DOI is present in pesticide_data.xlsx. The database supports cypermethrin biodegradation; exact metabolites are not standardized in the current source table.",
    ),
    row(
        "Thiacloprid",
        "Ensifer adhaerens TMX-23 thiacloprid nitrile-hydration evidence",
        "Ensifer adhaerens TMX-23",
        1,
        "Thiacloprid",
        "Thiacloprid transformation products not resolved in current database source",
        "Nitrile-hydratase-associated transformation",
        "Nitrile hydratase",
        "nhcA; nhpA; nhnR/nhnS; nhnH/nhnO",
        "GENETIC",
        "10.1111/jam.15172",
        "Database DOI record for thiacloprid degradation by Ensifer adhaerens TMX-23",
        "This DOI is present in pesticide_data.xlsx. The database supports genetic nitrile-hydratase-associated evidence; exact product names are not stored in the current source table.",
    ),
    row(
        "Propylene oxide",
        "Xanthobacter Py2 propylene oxide oxidoreductase evidence",
        "Xanthobacter strain Py2",
        1,
        "Propylene oxide",
        "Propylene oxide oxidation products",
        "Pyridine nucleotide-disulfide oxidoreductase-associated transformation",
        "Pyridine nucleotide-disulfide oxidoreductase",
        "X.ORF3",
        "PURIFIED_ENZYME",
        "10.1128/jb.178.22.6644-6646.1996",
        "Database DOI record for Xanthobacter Py2 propylene oxide oxidoreductase",
        "This DOI is present in pesticide_data.xlsx. The database supports enzymatic oxidoreductase evidence; the product label is conservative because exact metabolites are not standardized in the current source table.",
    ),
    row(
        "Dichlobenil",
        "Nocardia sp. DDN dichlobenil degradation evidence",
        "Nocardia sp. DDN",
        1,
        "Dichlobenil",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.1016/j.envpol.2009.06.003",
        "Database DOI record for dichlobenil degradation by Nocardia sp. DDN",
        "This DOI is present in pesticide_data.xlsx. The database supports microbial degradation; exact transformation products are not stored in the current source table.",
    ),
    row(
        "Tebuconazole",
        "Trametes versicolor tebuconazole laccase evidence",
        "Trametes versicolor ATCC 42530",
        1,
        "Tebuconazole",
        "Oxidative transformation products not resolved in current database source",
        "Laccase-associated oxidative degradation",
        "Laccase",
        "",
        "WHOLE_CELL",
        "10.1016/j.ecoenv.2020.110419",
        "Database DOI record for tebuconazole degradation by Trametes versicolor",
        "This DOI is present in pesticide_data.xlsx. The database supports fungal laccase-associated degradation; exact intermediates are not standardized in the current source table.",
    ),
    row(
        "Fludioxonil",
        "Pseudomonas and consortium fludioxonil degradation evidence",
        "Pseudomonas sp.; Ochrobactrum sp.; Comamonas sp.",
        1,
        "Fludioxonil",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.1016/j.jhazmat.2020.122545",
        "Database DOI record for fludioxonil degradation by bacterial isolates",
        "This DOI is present in pesticide_data.xlsx. The database supports bacterial degradation; no exact product names are stored in the current source table.",
    ),
    row(
        "Methomyl",
        "Aminobacter aminovorans MDW-2 methomyl hydrolase pathway",
        "Aminobacter aminovorans MDW-2",
        1,
        "Methomyl",
        "Methomyl C-N hydrolysis products",
        "Carbamate C-N bond hydrolysis",
        "Carbamate C-N hydrolase AmeH",
        "AmeH",
        "GENETIC",
        "10.1128/AEM.02005-20.",
        "Database DOI record for methomyl hydrolase AmeH",
        "This DOI is present in pesticide_data.xlsx. The database supports AmeH carbamate C-N hydrolase evidence; the product label is conservative until exact product names are standardized in the source data.",
    ),
]


SCREENING_DECISIONS = [
    {
        "pesticide": pesticide,
        "decision": "Integrated with conservative product label",
        "reason": "Reference DOI is present in pesticide_data.xlsx and/or PBDB_master_with_ids.xlsx. Product names were kept conservative where exact metabolites are not standardized in the current source data.",
    }
    for pesticide in sorted({r["pesticide"] for r in ROWS})
]


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 72)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(ROWS)
    decisions = pd.DataFrame(SCREENING_DECISIONS)
    df.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch22_final.csv", index=False)
    decisions.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch22_screening_decisions.csv", index=False)
    df.to_csv(MASTER_CSV, index=False)
    with pd.ExcelWriter(OUT_DIR / "pesticidedb_stepwise_pathway_batch22_final.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 22", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 22", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    print(f"rows={len(df)}")
    print(f"pesticides={df['pesticide'].nunique()}")


if __name__ == "__main__":
    main()
