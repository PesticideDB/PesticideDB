from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "stepwise_pathway_batch5_20260707"
MASTER_XLSX = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch5.xlsx"
MASTER_CSV = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch5.csv"


ROWS = [
    {
        "pesticide": "Mandipropamid",
        "pathway_name": "Mandipropamid degradation in Yangtze River water-sediment microcosm",
        "microorganism": "Yangtze River water-sediment microcosm",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Mandipropamid",
        "product": "CGA 380778",
        "reaction_label": "Hydrolysis",
        "enzyme": "Hydrolysis; enzyme not identified",
        "gene": "",
        "evidence_type": "PROPOSED",
        "doi": "10.1016/j.scitotenv.2023.164650",
        "reference_title": "The enantioselective environmental fate of mandipropamid in water-sediment microcosms",
        "source_pdf": "The enantioselective environmental fate of mandipropamid in water-sediment microcosms Distribution, degradation, degradation pathways and toxicity assessment..pdf",
        "evidence_note": "LC-TOF-MS/MS identified CGA 380778 as a sediment degradation product; Fig. 5 proposes hydrolysis from mandipropamid.",
    },
    {
        "pesticide": "Mandipropamid",
        "pathway_name": "Mandipropamid degradation in Yangtze River water-sediment microcosm",
        "microorganism": "Yangtze River water-sediment microcosm",
        "completeness": "PARTIAL",
        "step_order": 2,
        "substrate": "Mandipropamid",
        "product": "SYN 536638",
        "reaction_label": "Reduction",
        "enzyme": "Reduction; enzyme not identified",
        "gene": "",
        "evidence_type": "PROPOSED",
        "doi": "10.1016/j.scitotenv.2023.164650",
        "reference_title": "The enantioselective environmental fate of mandipropamid in water-sediment microcosms",
        "source_pdf": "The enantioselective environmental fate of mandipropamid in water-sediment microcosms Distribution, degradation, degradation pathways and toxicity assessment..pdf",
        "evidence_note": "LC-TOF-MS/MS identified SYN 536638; text and Fig. 5 propose reduction at the alkynyl group of mandipropamid.",
    },
    {
        "pesticide": "Mandipropamid",
        "pathway_name": "Mandipropamid degradation in Yangtze River water-sediment microcosm",
        "microorganism": "Yangtze River water-sediment microcosm",
        "completeness": "PARTIAL",
        "step_order": 3,
        "substrate": "Mandipropamid",
        "product": "NOA 458422",
        "reaction_label": "Hydrolysis",
        "enzyme": "Hydrolysis; enzyme not identified",
        "gene": "",
        "evidence_type": "PROPOSED",
        "doi": "10.1016/j.scitotenv.2023.164650",
        "reference_title": "The enantioselective environmental fate of mandipropamid in water-sediment microcosms",
        "source_pdf": "The enantioselective environmental fate of mandipropamid in water-sediment microcosms Distribution, degradation, degradation pathways and toxicity assessment..pdf",
        "evidence_note": "LC-TOF-MS/MS identified NOA 458422; Fig. 5 proposes hydrolysis from mandipropamid.",
    },
    {
        "pesticide": "Mandipropamid",
        "pathway_name": "Mandipropamid degradation in Yangtze River water-sediment microcosm",
        "microorganism": "Yangtze River water-sediment microcosm",
        "completeness": "PARTIAL",
        "step_order": 4,
        "substrate": "Mandipropamid",
        "product": "SYN 521195",
        "reaction_label": "Hydrolysis",
        "enzyme": "Hydrolysis; enzyme not identified",
        "gene": "",
        "evidence_type": "PROPOSED",
        "doi": "10.1016/j.scitotenv.2023.164650",
        "reference_title": "The enantioselective environmental fate of mandipropamid in water-sediment microcosms",
        "source_pdf": "The enantioselective environmental fate of mandipropamid in water-sediment microcosms Distribution, degradation, degradation pathways and toxicity assessment..pdf",
        "evidence_note": "LC-TOF-MS/MS identified SYN 521195; text proposes methoxy hydrolysis to generate the phenol moiety.",
    },
    {
        "pesticide": "Mandipropamid",
        "pathway_name": "Mandipropamid degradation in Yangtze River water-sediment microcosm",
        "microorganism": "Yangtze River water-sediment microcosm",
        "completeness": "PARTIAL",
        "step_order": 5,
        "substrate": "SYN 521195",
        "product": "SYN 539678",
        "reaction_label": "Reduction",
        "enzyme": "Reduction; enzyme not identified",
        "gene": "",
        "evidence_type": "PROPOSED",
        "doi": "10.1016/j.scitotenv.2023.164650",
        "reference_title": "The enantioselective environmental fate of mandipropamid in water-sediment microcosms",
        "source_pdf": "The enantioselective environmental fate of mandipropamid in water-sediment microcosms Distribution, degradation, degradation pathways and toxicity assessment..pdf",
        "evidence_note": "Fig. 5 proposes reduction from SYN 521195 to SYN 539678.",
    },
    {
        "pesticide": "Mandipropamid",
        "pathway_name": "Mandipropamid degradation in Yangtze River water-sediment microcosm",
        "microorganism": "Yangtze River water-sediment microcosm",
        "completeness": "PARTIAL",
        "step_order": 6,
        "substrate": "SYN 536638",
        "product": "NOA 458422",
        "reaction_label": "Hydrolysis",
        "enzyme": "Hydrolysis; enzyme not identified",
        "gene": "",
        "evidence_type": "PROPOSED",
        "doi": "10.1016/j.scitotenv.2023.164650",
        "reference_title": "The enantioselective environmental fate of mandipropamid in water-sediment microcosms",
        "source_pdf": "The enantioselective environmental fate of mandipropamid in water-sediment microcosms Distribution, degradation, degradation pathways and toxicity assessment..pdf",
        "evidence_note": "Text and Fig. 5 propose hydrolysis from SYN 536638 to NOA 458422.",
    },
    {
        "pesticide": "Mandipropamid",
        "pathway_name": "Mandipropamid degradation in Yangtze River water-sediment microcosm",
        "microorganism": "Yangtze River water-sediment microcosm",
        "completeness": "PARTIAL",
        "step_order": 7,
        "substrate": "SYN 536638",
        "product": "SYN 539678",
        "reaction_label": "Hydrolysis",
        "enzyme": "Hydrolysis; enzyme not identified",
        "gene": "",
        "evidence_type": "PROPOSED",
        "doi": "10.1016/j.scitotenv.2023.164650",
        "reference_title": "The enantioselective environmental fate of mandipropamid in water-sediment microcosms",
        "source_pdf": "The enantioselective environmental fate of mandipropamid in water-sediment microcosms Distribution, degradation, degradation pathways and toxicity assessment..pdf",
        "evidence_note": "Text and Fig. 5 propose hydrolysis from SYN 536638 to SYN 539678.",
    },
]


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 64)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(ROWS)
    df.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch5_final.csv", index=False)
    df.to_csv(MASTER_CSV, index=False)
    with pd.ExcelWriter(OUT_DIR / "pesticidedb_stepwise_pathway_batch5_final.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 5", index=False)
        style_workbook(writer)
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 5", index=False)
        style_workbook(writer)
    print(f"rows={len(df)}")
    print(f"pesticides={df['pesticide'].nunique()}")


if __name__ == "__main__":
    main()
