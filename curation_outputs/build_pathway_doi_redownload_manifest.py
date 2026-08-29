from __future__ import annotations

from pathlib import Path
from urllib.parse import quote

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
USB_ROOT = Path("/Volumes/WENTBROKE/Pesticide degradation database/pesticidepdf")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "pathway_doi_redownload_manifest_20260708"
SOURCE_PRIORITY = PROJECT_ROOT / "PesticideDB_Pathway_Source_Acquisition_Priority.csv"


def split_dois(value: str) -> list[str]:
    dois = []
    for part in str(value or "").split(" | "):
        doi = part.strip()
        if not doi or doi.startswith("+"):
            continue
        if doi not in dois:
            dois.append(doi)
    return dois


def safe_folder_name(name: str) -> str:
    replacements = {
        "DDT": "DDT(Dichlorodiphenyltrichloroethane)",
        "Chlorpyrifos-methyl": "Chlorpyrifos-Methyl",
        "Aldrin and Dieldrin": "Aldrin",
        "Dithiocarbamate": "Dithiocarbamates",
        "Cypermethrin": "Cypermethrins",
    }
    return replacements.get(name, name)


def doi_filename(doi: str, pesticide: str) -> str:
    cleaned = (
        doi.replace("https://doi.org/", "")
        .replace("/", "_")
        .replace(":", "_")
        .replace(";", "_")
    )
    return f"{pesticide}_{cleaned}.pdf".replace(" ", "_")


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 82)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    source = pd.read_csv(SOURCE_PRIORITY).fillna("")
    rows = []
    for _, row in source.iterrows():
        pesticide = row["pesticide"]
        folder = safe_folder_name(pesticide)
        folder_path = USB_ROOT / folder
        for doi in split_dois(row.get("database_doi_leads", "")):
            rows.append(
                {
                    "acquisition_priority": row["acquisition_priority"],
                    "pesticide": pesticide,
                    "priority_group": row["priority_group"],
                    "doi": doi,
                    "doi_url": f"https://doi.org/{doi}",
                    "google_scholar_url": f"https://scholar.google.com/scholar?q={quote(doi)}",
                    "target_usb_folder": str(folder_path),
                    "suggested_pdf_filename": doi_filename(doi, pesticide),
                    "database_record_count": row["database_record_count"],
                    "protein_record_count": row["protein_record_count"],
                    "database_gene_leads": row["database_gene_leads"],
                    "database_enzyme_leads": row["database_enzyme_leads"],
                    "database_microbe_leads": row["database_microbe_leads"],
                    "source_action": row["recommended_source_action"],
                    "pathway_extraction_rule": "Import only if the paper reports a substrate-to-product/intermediate transformation, reaction/enzyme/gene evidence, and organism/source context.",
                }
            )

    manifest_columns = [
        "acquisition_priority",
        "pesticide",
        "priority_group",
        "doi",
        "doi_url",
        "google_scholar_url",
        "target_usb_folder",
        "suggested_pdf_filename",
        "database_record_count",
        "protein_record_count",
        "database_gene_leads",
        "database_enzyme_leads",
        "database_microbe_leads",
        "source_action",
        "pathway_extraction_rule",
    ]
    manifest = pd.DataFrame(rows, columns=manifest_columns)
    if manifest.empty:
        summary = pd.DataFrame(
            columns=["acquisition_priority", "pesticide", "doi_count", "protein_record_count", "database_record_count"]
        )
    else:
        manifest = manifest.sort_values(
            ["acquisition_priority", "protein_record_count", "database_record_count", "pesticide", "doi"],
            ascending=[True, False, False, True, True],
        )
        summary = (
            manifest.groupby(["acquisition_priority", "pesticide"], dropna=False)
            .agg(doi_count=("doi", "count"), protein_record_count=("protein_record_count", "max"), database_record_count=("database_record_count", "max"))
            .reset_index()
            .sort_values(["acquisition_priority", "protein_record_count", "database_record_count", "pesticide"], ascending=[True, False, False, True])
        )
    csv_path = OUT_DIR / "pesticidedb_pathway_doi_redownload_manifest.csv"
    xlsx_path = OUT_DIR / "pesticidedb_pathway_doi_redownload_manifest.xlsx"
    manifest.to_csv(csv_path, index=False)
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        manifest.to_excel(writer, sheet_name="DOI Redownload Manifest", index=False)
        style_workbook(writer)
    root_csv = PROJECT_ROOT / "PesticideDB_Pathway_DOI_Redownload_Manifest.csv"
    root_xlsx = PROJECT_ROOT / "PesticideDB_Pathway_DOI_Redownload_Manifest.xlsx"
    manifest.to_csv(root_csv, index=False)
    with pd.ExcelWriter(root_xlsx, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        manifest.to_excel(writer, sheet_name="DOI Redownload Manifest", index=False)
        style_workbook(writer)
    print(csv_path)
    print(f"doi_rows={len(manifest)} pesticides={manifest['pesticide'].nunique() if 'pesticide' in manifest else 0}")
    print(summary.head(20).to_string(index=False))


if __name__ == "__main__":
    main()
