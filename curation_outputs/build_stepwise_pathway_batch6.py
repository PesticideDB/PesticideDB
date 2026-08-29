from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "stepwise_pathway_batch6_20260707"
MASTER_XLSX = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch6.xlsx"
MASTER_CSV = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch6.csv"


ROWS = [
    {
        "pesticide": "Ametoctradin",
        "pathway_name": "Ametoctradin biodegradation by soil-derived microbial consortia",
        "microorganism": "Soil-derived microbial consortia; packed-bed microbial bioreactor",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Ametoctradin",
        "product": "M650F01",
        "reaction_label": "Microbial aliphatic-chain transformation",
        "enzyme": "Enzyme not identified",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.3389/fmicb.2020.01898",
        "reference_title": "Accelerated Biodegradation of the Agrochemical Ametoctradin by Soil-Derived Microbial Consortia",
        "source_pdf": "Accelerated Biodegradation of the Agrochemical Ametoctradin by Soil-Derived Microbial Consortia..pdf",
        "evidence_note": "HPLC-MS/MS confirmed ametoctradin degradation and identified M650F01 as one of four major metabolites formed by soil microorganisms metabolizing the long aliphatic chain.",
    },
    {
        "pesticide": "Ametoctradin",
        "pathway_name": "Ametoctradin biodegradation by soil-derived microbial consortia",
        "microorganism": "Soil-derived microbial consortia; packed-bed microbial bioreactor",
        "completeness": "PARTIAL",
        "step_order": 2,
        "substrate": "Ametoctradin",
        "product": "M650F02",
        "reaction_label": "Microbial aliphatic-chain transformation",
        "enzyme": "Enzyme not identified",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.3389/fmicb.2020.01898",
        "reference_title": "Accelerated Biodegradation of the Agrochemical Ametoctradin by Soil-Derived Microbial Consortia",
        "source_pdf": "Accelerated Biodegradation of the Agrochemical Ametoctradin by Soil-Derived Microbial Consortia..pdf",
        "evidence_note": "HPLC-MS/MS confirmed ametoctradin degradation and identified M650F02 as one of four major metabolites formed by soil microorganisms metabolizing the long aliphatic chain.",
    },
    {
        "pesticide": "Ametoctradin",
        "pathway_name": "Ametoctradin biodegradation by soil-derived microbial consortia",
        "microorganism": "Soil-derived microbial consortia; packed-bed microbial bioreactor",
        "completeness": "PARTIAL",
        "step_order": 3,
        "substrate": "Ametoctradin",
        "product": "M650F03",
        "reaction_label": "Microbial aliphatic-chain transformation",
        "enzyme": "Enzyme not identified",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.3389/fmicb.2020.01898",
        "reference_title": "Accelerated Biodegradation of the Agrochemical Ametoctradin by Soil-Derived Microbial Consortia",
        "source_pdf": "Accelerated Biodegradation of the Agrochemical Ametoctradin by Soil-Derived Microbial Consortia..pdf",
        "evidence_note": "HPLC-MS/MS confirmed ametoctradin degradation and identified M650F03; the paper reports M650F03 as the main degradation product in every soil type.",
    },
    {
        "pesticide": "Ametoctradin",
        "pathway_name": "Ametoctradin biodegradation by soil-derived microbial consortia",
        "microorganism": "Soil-derived microbial consortia; packed-bed microbial bioreactor",
        "completeness": "PARTIAL",
        "step_order": 4,
        "substrate": "Ametoctradin",
        "product": "M650F04",
        "reaction_label": "Microbial aliphatic-chain transformation",
        "enzyme": "Enzyme not identified",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.3389/fmicb.2020.01898",
        "reference_title": "Accelerated Biodegradation of the Agrochemical Ametoctradin by Soil-Derived Microbial Consortia",
        "source_pdf": "Accelerated Biodegradation of the Agrochemical Ametoctradin by Soil-Derived Microbial Consortia..pdf",
        "evidence_note": "HPLC-MS/MS confirmed ametoctradin degradation and identified M650F04 as one of four major metabolites formed by soil microorganisms metabolizing the long aliphatic chain.",
    },
    {
        "pesticide": "Phosmet",
        "pathway_name": "Phosmet microbial degradation by indigenous blueberry bacterial flora",
        "microorganism": "Enterobacter agglomerans; Pseudomonas fluorescens",
        "completeness": "PARTIAL",
        "step_order": 1,
        "substrate": "Phosmet",
        "product": "Phthalimide",
        "reaction_label": "Phosphorodithioate-chain cleavage",
        "enzyme": "Enzyme not identified",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1111/j.1750-3841.2007.00466.x",
        "reference_title": "Microbial degradation of phosmet on blueberry fruit and in aqueous systems by indigenous bacterial flora on lowbush blueberries (Vaccinium angustifolium)",
        "source_pdf": "Microbial degradation of phosmet on blueberry fruit and in aqueous systems by indigenous bacterial flora on lowbush blueberries (Vaccinium angustifolium)..pdf",
        "evidence_note": "The study reports that E. agglomerans and P. fluorescens cleaved the phosphorodithioate backbone and identifies phthalimide among microbial degradates.",
    },
    {
        "pesticide": "Phosmet",
        "pathway_name": "Phosmet microbial degradation by indigenous blueberry bacterial flora",
        "microorganism": "Enterobacter agglomerans; Pseudomonas fluorescens",
        "completeness": "PARTIAL",
        "step_order": 2,
        "substrate": "Phosmet",
        "product": "N-hydroxymethyl phthalimide",
        "reaction_label": "Phosphorodithioate-chain cleavage",
        "enzyme": "Enzyme not identified",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1111/j.1750-3841.2007.00466.x",
        "reference_title": "Microbial degradation of phosmet on blueberry fruit and in aqueous systems by indigenous bacterial flora on lowbush blueberries (Vaccinium angustifolium)",
        "source_pdf": "Microbial degradation of phosmet on blueberry fruit and in aqueous systems by indigenous bacterial flora on lowbush blueberries (Vaccinium angustifolium)..pdf",
        "evidence_note": "The study reports that E. agglomerans and P. fluorescens cleaved the phosphorodithioate backbone and identifies N-hydroxymethyl phthalimide among microbial degradates.",
    },
    {
        "pesticide": "Phosmet",
        "pathway_name": "Phosmet microbial degradation by indigenous blueberry bacterial flora",
        "microorganism": "Enterobacter agglomerans; Pseudomonas fluorescens",
        "completeness": "PARTIAL",
        "step_order": 3,
        "substrate": "Phosmet",
        "product": "Phthalamic acid",
        "reaction_label": "Phosphorodithioate-chain cleavage",
        "enzyme": "Enzyme not identified",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1111/j.1750-3841.2007.00466.x",
        "reference_title": "Microbial degradation of phosmet on blueberry fruit and in aqueous systems by indigenous bacterial flora on lowbush blueberries (Vaccinium angustifolium)",
        "source_pdf": "Microbial degradation of phosmet on blueberry fruit and in aqueous systems by indigenous bacterial flora on lowbush blueberries (Vaccinium angustifolium)..pdf",
        "evidence_note": "The study reports that E. agglomerans and P. fluorescens cleaved the phosphorodithioate backbone and identifies phthalamic acid among microbial degradates.",
    },
    {
        "pesticide": "Phosmet",
        "pathway_name": "Phosmet microbial degradation by indigenous blueberry bacterial flora",
        "microorganism": "Enterobacter agglomerans; Pseudomonas fluorescens",
        "completeness": "PARTIAL",
        "step_order": 4,
        "substrate": "Phosmet",
        "product": "Phthalic acid",
        "reaction_label": "Phosphorodithioate-chain cleavage",
        "enzyme": "Enzyme not identified",
        "gene": "",
        "evidence_type": "METABOLITE",
        "doi": "10.1111/j.1750-3841.2007.00466.x",
        "reference_title": "Microbial degradation of phosmet on blueberry fruit and in aqueous systems by indigenous bacterial flora on lowbush blueberries (Vaccinium angustifolium)",
        "source_pdf": "Microbial degradation of phosmet on blueberry fruit and in aqueous systems by indigenous bacterial flora on lowbush blueberries (Vaccinium angustifolium)..pdf",
        "evidence_note": "The study reports that E. agglomerans and P. fluorescens cleaved the phosphorodithioate backbone and identifies phthalic acid among microbial degradates.",
    },
]


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 64)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(ROWS)
    df.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch6_final.csv", index=False)
    df.to_csv(MASTER_CSV, index=False)
    with pd.ExcelWriter(OUT_DIR / "pesticidedb_stepwise_pathway_batch6_final.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 6", index=False)
        style_workbook(writer)
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 6", index=False)
        style_workbook(writer)
    print(f"rows={len(df)}")
    print(f"pesticides={df['pesticide'].nunique()}")


if __name__ == "__main__":
    main()
