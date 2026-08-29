from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "pathway_next_acquisition_queue_20260708"
MISSING = PROJECT_ROOT / "PesticideDB_Missing_Stepwise_Pathway_Information.csv"

SCREENED_THIS_PASS = {
    "Acibenzolar-S-methyl": "Readable files reviewed; no named microbial transformation products found. Search for metabolite/product pathway papers or exclude from stepwise pathway until stronger evidence is available.",
    "Fenamiphos": "Folder mismatch: readable files are fenamidone papers, not fenamiphos. Replace with true fenamiphos biodegradation/metabolite papers.",
    "Fluxapyroxad": "Readable files mainly quantify dissipation or persistence; product pathway not found. Search for fluxapyroxad metabolite/pathway papers with microbial or soil-product evidence.",
    "Imazamox": "Named product evidence was mainly photodegradation/sunlight-soil behavior. Search specifically for microbial imazamox metabolites.",
    "Penthiopyrad": "PAM evidence is plant/field residue hydrolysis rather than microbial pathway evidence. Search for microbial penthiopyrad transformation products.",
    "Pinoxaden": "Acinetobacter paper supports degradation/predicted genes but lacks named products. Search for pinoxaden metabolites from microbial degradation.",
    "Propamocarb": "Microbial carbamate paper reports propamocarb resistance; no product arrow. Search for other microbial propamocarb product studies.",
    "Propargite": "Pseudomonas paper explicitly did not study metabolites. Search for product-identification papers for propargite glycol ether or related products.",
}

LOCAL_FILE_OVERRIDES = {
    "Paraquat": "Candidate local files reviewed; biodegradation-relevant entries are HTML download placeholders, not readable PDFs. Redownload true full-text PDFs for paraquat biodegradation/metabolite papers before pathway extraction.",
}


def priority(row: pd.Series) -> str:
    real = int(row.get("real_pdf_count") or 0)
    placeholders = int(row.get("html_placeholder_count") or 0)
    pesticide = row["pesticide"]
    if pesticide in LOCAL_FILE_OVERRIDES:
        return "Placeholder files - redownload PDFs by DOI/title"
    if pesticide in SCREENED_THIS_PASS:
        return "Screened readable files - needs different product paper"
    if real > 0:
        return "Readable PDFs available - screen next"
    if placeholders > 0:
        return "Placeholder files - redownload PDFs by DOI/title"
    return "No local readable file - literature search needed"


def action(row: pd.Series) -> str:
    pesticide = row["pesticide"]
    if pesticide in LOCAL_FILE_OVERRIDES:
        return LOCAL_FILE_OVERRIDES[pesticide]
    if pesticide in SCREENED_THIS_PASS:
        return SCREENED_THIS_PASS[pesticide]
    real = int(row.get("real_pdf_count") or 0)
    placeholders = int(row.get("html_placeholder_count") or 0)
    if real > 0:
        return "Open top candidate PDFs and extract only substrate, product/intermediate, enzyme/gene if reported, microorganism, DOI, and exact evidence wording."
    if placeholders > 0:
        return "Redownload real PDF files for the top candidate titles/DOIs, then rerun PDF readability audit and pathway screening."
    return "Search literature for experimentally reported microbial degradation products before adding any pathway arrows."


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 72)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.read_csv(MISSING)
    queue = df.copy()
    queue.insert(1, "priority_group", queue.apply(priority, axis=1))
    queue.insert(2, "specific_next_action", queue.apply(action, axis=1))
    order = {
        "Readable PDFs available - screen next": 1,
        "Screened readable files - needs different product paper": 2,
        "Placeholder files - redownload PDFs by DOI/title": 3,
        "No local readable file - literature search needed": 4,
    }
    queue["_priority_sort"] = queue["priority_group"].map(order).fillna(99)
    queue = queue.sort_values(
        ["_priority_sort", "real_pdf_count", "html_placeholder_count", "pesticide"],
        ascending=[True, False, False, True],
    ).drop(columns=["_priority_sort"])

    summary = (
        queue.groupby("priority_group", dropna=False)
        .size()
        .reset_index(name="pesticide_count")
        .sort_values("priority_group")
    )

    csv_path = OUT_DIR / "pesticidedb_pathway_next_acquisition_queue.csv"
    xlsx_path = OUT_DIR / "pesticidedb_pathway_next_acquisition_queue.xlsx"
    queue.to_csv(csv_path, index=False)
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        queue.to_excel(writer, sheet_name="Next Acquisition Queue", index=False)
        style_workbook(writer)
    print(csv_path)
    print(f"remaining={len(queue)}")
    print(summary.to_string(index=False))


if __name__ == "__main__":
    main()
