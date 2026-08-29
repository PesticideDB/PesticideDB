from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path("/Users/nana/Desktop/PepDB/PepDatabase")
OUT_DIR = PROJECT_ROOT / "curation_outputs" / "stepwise_pathway_batch21_20260712"
MASTER_XLSX = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch21.xlsx"
MASTER_CSV = PROJECT_ROOT / "PesticideDB_Stepwise_Pathway_Master_Batch21.csv"


def row(
    pesticide,
    pathway_name,
    microorganism,
    step_order,
    substrate,
    product,
    reaction_label,
    enzyme,
    gene,
    evidence_type,
    doi,
    reference_title,
    note,
    completeness="PARTIAL",
):
    return {
        "pesticide": pesticide,
        "pathway_name": pathway_name,
        "microorganism": microorganism,
        "completeness": completeness,
        "step_order": step_order,
        "substrate": substrate,
        "product": product,
        "reaction_label": reaction_label,
        "enzyme": enzyme,
        "gene": gene,
        "evidence_type": evidence_type,
        "doi": doi,
        "reference_title": reference_title,
        "source_pdf": "Database reference from pesticide_data.xlsx / PBDB_master_with_ids.xlsx",
        "evidence_note": note,
    }


UNRESOLVED_PRODUCT = "Microbial transformation products not resolved in current database source"


ROWS = [
    row(
        "Dimethoate",
        "Sphingomonas sp. DC-6 dimethoate amidohydrolase pathway",
        "Sphingomonas sp. DC-6",
        1,
        "Dimethoate",
        "Dimethoate hydrolysis products",
        "Amidohydrolase-mediated hydrolysis",
        "Dimethoate amidohydrolase",
        "dmhA",
        "PURIFIED_ENZYME",
        "10.1007/s10529-015-2027-6",
        "Database DOI record for dimethoate amidohydrolase in Sphingomonas sp. DC-6",
        "This DOI is present in pesticide_data.xlsx and PBDB_master_with_ids.xlsx. The database supports purified/kinetic dimethoate amidohydrolase evidence; the exact product names are kept conservative because they are not standardized in the current source table.",
    ),
    row(
        "Pinoxaden",
        "Acinetobacter pittobacter T4P20-1 pinoxaden degradation evidence",
        "Acinetobacter pittobacter strain T4P20-1",
        1,
        "Pinoxaden",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation with predicted associated genes",
        "Predicted degradation-associated proteins",
        "psd; tgt; pyrF; hemF; phnX; folE",
        "WHOLE_CELL",
        "10.21203/rs.3.rs-1375498/v1",
        "Database DOI record for pinoxaden degradation by Acinetobacter pittobacter strain T4P20-1",
        "This DOI is present in pesticide_data.xlsx. The current database source supports degradation and predicted genes, but does not store named transformation products; therefore this is represented as a conservative evidence node rather than a complete metabolite pathway.",
    ),
    row(
        "Fenamiphos",
        "Microbacterium esteraromaticum fenamiphos hydrolysis evidence",
        "Microbacterium esteraromaticum",
        1,
        "Fenamiphos",
        "Fenamiphos hydrolysis products",
        "Organophosphorus hydrolase / esterase activity",
        "Organophosphorus hydrolase / esterase",
        "",
        "WHOLE_CELL",
        "10.1016/j.biortech.2008.12.043",
        "Database DOI record for fenamiphos degradation by Microbacterium esteraromaticum",
        "This DOI is present in pesticide_data.xlsx. The database supports hydrolytic degradation evidence; product labels remain conservative until exact metabolites are curated from the source article.",
    ),
    row(
        "Phosphonic acid",
        "Escherichia coli phosphonic acid C-P lyase pathway",
        "Escherichia coli",
        1,
        "Phosphonic acid",
        "C-P bond cleavage products",
        "C-P lyase-mediated phosphonate utilization",
        "C-P lyase",
        "phnC-phnP",
        "GENETIC",
        "10.1128/JB.01131-09",
        "Database DOI record for phn operon-associated phosphonic acid utilization",
        "This DOI is present in pesticide_data.xlsx. The database supports genetic C-P lyase evidence; products are kept as a conservative cleavage-product label because the database source does not define a single pesticide-style metabolite.",
    ),
    row(
        "Isoxaflutole",
        "White-rot fungus isoxaflutole oxidative degradation evidence",
        "Phanerochaete chrysosporium (ATCC 24725); Trametes versicolor",
        1,
        "Isoxaflutole",
        "Oxidative transformation products not resolved in current database source",
        "Laccase/peroxidase-associated oxidative transformation",
        "Laccase; lignin peroxidase; manganese peroxidase",
        "EC 1.10.3.2; LiP; MnP; bphA1; benA; bph; p450; ppah",
        "WHOLE_CELL",
        "10.1021/jf000397q",
        "Database DOI record for fungal isoxaflutole degradation",
        "This DOI is present in pesticide_data.xlsx. The database supports fungal oxidative degradation evidence; exact intermediates are not standardized in the current source table.",
    ),
    row(
        "Quinclorac",
        "Burkholderia cepacia WZ1 quinclorac aromatic-ring degradation evidence",
        "Burkholderia cepacia WZ1",
        1,
        "Quinclorac",
        "Chlorinated aromatic ring-cleavage products",
        "Dioxygenase/reductase-associated aromatic degradation",
        "Phthalate dioxygenase reductase; chlorocatechol 1,2-dioxygenase",
        "",
        "WHOLE_CELL",
        "10.1016/j.soilbio.2008.10.026",
        "Database DOI record for quinclorac degradation by Burkholderia cepacia WZ1",
        "This DOI is present in pesticide_data.xlsx. The database supports dioxygenase-linked quinclorac degradation; the product label is conservative because named metabolites are not stored in the current source table.",
    ),
    row(
        "Cyfluthrin/beta-cyfluthrin",
        "Bacillus subtilis BSF01 beta-cyfluthrin carboxylesterase pathway",
        "Bacillus subtilis BSF01",
        1,
        "Cyfluthrin/beta-cyfluthrin",
        "Pyrethroid ester-cleavage products",
        "Carboxylesterase-mediated ester hydrolysis",
        "Carboxylesterase assay system",
        "cesB",
        "GENETIC",
        "10.3389/fbioe.2020.00889",
        "Database DOI record for Bacillus subtilis BSF01 beta-cyfluthrin carboxylesterase",
        "This DOI is present in pesticide_data.xlsx. The database supports cesB-linked carboxylesterase activity; product names are kept broad until exact pyrethroid cleavage metabolites are curated from the article.",
    ),
    row(
        "Profenofos",
        "Cupriavidus nantongensis X1 profenofos organophosphate hydrolase pathway",
        "Cupriavidus nantongensis X1T",
        1,
        "Profenofos",
        "Profenofos hydrolysis products",
        "Organophosphate hydrolase-mediated hydrolysis",
        "Organophosphate hydrolase OpdB",
        "opdB",
        "GENETIC",
        "10.1021/acs.jafc.0c00132",
        "Database DOI record for OpdB-mediated profenofos degradation",
        "This DOI is present in pesticide_data.xlsx. The database supports OpdB-linked profenofos hydrolysis; exact product naming is deferred until metabolite names are standardized in the source data.",
    ),
    row(
        "Endrin",
        "Burkholderia and Cupriavidus endrin degradation evidence",
        "Burkholderia sp. strain MED-7; Cupriavidus sp. strain MED-5",
        1,
        "Endrin",
        UNRESOLVED_PRODUCT,
        "Whole-cell degradation",
        "",
        "",
        "WHOLE_CELL",
        "10.1007/s00253-008-1670-4",
        "Database DOI record for bacterial endrin degradation",
        "This DOI is present in pesticide_data.xlsx. The database supports whole-cell endrin degradation but does not store named intermediates in the current source table.",
    ),
    row(
        "Phorate",
        "Ralstonia eutropha AAJ1 phorate phosphatase degradation evidence",
        "Ralstonia eutropha AAJ1",
        1,
        "Phorate",
        "Phorate phosphoester-cleavage products",
        "Phosphomonoesterase/phosphodiesterase-associated transformation",
        "Phosphomonoesterase; phosphodiesterase",
        "",
        "WHOLE_CELL",
        "10.1111/j.1472-765X.2009.02631.x",
        "Database DOI record for phosphatase-associated phorate degradation",
        "This DOI is present in pesticide_data.xlsx. The database supports phosphatase-associated degradation evidence; product naming is conservative because exact metabolites are not stored in the current source table.",
    ),
]


SCREENING_DECISIONS = [
    {
        "pesticide": pesticide,
        "decision": "Integrated with conservative product label",
        "reason": "Reference DOI is present in pesticide_data.xlsx and/or PBDB_master_with_ids.xlsx. Product names were kept conservative where exact metabolites are not standardized in the current source data.",
    }
    for pesticide in sorted({r["pesticide"] for r in ROWS})
]


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 72)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(ROWS)
    decisions = pd.DataFrame(SCREENING_DECISIONS)
    df.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch21_final.csv", index=False)
    decisions.to_csv(OUT_DIR / "pesticidedb_stepwise_pathway_batch21_screening_decisions.csv", index=False)
    df.to_csv(MASTER_CSV, index=False)
    with pd.ExcelWriter(OUT_DIR / "pesticidedb_stepwise_pathway_batch21_final.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 21", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    with pd.ExcelWriter(MASTER_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stepwise Pathway Batch 21", index=False)
        decisions.to_excel(writer, sheet_name="Screening Decisions", index=False)
        style_workbook(writer)
    print(f"rows={len(df)}")
    print(f"pesticides={df['pesticide'].nunique()}")


if __name__ == "__main__":
    main()
