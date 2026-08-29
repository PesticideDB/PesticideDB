from __future__ import annotations

import json
import os
import re
import sys
import time
from collections import defaultdict
from pathlib import Path

import django
import pandas as pd
from Bio import Entrez
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.append(str(PROJECT_ROOT))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "PepDatabase.settings")
django.setup()

from base.models import Pesticide  # noqa: E402


OUTPUT_DIR = PROJECT_ROOT / "curation_outputs"
OUTPUT_PATH = OUTPUT_DIR / "pesticide_degrading_microorganism_taxonomy.xlsx"
CACHE_PATH = OUTPUT_DIR / "taxonomy_lookup_cache.json"
Entrez.email = os.environ.get("NCBI_ENTREZ_EMAIL", "gurungsaru634@gmail.com")
Entrez.tool = "PesticideDBTaxonomyExport"

TARGET_RANKS = {
    "superkingdom": "Domain",
    "phylum": "Phylum",
    "class": "Class",
    "order": "Order",
    "family": "Family",
    "genus": "Genus",
    "species": "Species",
}

AMBIGUOUS_TERMS = {
    "",
    "-",
    "mixed microbe",
    "not specified",
    "unknown",
    "uncultured bacterium",
}


def clean(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def canonical(value) -> str:
    return clean(value).casefold()


def doi_from_record(record: Pesticide) -> str:
    doi = clean(record.doi)
    if doi:
        return doi
    reference = clean(record.reference)
    match = re.search(r"10\.\d{4,9}/\S+", reference)
    if not match:
        return ""
    return match.group(0).rstrip(".,;)")


def load_cache() -> dict[str, dict]:
    if CACHE_PATH.exists():
        return json.loads(CACHE_PATH.read_text())
    return {}


def save_cache(cache: dict[str, dict]) -> None:
    CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    CACHE_PATH.write_text(json.dumps(cache, indent=2, sort_keys=True))


def name_candidates(name: str) -> list[str]:
    cleaned = clean(name)
    candidates = []

    replacements = [
        (r"\bspecies\b", "sp."),
        (r"\bsp\s+", "sp. "),
        (r"\bspp\.\b", "spp."),
    ]
    normalized = cleaned
    for pattern, replacement in replacements:
        normalized = re.sub(pattern, replacement, normalized, flags=re.IGNORECASE)

    candidates.append(normalized)

    # Remove culture descriptors and strain labels when exact strain is absent from NCBI.
    stripped = re.sub(
        r"\b(strain|isolate|sp\.|spp\.|dominant in|in Tunisian|consortium|community|enrichment)\b.*$",
        "",
        normalized,
        flags=re.IGNORECASE,
    ).strip(" ,;:-")
    if stripped and stripped != normalized:
        candidates.append(stripped)

    words = normalized.split()
    if len(words) >= 2 and words[1].lower() not in {"sp.", "spp.", "species"}:
        candidates.append(" ".join(words[:2]))
    if words:
        candidates.append(words[0])

    unique = []
    seen = set()
    for candidate in candidates:
        key = canonical(candidate)
        if key and key not in seen and key not in AMBIGUOUS_TERMS:
            unique.append(candidate)
            seen.add(key)
    return unique


def parsed_binomial(name: str) -> tuple[str, str]:
    cleaned = clean(name)
    cleaned = re.sub(r"^engineered\s+", "", cleaned, flags=re.IGNORECASE)
    words = cleaned.split()
    if not words:
        return "", ""
    genus = words[0].strip("()[],:;")
    species = ""
    if len(words) >= 2 and words[1].lower() not in {"sp.", "spp.", "species", "strain", "isolate"}:
        species = words[1].strip("()[],:;")
    if not re.fullmatch(r"[A-Z][A-Za-z-]+", genus):
        return "", species
    return genus, species


def taxonomy_is_compatible(query_name: str, taxonomy: dict) -> bool:
    expected_genus, expected_species = parsed_binomial(query_name)
    if not expected_genus:
        return True

    lineage = clean(taxonomy.get("Taxonomy", ""))
    genus = clean(taxonomy.get("Genus", ""))
    species = clean(taxonomy.get("Species", ""))
    matched_name = clean(taxonomy.get("Matched_name", ""))

    if canonical(genus) == canonical(expected_genus):
        return True

    # Accept species-level synonym/reclassification only when the submitted
    # species epithet is still present in the NCBI species or matched name.
    if expected_species:
        expected = canonical(expected_species)
        species_tokens = {canonical(part) for part in species.split()}
        matched_tokens = {canonical(part) for part in matched_name.split()}
        if expected in species_tokens or expected in matched_tokens:
            return True

    # Generic "Pseudomonas sp." records should not silently become unrelated
    # genera just because an arbitrary strain number matched elsewhere.
    if f"; {expected_genus};" in f"{lineage};":
        return True
    return False


def entrez_search_taxid(candidate: str) -> str:
    query = f'"{candidate}"[Scientific Name]'
    with Entrez.esearch(db="taxonomy", term=query, retmode="xml", retmax=5) as handle:
        result = Entrez.read(handle)
    ids = result.get("IdList", [])
    if ids:
        return ids[0]

    with Entrez.esearch(db="taxonomy", term=candidate, retmode="xml", retmax=5) as handle:
        result = Entrez.read(handle)
    ids = result.get("IdList", [])
    return ids[0] if ids else ""


def entrez_fetch_taxonomy(taxid: str) -> dict:
    with Entrez.efetch(db="taxonomy", id=taxid, retmode="xml") as handle:
        records = Entrez.read(handle)
    if not records:
        return {}
    record = records[0]
    lineage = {rank: "" for rank in TARGET_RANKS.values()}
    for item in record.get("LineageEx", []):
        rank = item.get("Rank")
        if rank in TARGET_RANKS:
            lineage[TARGET_RANKS[rank]] = item.get("ScientificName", "")

    if record.get("Rank") in TARGET_RANKS:
        lineage[TARGET_RANKS[record["Rank"]]] = record.get("ScientificName", "")

    return {
        "TaxID": str(record.get("TaxId", "")),
        "Taxonomy": record.get("Lineage", ""),
        **lineage,
        "Matched_name": record.get("ScientificName", ""),
        "Matched_rank": record.get("Rank", ""),
    }


def lookup_taxonomy(name: str, cache: dict[str, dict]) -> dict:
    key = canonical(name)
    if key in cache:
        cached = cache[key]
        if cached.get("Taxonomy_status") == "matched" and not taxonomy_is_compatible(name, cached):
            del cache[key]
        else:
            return cached

    if key in AMBIGUOUS_TERMS:
        cache[key] = empty_taxonomy("ambiguous microorganism name")
        return cache[key]

    for candidate in name_candidates(name):
        try:
            taxid = entrez_search_taxid(candidate)
            time.sleep(0.34)
            if taxid:
                taxonomy = entrez_fetch_taxonomy(taxid)
                time.sleep(0.34)
                if not taxonomy_is_compatible(candidate, taxonomy):
                    continue
                taxonomy["Lookup_name"] = candidate
                taxonomy["Taxonomy_status"] = "matched"
                cache[key] = taxonomy
                return taxonomy
        except Exception as exc:
            cache[key] = empty_taxonomy(f"lookup error: {exc}")
            save_cache(cache)
            return cache[key]

    cache[key] = empty_taxonomy("no NCBI Taxonomy match")
    return cache[key]


def empty_taxonomy(status: str) -> dict:
    return {
        "TaxID": "",
        "Taxonomy": "",
        "Domain": "",
        "Phylum": "",
        "Class": "",
        "Order": "",
        "Family": "",
        "Genus": "",
        "Species": "",
        "Matched_name": "",
        "Matched_rank": "",
        "Lookup_name": "",
        "Taxonomy_status": status,
    }


def build_rows(cache: dict[str, dict]) -> list[dict]:
    records = (
        Pesticide.objects.exclude(microorganism__isnull=True)
        .exclude(microorganism="")
        .order_by("microorganism", "pesticide", "publication_year", "id")
    )
    rows = []
    for record in records:
        microorganism = clean(record.microorganism)
        taxonomy = lookup_taxonomy(microorganism, cache)
        rows.append(
            {
                "Microorganism": microorganism,
                "DOI": doi_from_record(record),
                "Reference": clean(record.reference),
                "Pesticide": clean(record.pesticide),
                "Publication_year": record.publication_year or "",
                "Culture_type": clean(record.culture_type) or "Individual strain",
                **taxonomy,
            }
        )
        if len(rows) % 25 == 0:
            save_cache(cache)
            print(f"processed_records={len(rows)} cache_entries={len(cache)}", flush=True)
    save_cache(cache)
    return rows


def is_generic_name(name: str) -> bool:
    lowered = canonical(name)
    return bool(
        re.fullmatch(r"[a-z]+", lowered)
        or re.fullmatch(r"[a-z]+ (sp\.|spp\.|species)", lowered)
        or "dominant in" in lowered
        or "community" in lowered
        or "consortium" in lowered
    )


def build_unique_rows(all_rows: list[dict]) -> list[dict]:
    grouped = defaultdict(list)
    for row in all_rows:
        microorganism = row["Microorganism"]
        doi = canonical(row["DOI"]) or canonical(row["Reference"])
        if is_generic_name(microorganism):
            key = (canonical(microorganism), doi)
        else:
            key = (canonical(microorganism),)
        grouped[key].append(row)

    unique_rows = []
    for rows in grouped.values():
        rows = sorted(
            rows,
            key=lambda row: (
                int(row["Publication_year"]) if str(row["Publication_year"]).isdigit() else -1,
                row["DOI"],
                row["Pesticide"],
            ),
            reverse=True,
        )
        chosen = rows[0].copy()
        chosen["All_DOIs_for_entry"] = "; ".join(
            sorted({row["DOI"] for row in rows if row["DOI"]})
        )
        chosen["All_pesticides_for_entry"] = "; ".join(
            sorted({row["Pesticide"] for row in rows if row["Pesticide"]})
        )
        chosen["Record_count_collapsed"] = len(rows)
        unique_rows.append(chosen)
    unique_rows.sort(key=lambda row: (canonical(row["Microorganism"]), canonical(row["DOI"])))
    return unique_rows


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in ws.columns:
            letter = get_column_letter(column_cells[0].column)
            max_len = max(len(str(cell.value or "")) for cell in column_cells[:200])
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 45)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(path)


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    cache = load_cache()
    all_rows = build_rows(cache)
    unique_rows = build_unique_rows(all_rows)

    columns = [
        "Microorganism",
        "DOI",
        "Reference",
        "Pesticide",
        "Publication_year",
        "Culture_type",
        "TaxID",
        "Taxonomy",
        "Domain",
        "Phylum",
        "Class",
        "Order",
        "Family",
        "Genus",
        "Species",
        "Matched_name",
        "Matched_rank",
        "Lookup_name",
        "Taxonomy_status",
    ]
    unique_columns = columns + [
        "All_DOIs_for_entry",
        "All_pesticides_for_entry",
        "Record_count_collapsed",
    ]
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        pd.DataFrame(all_rows, columns=columns).to_excel(
            writer, index=False, sheet_name="Sheet1_All_records"
        )
        pd.DataFrame(unique_rows, columns=unique_columns).to_excel(
            writer, index=False, sheet_name="Sheet2_Deduplicated"
        )
    format_workbook(OUTPUT_PATH)
    print(f"output={OUTPUT_PATH}")
    print(f"all_records={len(all_rows)}")
    print(f"deduplicated_rows={len(unique_rows)}")
    print(f"taxonomy_cache_entries={len(cache)}")


if __name__ == "__main__":
    main()
